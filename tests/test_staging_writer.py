"""
Regression tests for staging_writer.py's Title Case header-row display
(title_case_label) - the header row written to master.xlsx/staging/
export.xlsx should read as human-friendly labels ("Internal Ref") while the
underlying column identity used for every round-trip read stays the real
snake_case field name ("internal_ref"), since master_merge.py and friends
match/edit fields by that exact name.

Run with:

    .venv\\Scripts\\python.exe -m unittest tests.test_staging_writer -v
"""

import sys
import unittest
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from schema import ListingRow
from staging_writer import read_xlsx_with_hyperlinks, title_case_label, write_rows_to_xlsx


class TitleCaseLabelTests(unittest.TestCase):
    def test_documented_examples(self):
        self.assertEqual(title_case_label("internal_ref"), "Internal Ref")
        self.assertEqual(title_case_label("brochure_link"), "Brochure Link")
        self.assertEqual(title_case_label("desks_min"), "Desks Min")
        self.assertEqual(title_case_label("size_sqft"), "Size Sqft")

    def test_single_word_field(self):
        self.assertEqual(title_case_label("building"), "Building")


class WriteRowsToXlsxHeaderTests(unittest.TestCase):
    def _write_and_load_workbook(self, rows):
        buffer = BytesIO()
        write_rows_to_xlsx(rows, buffer)
        buffer.seek(0)
        return load_workbook(buffer)

    def test_header_row_is_title_case(self):
        wb = self._write_and_load_workbook([ListingRow(building="A", provider="P1")])
        ws = wb.active
        header_values = [cell.value for cell in ws[1]]

        self.assertIn("Internal Ref", header_values)
        self.assertIn("Brochure Link", header_values)
        self.assertIn("Desks Min", header_values)
        self.assertIn("Size Sqft", header_values)
        # None of the raw snake_case names should appear as header text.
        self.assertNotIn("internal_ref", header_values)
        self.assertNotIn("brochure_link", header_values)

    def test_header_order_still_matches_schema_field_order(self):
        # read_xlsx_with_hyperlinks relies on this positional match, not on
        # parsing the (now cosmetic) header text - see its docstring.
        wb = self._write_and_load_workbook([ListingRow(building="A", provider="P1")])
        ws = wb.active
        header_values = [cell.value for cell in ws[1]]
        expected = [title_case_label(f) for f in ListingRow.model_fields.keys()]

        self.assertEqual(header_values, expected)


class RoundTripTests(unittest.TestCase):
    """The critical invariant: Title Case is display-only - reading a
    written file back must still yield real snake_case column names and
    unchanged values, or every downstream field lookup breaks."""

    def _round_trip(self, rows):
        buffer = BytesIO()
        write_rows_to_xlsx(rows, buffer)
        return read_xlsx_with_hyperlinks(buffer.getvalue())

    def test_column_names_are_snake_case_not_title_case(self):
        df = self._round_trip([ListingRow(building="A", provider="P1")])

        self.assertIn("internal_ref", df.columns)
        self.assertIn("brochure_link", df.columns)
        self.assertNotIn("Internal Ref", df.columns)
        self.assertNotIn("Brochure Link", df.columns)

    def test_column_names_match_listingrow_fields_exactly(self):
        df = self._round_trip([ListingRow(building="A", provider="P1")])
        self.assertEqual(list(df.columns), list(ListingRow.model_fields.keys()))

    def test_values_round_trip_correctly(self):
        row = ListingRow(building="40 New Bond Street", provider="Workplace Plus", size_sqft=5000.0)
        df = self._round_trip([row])

        self.assertEqual(df.iloc[0]["building"], "40 New Bond Street")
        self.assertEqual(df.iloc[0]["provider"], "Workplace Plus")
        self.assertEqual(df.iloc[0]["size_sqft"], 5000.0)

    def test_hyperlink_column_still_round_trips_the_real_url_not_display_text(self):
        row = ListingRow(building="A", provider="P1", brochure_link="https://example.com/brochure.pdf")
        df = self._round_trip([row])

        self.assertEqual(df.iloc[0]["brochure_link"], "https://example.com/brochure.pdf")

    def test_hidden_columns_still_present_in_the_round_tripped_data(self):
        # Hidden in Excel (column width), never dropped from the data.
        row = ListingRow(building="A", provider="P1", property_id="prop-123", source_file="a.pdf")
        df = self._round_trip([row])

        self.assertEqual(df.iloc[0]["property_id"], "prop-123")
        self.assertEqual(df.iloc[0]["source_file"], "a.pdf")

    def test_brochure_and_floorplan_links_round_trip_distinctly_never_swapped(self):
        row = ListingRow(
            building="A", provider="P1",
            brochure_link="https://app.box.com/s/brochure123",
            floorplan_link="https://app.box.com/s/floorplan456",
        )
        df = self._round_trip([row])

        self.assertEqual(df.iloc[0]["brochure_link"], "https://app.box.com/s/brochure123")
        self.assertEqual(df.iloc[0]["floorplan_link"], "https://app.box.com/s/floorplan456")

    def test_a_genuinely_missing_floorplan_link_stays_blank_through_the_round_trip(self):
        row = ListingRow(building="A", provider="P1", brochure_link="https://example.com/brochure.pdf")
        df = self._round_trip([row])

        self.assertEqual(df.iloc[0]["brochure_link"], "https://example.com/brochure.pdf")
        self.assertTrue(pd.isna(df.iloc[0]["floorplan_link"]))


class HyperlinkDisplayTextTests(unittest.TestCase):
    """The written .xlsx cell must show a short, correctly-worded label per
    link TYPE, never the raw URL and never the wrong field's label."""

    def _cell_values(self, rows, field):
        buffer = BytesIO()
        write_rows_to_xlsx(rows, buffer)
        buffer.seek(0)
        wb = load_workbook(buffer)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        col_idx = headers.index(title_case_label(field)) + 1
        return [ws.cell(row=r, column=col_idx).value for r in range(2, ws.max_row + 1)]

    def test_brochure_link_shows_open_brochure(self):
        row = ListingRow(building="A", brochure_link="https://example.com/brochure.pdf")
        self.assertEqual(self._cell_values([row], "brochure_link"), ["Open brochure"])

    def test_floorplan_link_shows_open_floor_plan(self):
        row = ListingRow(building="A", floorplan_link="https://example.com/floorplan.pdf")
        self.assertEqual(self._cell_values([row], "floorplan_link"), ["Open floor plan"])

    def test_blank_brochure_link_shows_nothing_no_button(self):
        row = ListingRow(building="A", brochure_link=None)
        self.assertEqual(self._cell_values([row], "brochure_link"), [None])

    def test_blank_floorplan_link_shows_nothing_no_button(self):
        row = ListingRow(building="A", floorplan_link=None)
        self.assertEqual(self._cell_values([row], "floorplan_link"), [None])

    def test_a_row_with_both_shows_the_correct_distinct_label_for_each(self):
        row = ListingRow(
            building="A", brochure_link="https://example.com/b.pdf", floorplan_link="https://example.com/f.pdf",
        )
        self.assertEqual(self._cell_values([row], "brochure_link"), ["Open brochure"])
        self.assertEqual(self._cell_values([row], "floorplan_link"), ["Open floor plan"])


class BrokenBrochureLinkDisplayTests(unittest.TestCase):
    """
    A row whose brochure_link_broken is True must show a short, plain,
    grey/italic "Broken link" label instead of the normal clickable
    "Open brochure" - see BROKEN_LINK_DISPLAY_TEXT/BROKEN_LINK_FONT's own
    comment for why the cell keeps its real hyperlink underneath anyway
    (Option A of the two considered - the real URL must never be lost on
    a later read-back, see RoundTripTests below for the actual proof).
    """

    def _cell(self, row):
        buffer = BytesIO()
        write_rows_to_xlsx([row], buffer)
        buffer.seek(0)
        wb = load_workbook(buffer)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        col_idx = headers.index(title_case_label("brochure_link")) + 1
        return ws.cell(row=2, column=col_idx)

    def test_broken_link_shows_the_broken_label_not_open_brochure(self):
        row = ListingRow(building="A", brochure_link="https://example.com/dead.pdf", brochure_link_broken=True)
        cell = self._cell(row)
        self.assertEqual(cell.value, "Broken link")

    def test_broken_link_cell_still_carries_the_real_hyperlink_target(self):
        row = ListingRow(building="A", brochure_link="https://example.com/dead.pdf", brochure_link_broken=True)
        cell = self._cell(row)
        self.assertEqual(cell.hyperlink.target, "https://example.com/dead.pdf")

    def test_broken_link_uses_the_broken_font_not_the_normal_hyperlink_font(self):
        row = ListingRow(building="A", brochure_link="https://example.com/dead.pdf", brochure_link_broken=True)
        cell = self._cell(row)
        self.assertEqual(cell.font.color.rgb, "00808080")
        self.assertTrue(cell.font.italic)
        self.assertFalse(cell.font.underline)

    def test_a_working_link_is_completely_unaffected_false(self):
        row = ListingRow(building="A", brochure_link="https://example.com/fine.pdf", brochure_link_broken=False)
        cell = self._cell(row)
        self.assertEqual(cell.value, "Open brochure")

    def test_a_never_rechecked_link_is_completely_unaffected_none(self):
        row = ListingRow(building="A", brochure_link="https://example.com/unknown.pdf", brochure_link_broken=None)
        cell = self._cell(row)
        self.assertEqual(cell.value, "Open brochure")

    def test_floorplan_link_is_never_affected_by_brochure_link_broken(self):
        # brochure_link_broken is scoped to brochure_link only - a row's
        # floorplan_link cell must never show "Broken link" because of it.
        row = ListingRow(
            building="A", brochure_link="https://example.com/dead.pdf", brochure_link_broken=True,
            floorplan_link="https://example.com/plan.pdf",
        )
        buffer = BytesIO()
        write_rows_to_xlsx([row], buffer)
        buffer.seek(0)
        wb = load_workbook(buffer)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        col_idx = headers.index(title_case_label("floorplan_link")) + 1
        self.assertEqual(ws.cell(row=2, column=col_idx).value, "Open floor plan")

    def test_broken_link_url_still_round_trips_correctly(self):
        # The actual point of keeping the hyperlink (Option A) - a later
        # load of this same file (every merge run does this) must recover
        # the real URL, never the literal "Broken link" display text.
        row = ListingRow(building="A", brochure_link="https://example.com/dead.pdf", brochure_link_broken=True)
        buffer = BytesIO()
        write_rows_to_xlsx([row], buffer)
        df = read_xlsx_with_hyperlinks(buffer.getvalue())
        self.assertEqual(df.iloc[0]["brochure_link"], "https://example.com/dead.pdf")

    def test_brochure_link_broken_column_is_hidden_not_shown_as_a_raw_flag(self):
        # brochure_link_broken is pipeline diagnostics, never something a
        # reviewer should see as a raw True/False/blank column next to
        # brochure_link - see HIDDEN_COLUMNS' own comment on why it's
        # hidden rather than dropped (same treatment as source_file/
        # property_id), still there to unhide/read directly if ever
        # needed, but not surfaced by default.
        row = ListingRow(building="A", brochure_link="https://example.com/dead.pdf", brochure_link_broken=True)
        buffer = BytesIO()
        write_rows_to_xlsx([row], buffer)
        buffer.seek(0)
        wb = load_workbook(buffer)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        col_idx = headers.index(title_case_label("brochure_link_broken")) + 1
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        self.assertTrue(ws.column_dimensions[column_letter].hidden)


class LegacyColumnCompatibilityTests(unittest.TestCase):
    """Regression coverage for a real bug found while verifying this exact
    scenario against the real data/master.xlsx: that file predates the
    property_id field (inserted mid-schema, not at the end) AND still has
    the six now-removed range columns (which sat between rent_psf and
    brochure_link, not at the end either) - reading columns by fixed
    position against the CURRENT schema silently misaligned every column
    from "lat" onward (property_id's slot) into the WRONG field, and in
    this test's case raises a pydantic ValidationError rather than reading
    correctly. read_xlsx_with_hyperlinks must read each column by its own
    header text instead, immune to reordering/missing/extra columns."""

    # The real, historical column order/header text of data/master.xlsx -
    # no property_id, desks_max before desks_min, and the six now-removed
    # range columns sitting in the middle rather than the end. Snake_case
    # header text (this predates the Title Case display feature too) -
    # _label_to_field_name must be a no-op round-trip for text that's
    # already snake_case, not just for "Title Case" text.
    _LEGACY_HEADERS = [
        "internal_ref", "provider", "address_1", "postcode", "source_file", "lat", "lng", "submarket",
        "building", "floor_unit", "size_sqft", "desks_max", "desks_min", "size_sqft_min", "size_sqft_max",
        "rent_pcm", "rent_psf", "rent_psf_min", "rent_psf_max", "rent_pcm_min", "rent_pcm_max",
        "brochure_link", "special_features", "state_of_space", "contacts",
    ]

    def _legacy_workbook_bytes(self) -> bytes:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(self._LEGACY_HEADERS)
        row = {
            "internal_ref": "Breezblok", "provider": "Breezblok", "building": "John Stow House",
            "floor_unit": "Office 302", "size_sqft": 1750, "desks_max": 32, "rent_pcm": 18000,
            "rent_psf": 123.43, "lat": 51.5147, "lng": -0.0785, "submarket": "City of London",
        }
        ws.append([row.get(h) for h in self._LEGACY_HEADERS])
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def test_reading_a_legacy_reordered_file_does_not_crash(self):
        df = read_xlsx_with_hyperlinks(self._legacy_workbook_bytes())
        self.assertEqual(len(df), 1)

    def test_legacy_columns_land_in_their_correct_field_not_shifted(self):
        # The actual bug: a fixed-position read put "lat"'s value into
        # "property_id", "submarket"'s value into "lng", etc.
        df = read_xlsx_with_hyperlinks(self._legacy_workbook_bytes())
        row = df.iloc[0]

        self.assertEqual(row["building"], "John Stow House")
        self.assertEqual(row["floor_unit"], "Office 302")
        self.assertEqual(row["size_sqft"], 1750)
        self.assertEqual(row["lat"], 51.5147)
        self.assertEqual(row["submarket"], "City of London")
        self.assertEqual(row["rent_psf"], 123.43)

    def test_legacy_removed_range_columns_are_dropped_by_listingrow_construction(self):
        # read_xlsx_with_hyperlinks itself is a generic xlsx->DataFrame
        # reader - it correctly reads "size_sqft_min" back as a real column
        # (the file genuinely has it) rather than guessing which fields are
        # "current schema". Dropping fields ListingRow no longer has is
        # ListingRow's job (pydantic's default extra="ignore"), exercised
        # via dataframe_to_listing_rows here.
        from storage.file_store import dataframe_to_listing_rows

        df = read_xlsx_with_hyperlinks(self._legacy_workbook_bytes())
        self.assertIn("size_sqft_min", df.columns)  # genuinely present in the file - read correctly

        rows = dataframe_to_listing_rows(df)
        for removed_field in ("size_sqft_min", "size_sqft_max", "rent_psf_min", "rent_psf_max", "rent_pcm_min", "rent_pcm_max"):
            self.assertFalse(hasattr(rows[0], removed_field))

    def test_legacy_row_missing_property_id_entirely_converts_cleanly(self):
        # No property_id column at all in this historical shape - must not
        # KeyError, and the resulting ListingRow just gets property_id=None.
        from storage.file_store import dataframe_to_listing_rows

        df = read_xlsx_with_hyperlinks(self._legacy_workbook_bytes())
        rows = dataframe_to_listing_rows(df)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].building, "John Stow House")
        self.assertIsNone(rows[0].property_id)

    def test_dataframe_to_listing_rows_ignores_unknown_columns(self):
        # A separate guarantee from read_xlsx_with_hyperlinks' own fix above:
        # even if some other caller hands dataframe_to_listing_rows a
        # DataFrame that genuinely still has an old/unknown column (pydantic's
        # default extra="ignore" on ListingRow(**cleaned)), it must not raise.
        import pandas as pd

        from storage.file_store import dataframe_to_listing_rows

        df = pd.DataFrame([{
            "building": "A", "provider": "P1", "rent_psf_min": 190.0, "size_sqft_min": 2123.0,
        }])
        rows = dataframe_to_listing_rows(df)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].building, "A")
        self.assertFalse(hasattr(rows[0], "rent_psf_min"))


if __name__ == "__main__":
    unittest.main()
