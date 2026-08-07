"""
Regression test for app.py's handling of a hidden, non-authoritative rollup
sheet (see extract_spreadsheet_gemini.is_non_authoritative_rollup_sheet) -
the real, confirmed Copthall Estates Availability.xlsx shape: a hidden
"Portfolio" sheet (a flat Area/Station/Building/Office/... table with zero
real hyperlinks anywhere, roughly a year stale) sitting alongside genuinely
current, VISIBLE per-area sheets (City, Mid Town, Westend Soho, Blackfriars).

Runs the real app.py end-to-end via Streamlit's AppTest, with extract_
spreadsheet_gemini's Gemini call mocked - never calls the real API (same
principle as test_app_upload_missing_brochure_link.py).

Run with:
    .venv\\Scripts\\python.exe -m unittest tests.test_app_upload_hidden_portfolio_sheet -v
"""

import sys
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook
from streamlit.testing.v1 import AppTest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from storage.file_store import get_staging_fully_occupied_buildings, list_pending_staging_files, load_staging_as_dataframe

BASE = Path(__file__).resolve().parent.parent


def _clear_pending():
    for p in list_pending_staging_files():
        (BASE / p).unlink(missing_ok=True)
        (BASE / p).with_suffix(".meta.json").unlink(missing_ok=True)


def _city_block(ws):
    # A genuine shape (b) repeating block with real Floorplans/Brochure
    # hyperlinks - what every genuinely current Copthall area sheet has.
    ws.append([None, "50 Gresham Street - Bank"])
    ws.append([None, "A well-located City office building."])
    ws.append([
        None, "50 Gresham Street, EC2V 7AY", None, None, None,
        "Download Floorplans", None, None, "Download Brochure",
    ])
    ws["F3"].hyperlink = "https://example.com/floorplans"
    ws["I3"].hyperlink = "https://example.com/brochure.pdf"
    ws.append([None, "Office", "Sq.Ft", "Rent PCM", "Available From"])
    ws.append([None, "2nd Floor", 973.0, 14600.0, "Now"])


def _portfolio_rollup_rows(ws):
    # The real confirmed Copthall Portfolio shape: a title row (so its real
    # header row is NOT row 1 - header-mapping never even looks past row 1,
    # so this - exactly like the real file - falls to the Gemini fallback
    # path, where the hidden/no-download-line check actually applies),
    # then a flat table with no hyperlinks anywhere, overlapping (but
    # differently-named/differently-floored) building data.
    ws.append([None, "Copthall Estates Availibility - Updated 05/08/25"])
    ws.append([None, "Area", "Station", "Building", "Office", "Sq.Ft"])
    ws.append([None, "City", "Bank", "50 Gresham", "3rd Floor", 973])


def _build_workbook_with_hidden_portfolio() -> bytes:
    wb = Workbook()
    ws_city = wb.active
    ws_city.title = "City"
    _city_block(ws_city)

    ws_portfolio = wb.create_sheet("Portfolio")
    _portfolio_rollup_rows(ws_portfolio)
    ws_portfolio.sheet_state = "hidden"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_workbook_with_visible_portfolio_named_sheet() -> bytes:
    # A different provider's own genuinely current, VISIBLE sheet that
    # happens to be named "Portfolio" and DOES carry real hyperlinked
    # blocks - must never be skipped just because of its name or because
    # some OTHER provider's "Portfolio" sheet is confirmed non-authoritative.
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio"
    _city_block(ws)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _fully_occupied_block(ws):
    ws.append([None, "21 Whitefriars - Blackfriars"])
    ws.append([None, "A boutique riverside office building."])
    ws.append([None, "21 Whitefriars, EC4Y 8JJ", None, None, None, "Download Floorplans"])
    ws.append([None, "Office", "Sq.Ft", "Rent PCM", "Available From"])
    ws.append([None, "Fully Occupied"])


def _build_workbook_with_hidden_portfolio_and_fully_occupied_sheet() -> bytes:
    wb = Workbook()
    ws_city = wb.active
    ws_city.title = "City"
    _city_block(ws_city)

    ws_portfolio = wb.create_sheet("Portfolio")
    _portfolio_rollup_rows(ws_portfolio)
    ws_portfolio.sheet_state = "hidden"

    ws_blackfriars = wb.create_sheet("Blackfriars")
    _fully_occupied_block(ws_blackfriars)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


CITY_RAW_RESPONSE = {
    "provider": "Copthall Estates",
    "contacts": None,
    "fully_occupied_buildings": [],
    "units": [
        {
            "building": "50 Gresham Street", "floor_unit": "2nd Floor",
            "size_sqft": 973, "rent_pcm": 14600,
            "brochure_link": "https://example.com/brochure.pdf",
        },
    ],
}


class HiddenPortfolioSheetSkippedTests(unittest.TestCase):
    def setUp(self):
        _clear_pending()

    def tearDown(self):
        _clear_pending()

    def test_hidden_portfolio_sheet_is_skipped_city_sheet_still_extracted(self):
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=CITY_RAW_RESPONSE) as mock_call_gemini:
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()

            at.file_uploader[0].upload(
                "Copthall.xlsx", _build_workbook_with_hidden_portfolio(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            at.run()

            extract_buttons = [b for b in at.button if b.label == "Extract"]
            extract_buttons[0].click().run()
            self.assertFalse(at.exception)

        # Never called for the hidden Portfolio sheet - it's skipped BEFORE
        # any Gemini call, not filtered out after the fact.
        self.assertEqual(mock_call_gemini.call_count, 1)

        info_text = "".join(i.value for i in at.info)
        self.assertIn("Portfolio", info_text)
        self.assertIn("hidden", info_text.lower())

        pending = list_pending_staging_files()
        self.assertEqual(len(pending), 1)
        df = load_staging_as_dataframe(pending[0])
        self.assertEqual(len(df), 1)

    def test_no_duplicate_when_portfolio_and_area_sheet_both_describe_the_same_building(self):
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=CITY_RAW_RESPONSE):
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()

            at.file_uploader[0].upload(
                "Copthall.xlsx", _build_workbook_with_hidden_portfolio(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            at.run()

            extract_buttons = [b for b in at.button if b.label == "Extract"]
            extract_buttons[0].click().run()
            self.assertFalse(at.exception)

        pending = list_pending_staging_files()
        df = load_staging_as_dataframe(pending[0])
        # Exactly one "50 Gresham..." row - the City sheet's own 2nd Floor,
        # never a second copy from Portfolio's differently-named/differently
        # -floored "50 Gresham" - "3rd Floor" entry.
        gresham_rows = df[df["building"].str.contains("Gresham", na=False)]
        self.assertEqual(len(gresham_rows), 1)
        self.assertEqual(gresham_rows.iloc[0]["floor_unit"], "2nd Floor")

    def test_visible_portfolio_named_sheet_from_another_provider_is_not_skipped(self):
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=CITY_RAW_RESPONSE) as mock_call_gemini:
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()

            at.file_uploader[0].upload(
                "OtherProvider.xlsx", _build_workbook_with_visible_portfolio_named_sheet(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            at.run()

            extract_buttons = [b for b in at.button if b.label == "Extract"]
            extract_buttons[0].click().run()
            self.assertFalse(at.exception)

        mock_call_gemini.assert_called_once()
        pending = list_pending_staging_files()
        self.assertEqual(len(pending), 1)
        df = load_staging_as_dataframe(pending[0])
        self.assertEqual(len(df), 1)

        info_text = "".join(i.value for i in at.info)
        self.assertNotIn("hidden", info_text.lower())

    def test_fully_occupied_signal_from_authoritative_sheet_still_flows_alongside_a_skipped_portfolio_sheet(self):
        # Proves the two features compose correctly: the hidden Portfolio
        # sheet contributes nothing at all (not even to fully_occupied_
        # buildings), while a DIFFERENT, genuinely authoritative sheet's own
        # Fully Occupied signal (see master_merge.find_stale_candidates)
        # still reaches staging untouched.
        def _gemini_side_effect(client, prompt, parts):
            text = parts[0]
            if "Whitefriars" in text:
                return {
                    "provider": "Copthall Estates", "contacts": None,
                    "fully_occupied_buildings": ["21 Whitefriars"], "units": [],
                }
            return CITY_RAW_RESPONSE

        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", side_effect=_gemini_side_effect) as mock_call_gemini:
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()

            at.file_uploader[0].upload(
                "Copthall.xlsx", _build_workbook_with_hidden_portfolio_and_fully_occupied_sheet(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            at.run()

            extract_buttons = [b for b in at.button if b.label == "Extract"]
            extract_buttons[0].click().run()
            self.assertFalse(at.exception)

        # Called for City and Blackfriars, never for the hidden Portfolio sheet.
        self.assertEqual(mock_call_gemini.call_count, 2)

        pending = list_pending_staging_files()
        self.assertEqual(len(pending), 1)
        fully_occupied = get_staging_fully_occupied_buildings(pending[0])
        self.assertEqual(fully_occupied, [{"provider": "Copthall Estates", "building": "21 Whitefriars"}])

    def test_normal_tabular_sheet_never_reaches_the_rollup_check_at_all(self):
        # A plain single-consistent-table sheet (Kitt's/UNION shape - a real
        # header row at row 1, "Building" maps automatically) is fully
        # resolved by header-mapping alone and never falls to the Gemini
        # fallback branch is_non_authoritative_rollup_sheet lives in - so
        # Gemini (and this check) is never even reached, regardless of
        # sheet name or hidden state.
        wb = Workbook()
        ws = wb.active
        ws.title = "Portfolio"
        ws.append(["Building", "Floor/Unit", "Size (sq ft)", "Rent PCM"])
        ws.append(["40 New Bond Street", "3rd Floor", 2000, 15000])
        ws.sheet_state = "hidden"
        # openpyxl (like Excel itself) refuses to save a workbook whose only
        # sheet is hidden - an unrelated second, visible, empty sheet keeps
        # this fixture valid without changing what's actually under test.
        wb.create_sheet("Other")
        buffer = BytesIO()
        wb.save(buffer)

        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini") as mock_call_gemini:
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()

            at.file_uploader[0].upload(
                "tabular.xlsx", buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            at.run()

            extract_buttons = [b for b in at.button if b.label == "Extract"]
            extract_buttons[0].click().run()
            self.assertFalse(at.exception)

        mock_call_gemini.assert_not_called()
        pending = list_pending_staging_files()
        df = load_staging_as_dataframe(pending[0])
        self.assertEqual(len(df), 1)
        self.assertEqual(df.iloc[0]["building"], "40 New Bond Street")


if __name__ == "__main__":
    unittest.main()
