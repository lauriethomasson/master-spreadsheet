"""
Regression tests for app.py's ambiguous-sheet fallback - a THIRD outcome
alongside the existing confident "auto_skip" (see extract_spreadsheet_
gemini.is_non_authoritative_rollup_sheet, unchanged) and ordinary
"authoritative" processing: a sheet with some meaningful but not-confident
evidence of being non-authoritative (see classify_sheet_for_extraction) is
surfaced to a human for an explicit Include/Skip decision BEFORE Extract is
even clickable, rather than being silently guessed either way.

Runs the real app.py end-to-end via Streamlit's AppTest, with extract_
spreadsheet_gemini's Gemini call mocked - never calls the real API (same
principle as test_app_upload_hidden_portfolio_sheet.py).

Run with:
    .venv\\Scripts\\python.exe -m unittest tests.test_app_upload_ambiguous_sheet -v
"""

import hashlib
import sys
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook
from streamlit.testing.v1 import AppTest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import app
from storage.file_store import list_pending_staging_files, load_staging_as_dataframe

BASE = Path(__file__).resolve().parent.parent


def _clear_pending():
    for p in list_pending_staging_files():
        (BASE / p).unlink(missing_ok=True)
        (BASE / p).with_suffix(".meta.json").unlink(missing_ok=True)


def _city_block(ws, updated="03/08/2026"):
    ws.append([None, f"Copthall Estates City Availability - Updated {updated}"])
    ws.append([None, "50 Gresham Street - Bank"])
    ws.append([None, "A well-located City office building."])
    ws.append([
        None, "50 Gresham Street, EC2V 7AY", None, None, None,
        "Download Floorplans", None, None, "Download Brochure",
    ])
    ws["F4"].hyperlink = "https://example.com/floorplans"
    ws["I4"].hyperlink = "https://example.com/brochure.pdf"
    ws.append([None, "Office", "Sq.Ft", "Rent PCM", "Available From"])
    ws.append([None, "2nd Floor", 973.0, 14600.0, "Now"])


def _flat_rollup_table(ws, title_line="Copthall Estates Availibility - Updated 05/08/25"):
    ws.append([None, title_line])
    ws.append([None, "Area", "Station", "Building", "Office", "Sq.Ft", "Price Per Sq.Ft"])
    for i in range(10):
        ws.append([None, "City", "Bank", f"Building {i}", "3rd Floor", 1000 + i, 150])


def _hidden_but_structurally_legit_block(ws):
    # Real per-building hyperlinked block (shape (b)) - just hidden, which
    # alone is meaningful-but-not-confident evidence (see classify_sheet_
    # for_extraction's own "hidden" reason).
    ws.append([None, "11 Cursitor Street - Mid Town"])
    ws.append([None, "A period office building near Chancery Lane."])
    ws.append([None, "11 Cursitor Street, EC4A 1LT", None, None, None, "Download Floorplans"])
    ws.append([None, "Office", "Sq.Ft", "Rent PCM", "Available From"])
    ws.append([None, "1st Floor", 1200.0, 16000.0, "Now"])


def _build_workbook(sheets: dict) -> bytes:
    """sheets: {sheet_name: (builder_fn, sheet_state)}"""
    wb = Workbook()
    first = True
    for sheet_name, (builder, state) in sheets.items():
        ws = wb.active if first else wb.create_sheet(sheet_name)
        if first:
            ws.title = sheet_name
            first = False
        builder(ws)
        ws.sheet_state = state
    return _save(wb)


def _save(wb) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _kitts_style_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Availability"
    ws.append(["Building", "Floor/Unit", "Size (sq ft)", "Rent PCM"])
    ws.append(["40 New Bond Street", "3rd Floor", 2000, 15000])
    return _save(wb)


CITY_RAW_RESPONSE = {
    "provider": "Copthall Estates",
    "contacts": None,
    "fully_occupied_buildings": [],
    "units": [
        {
            "building": "50 Gresham Street", "floor_unit": "2nd Floor",
            "size_sqft": 973, "rent_pcm": 14600,
            "brochure_link": "https://example.com/brochure.pdf",
        },
    ],
}

CURSITOR_RAW_RESPONSE = {
    "provider": "Copthall Estates",
    "contacts": None,
    "fully_occupied_buildings": [],
    "units": [
        {
            "building": "11 Cursitor Street", "floor_unit": "1st Floor",
            "size_sqft": 1200, "rent_pcm": 16000,
        },
    ],
}


def _file_hash(file_bytes: bytes) -> str:
    return hashlib.sha256(file_bytes).hexdigest()


class AmbiguousSheetPromptTests(unittest.TestCase):
    def setUp(self):
        _clear_pending()

    def tearDown(self):
        _clear_pending()

    def _upload(self, at, filename, file_bytes):
        at.file_uploader[0].upload(
            filename, file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        at.run()

    def test_ambiguous_visible_portfolio_like_sheet_produces_a_prompt(self):
        file_bytes = _build_workbook({
            "City": (_city_block, "visible"),
            "Portfolio": (_flat_rollup_table, "visible"),
        })

        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=CITY_RAW_RESPONSE):
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()
            self._upload(at, "Copthall.xlsx", file_bytes)

            self.assertFalse(at.exception)
            markdown_text = "".join(m.value for m in at.markdown)
            self.assertIn("Portfolio", markdown_text)

            radios = [r for r in at.radio if 'What should we do with "Portfolio"?' == r.label]
            self.assertEqual(len(radios), 1)

    def test_ambiguous_hidden_but_potentially_legitimate_sheet_produces_a_prompt(self):
        file_bytes = _build_workbook({
            "City": (_city_block, "visible"),
            "Extra": (_hidden_but_structurally_legit_block, "hidden"),
        })

        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=CITY_RAW_RESPONSE):
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()
            self._upload(at, "Copthall.xlsx", file_bytes)

            self.assertFalse(at.exception)
            radios = [r for r in at.radio if 'What should we do with "Extra"?' == r.label]
            self.assertEqual(len(radios), 1)

    def test_ordinary_authoritative_sheet_produces_no_prompt(self):
        file_bytes = _build_workbook({"City": (_city_block, "visible")})

        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=CITY_RAW_RESPONSE):
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()
            self._upload(at, "Copthall.xlsx", file_bytes)

            self.assertFalse(at.exception)
            self.assertEqual(len(at.radio), 0)
            extract_buttons = [b for b in at.button if b.label == "Extract"]
            self.assertFalse(extract_buttons[0].disabled)

    def test_kitts_style_tabular_upload_produces_no_prompt_and_is_not_disabled(self):
        with patch("extract_spreadsheet_gemini.get_client") as mock_client, \
             patch("extract_spreadsheet_gemini.call_gemini") as mock_gemini:
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()
            self._upload(at, "Kitts.xlsx", _kitts_style_workbook())

            self.assertFalse(at.exception)
            self.assertEqual(len(at.radio), 0)
            extract_buttons = [b for b in at.button if b.label == "Extract"]
            self.assertFalse(extract_buttons[0].disabled)
            mock_client.assert_not_called()
            mock_gemini.assert_not_called()

    def test_hidden_sheet_explanatory_text_is_shown(self):
        file_bytes = _build_workbook({
            "City": (_city_block, "visible"),
            "Portfolio": (_flat_rollup_table, "visible"),
        })

        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=CITY_RAW_RESPONSE):
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()
            self._upload(at, "Copthall.xlsx", file_bytes)

            self.assertFalse(at.exception)
            all_text = "".join(c.value for c in at.caption) + "".join(m.value for m in at.markdown)
            self.assertIn("hidden sheets", all_text.lower())

    def test_auto_skipped_hidden_sheet_produces_no_upload_time_message(self):
        # A confidently non-authoritative sheet (real Copthall Portfolio
        # shape: hidden + zero download lines) is never a prompt, and - per
        # explicit request - no longer gets any upload-time informational
        # message either: it's silently skipped, exactly like any other
        # fully-automatic decision. The Extract-time "hidden, non-
        # authoritative rollup sheet — skipped" message (unchanged, shown
        # only once Extract actually runs) is covered separately by
        # test_app_upload_hidden_portfolio_sheet.py.
        file_bytes = _build_workbook({
            "City": (_city_block, "visible"),
            "Portfolio": (_flat_rollup_table, "hidden"),
        })

        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=CITY_RAW_RESPONSE):
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()
            self._upload(at, "Copthall.xlsx", file_bytes)

            self.assertFalse(at.exception)
            self.assertEqual(len(at.radio), 0)  # never a prompt for this one
            info_text = "".join(i.value for i in at.info)
            self.assertNotIn("Portfolio", info_text)

            extract_buttons = [b for b in at.button if b.label == "Extract"]
            self.assertFalse(extract_buttons[0].disabled)

    def test_another_providers_legitimate_visible_portfolio_sheet_is_not_auto_skipped(self):
        file_bytes = _build_workbook({"Portfolio": (_city_block, "visible")})

        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=CITY_RAW_RESPONSE) as mock_gemini:
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()
            self._upload(at, "OtherProvider.xlsx", file_bytes)
            self.assertFalse(at.exception)

            info_text = "".join(i.value for i in at.info)
            self.assertNotIn("Skipped", info_text)

            extract_buttons = [b for b in at.button if b.label == "Extract"]
            extract_buttons[0].click().run()
            self.assertFalse(at.exception)
            mock_gemini.assert_called_once()

    def test_multiple_ambiguous_sheets_require_separate_decisions(self):
        file_bytes = _build_workbook({
            "City": (_city_block, "visible"),
            "Portfolio": (_flat_rollup_table, "visible"),
            "Extra": (_hidden_but_structurally_legit_block, "hidden"),
        })

        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=CITY_RAW_RESPONSE):
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()
            self._upload(at, "Copthall.xlsx", file_bytes)
            self.assertFalse(at.exception)

            extract_buttons = [b for b in at.button if b.label == "Extract"]
            self.assertTrue(extract_buttons[0].disabled)

            file_hash = _file_hash(file_bytes)
            portfolio_key = app._sheet_decision_key(file_hash, "Portfolio")
            extra_key = app._sheet_decision_key(file_hash, "Extra")

            # Resolving only ONE of the two must still leave Extract disabled.
            at.radio(key=portfolio_key).set_value(app.SHEET_DECISION_SKIP)
            at.run()
            extract_buttons = [b for b in at.button if b.label == "Extract"]
            self.assertTrue(extract_buttons[0].disabled)

            at.radio(key=extra_key).set_value(app.SHEET_DECISION_INCLUDE)
            at.run()
            extract_buttons = [b for b in at.button if b.label == "Extract"]
            self.assertFalse(extract_buttons[0].disabled)


class AmbiguousSheetDecisionEffectTests(unittest.TestCase):
    def setUp(self):
        _clear_pending()

    def tearDown(self):
        _clear_pending()

    def test_user_chooses_skip_gemini_never_called_for_that_sheet(self):
        file_bytes = _build_workbook({
            "City": (_city_block, "visible"),
            "Portfolio": (_flat_rollup_table, "visible"),
        })
        file_hash = _file_hash(file_bytes)

        def _gemini_side_effect(client, prompt, parts):
            text = parts[0]
            self.assertNotIn("Building 0", text)  # Portfolio's own rollup text never reaches Gemini
            return CITY_RAW_RESPONSE

        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", side_effect=_gemini_side_effect) as mock_gemini:
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()
            at.file_uploader[0].upload(
                "Copthall.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            at.run()

            at.radio(key=app._sheet_decision_key(file_hash, "Portfolio")).set_value(app.SHEET_DECISION_SKIP)
            at.run()

            extract_buttons = [b for b in at.button if b.label == "Extract"]
            self.assertFalse(extract_buttons[0].disabled)
            extract_buttons[0].click().run()
            self.assertFalse(at.exception)

        mock_gemini.assert_called_once()  # only for City

        pending = list_pending_staging_files()
        df = load_staging_as_dataframe(pending[0])
        self.assertEqual(len(df), 1)
        self.assertEqual(df.iloc[0]["building"], "50 Gresham Street")

    def test_user_chooses_include_normal_extraction_runs(self):
        file_bytes = _build_workbook({
            "City": (_city_block, "visible"),
            "Portfolio": (_flat_rollup_table, "visible"),
        })
        file_hash = _file_hash(file_bytes)

        rollup_raw_response = {
            "provider": "Copthall Estates", "contacts": None, "fully_occupied_buildings": [],
            "units": [
                {"building": "Building 0", "floor_unit": "3rd Floor", "size_sqft": 1000, "rent_pcm": 15000},
            ],
        }

        def _gemini_side_effect(client, prompt, parts):
            text = parts[0]
            if "Building 0" in text:
                return rollup_raw_response
            return CITY_RAW_RESPONSE

        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", side_effect=_gemini_side_effect) as mock_gemini:
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()
            at.file_uploader[0].upload(
                "Copthall.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            at.run()

            at.radio(key=app._sheet_decision_key(file_hash, "Portfolio")).set_value(app.SHEET_DECISION_INCLUDE)
            at.run()

            extract_buttons = [b for b in at.button if b.label == "Extract"]
            self.assertFalse(extract_buttons[0].disabled)
            extract_buttons[0].click().run()
            self.assertFalse(at.exception)

        self.assertEqual(mock_gemini.call_count, 2)  # City AND Portfolio, since included
        pending = list_pending_staging_files()
        df = load_staging_as_dataframe(pending[0])
        self.assertEqual(len(df), 2)
        self.assertIn("Building 0", set(df["building"]))

    def test_extract_disabled_message_names_the_sheet_still_needing_a_decision(self):
        file_bytes = _build_workbook({
            "City": (_city_block, "visible"),
            "Portfolio": (_flat_rollup_table, "visible"),
        })

        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=CITY_RAW_RESPONSE):
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()
            at.file_uploader[0].upload(
                "Copthall.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            at.run()

            warning_text = "".join(w.value for w in at.warning)
            self.assertIn("Portfolio", warning_text)


class AmbiguousSheetDecisionScopingTests(unittest.TestCase):
    def setUp(self):
        _clear_pending()

    def tearDown(self):
        _clear_pending()

    def test_decision_does_not_contaminate_a_different_workbook(self):
        first_bytes = _build_workbook({
            "City": (_city_block, "visible"),
            "Portfolio": (_flat_rollup_table, "visible"),
        })
        second_bytes = _build_workbook({
            "City": (lambda ws: _city_block(ws, updated="10/08/2026"), "visible"),
            "Portfolio": (lambda ws: _flat_rollup_table(ws, title_line="A Different Portfolio - Updated 01/01/25"), "visible"),
        })
        self.assertNotEqual(first_bytes, second_bytes)

        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=CITY_RAW_RESPONSE):
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()

            at.file_uploader[0].upload(
                "Copthall.xlsx", first_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            at.run()
            at.radio(key=app._sheet_decision_key(_file_hash(first_bytes), "Portfolio")).set_value(
                app.SHEET_DECISION_SKIP
            )
            at.run()

            # Uploading a DIFFERENT file (different bytes) must start with
            # NO decision recorded for its own "Portfolio" sheet, even
            # though the sheet name is identical - content-hash-scoped, not
            # name-scoped.
            at.file_uploader[0].upload(
                "Copthall2.xlsx", second_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            at.run()

            extract_buttons = [b for b in at.button if b.label == "Extract"]
            self.assertTrue(extract_buttons[0].disabled)


class SpreadsheetContentHashTests(unittest.TestCase):
    """app._spreadsheet_content_hash - the cache-identity function folding
    ambiguous-sheet decisions into the existing content_hash formula (see
    find_previous_upload_by_hash) so a decision change can never silently
    reuse a result cached under a different one."""

    def test_same_bytes_different_decisions_hash_differently(self):
        file_bytes = b"pretend workbook bytes"
        skip_hash = app._spreadsheet_content_hash(file_bytes, {"Portfolio": app.SHEET_DECISION_SKIP})
        include_hash = app._spreadsheet_content_hash(file_bytes, {"Portfolio": app.SHEET_DECISION_INCLUDE})

        self.assertNotEqual(skip_hash, include_hash)

    def test_same_bytes_same_decisions_hash_identically(self):
        file_bytes = b"pretend workbook bytes"
        hash_a = app._spreadsheet_content_hash(file_bytes, {"Portfolio": app.SHEET_DECISION_SKIP})
        hash_b = app._spreadsheet_content_hash(file_bytes, {"Portfolio": app.SHEET_DECISION_SKIP})

        self.assertEqual(hash_a, hash_b)

    def test_no_ambiguous_sheets_matches_the_pre_existing_formula(self):
        file_bytes = b"pretend workbook bytes"
        result = app._spreadsheet_content_hash(file_bytes, {})
        expected = hashlib.sha256(
            app._SPREADSHEET_LOGIC_FINGERPRINT.encode("utf-8") + b"\0" + file_bytes + b"\0" + b"{}"
        ).hexdigest()

        self.assertEqual(result, expected)

    def test_different_bytes_same_decisions_hash_differently(self):
        hash_a = app._spreadsheet_content_hash(b"bytes A", {"Portfolio": app.SHEET_DECISION_SKIP})
        hash_b = app._spreadsheet_content_hash(b"bytes B", {"Portfolio": app.SHEET_DECISION_SKIP})

        self.assertNotEqual(hash_a, hash_b)


if __name__ == "__main__":
    unittest.main()
