"""
Regression tests for app._rows_dropped_as_duplicate_sheet_extractions - a
real, confirmed Kitt's "Kitts_Availability_External.xlsx" gap: two header-
mapped sheets ("Source_Availability", an internal working tracker pulling
live via IMPORTRANGE; "Live Availability", the polished published-for-
external-use view) independently listing the exact same real units, so
every sheet being processed unconditionally and independently (deliberate,
correct for a real Copthall Estates file's own 4 genuinely different-area
sheets) produced two rows per shared unit for Kitt's. Most of those pairs
matched silently at Review & Master's own intra-batch duplicate detection
(byte-identical text); "8 Laurence Pountney Hill" had a real difference
(one sheet's own "Space video" cell has a genuine embedded YouTube
hyperlink, the other doesn't), correctly but unnecessarily tripping a
"possible duplicate listings, can't safely tell" human-review card for
what was never two independently-entered listings at all.

Fix: detect actual (building, floor_unit) row overlap BETWEEN header-mapped
sheets in the same file (never a sheet-naming heuristic - see app.py's own
_rows_dropped_as_duplicate_sheet_extractions docstring for why that
wouldn't generalize), and when a large majority of one sheet's own rows are
also found on another, drop the redundant duplicate - keeping the richer
row via master_merge.richest_listing_index, the same tie-break Review &
Master's own "Same listing — merge" choice already trusts.

Run with:
    .venv\\Scripts\\python.exe -m unittest tests.test_app_upload_duplicate_sheets -v
"""

import sys
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook
from streamlit.testing.v1 import AppTest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import app
from schema import ListingRow
from storage.file_store import list_pending_staging_files, load_staging_as_dataframe

# The bare `import app` above executes app.py's top-level page code once
# with no active Streamlit runtime, which taints every subsequent AppTest
# run of app.py in this process unless cleared - see tests/test_app_upload_
# paste_a_link.py's own copy of this same fix for the full explanation.
from streamlit.delta_generator_singletons import get_dg_singleton_instance

get_dg_singleton_instance().main_dg._form_data = None

BASE = Path(__file__).resolve().parent.parent


def _clear_pending():
    for p in list_pending_staging_files():
        (BASE / p).unlink(missing_ok=True)
        (BASE / p).with_suffix(".meta.json").unlink(missing_ok=True)


class RowsDroppedAsDuplicateSheetExtractionsTests(unittest.TestCase):
    """Direct, non-AppTest unit tests of the pure overlap-detection logic."""

    def _row(self, building, floor_unit=None, **fields):
        return ListingRow(building=building, floor_unit=floor_unit, **fields)

    def test_a_single_sheet_is_never_touched(self):
        sheet = [self._row("8 Laurence Pountney Hill", "G"), self._row("8 Laurence Pountney Hill", "2nd")]
        self.assertEqual(app._rows_dropped_as_duplicate_sheet_extractions([sheet]), set())

    def test_high_overlap_sheets_merge_special_features_onto_the_surviving_row(self):
        # Mirrors the real 8 Laurence Pountney Hill case: identical size/
        # rent on both sheets (richest_listing_index alone ties), but one
        # sheet's row has real special_features text (the embedded video
        # link) the other's lacks - richest_listing_index alone would pick
        # arbitrarily and could silently lose that text; the merge step
        # must fold it onto whichever row survives instead.
        sparse = self._row("8 Laurence Pountney Hill", "G", size_sqft=1200)
        rich = self._row("8 Laurence Pountney Hill", "G", size_sqft=1200, special_features="Space video: https://youtu.be/abc")
        sheet_a = [sparse]
        sheet_b = [rich]

        dropped = app._rows_dropped_as_duplicate_sheet_extractions([sheet_a, sheet_b])

        # Exactly one of the two is dropped, and the survivor carries the
        # real special_features text either way - which physical row
        # object survives (a richness tie resolves to the earliest index)
        # is an implementation detail this test deliberately doesn't pin.
        self.assertEqual(len(dropped), 1)
        survivor = rich if id(rich) not in dropped else sparse
        self.assertIn("Space video", survivor.special_features)

    def test_rows_unique_to_one_sheet_are_never_dropped(self):
        # 3 of 4 rows overlap (75%, clears the threshold) - the sheets are
        # judged duplicative overall, but each sheet's own genuinely unique
        # row (no counterpart on the other sheet) must still survive.
        shared_a1 = self._row("Building A", "1st")
        shared_a2 = self._row("Building A", "2nd")
        shared_a3 = self._row("Building A", "3rd")
        unique_to_source = self._row("Building A", "4th")

        shared_b1 = self._row("Building A", "1st")
        shared_b2 = self._row("Building A", "2nd")
        shared_b3 = self._row("Building A", "3rd")
        unique_to_live = self._row("Building A", "5th")

        sheet_source = [shared_a1, shared_a2, shared_a3, unique_to_source]
        sheet_live = [shared_b1, shared_b2, shared_b3, unique_to_live]

        dropped = app._rows_dropped_as_duplicate_sheet_extractions([sheet_source, sheet_live])

        # Exactly one member of each shared pair dropped, never either
        # unique row.
        self.assertEqual(len(dropped), 3)
        for pair in ((shared_a1, shared_b1), (shared_a2, shared_b2), (shared_a3, shared_b3)):
            self.assertEqual(sum(id(r) in dropped for r in pair), 1)
        self.assertNotIn(id(unique_to_source), dropped)
        self.assertNotIn(id(unique_to_live), dropped)

    def test_low_overlap_sheets_are_left_completely_alone(self):
        # The real Copthall Estates shape: 4 genuinely different-area
        # sheets, essentially no cross-sheet overlap. One coincidentally
        # shared building among many genuinely different ones must never
        # clear the majority-overlap threshold.
        sheet_city = [self._row("50 Gresham Street", "2nd"), self._row("1 London Wall", "3rd"), self._row("2 London Wall", "4th")]
        sheet_midtown = [self._row("50 Gresham Street", "2nd"), self._row("Holborn Gate", "1st"), self._row("Kingsway House", "5th")]

        dropped = app._rows_dropped_as_duplicate_sheet_extractions([sheet_city, sheet_midtown])

        self.assertEqual(dropped, set())

    def test_blank_building_rows_are_never_treated_as_a_shared_identity(self):
        # building is a required str on ListingRow - "" (never None) is
        # the real shape a genuinely blank one takes.
        blank_a = self._row("", "1st")
        blank_b = self._row("", "1st")
        # Give both sheets a real, matching pair too, so overlap-ratio
        # alone can't explain an empty result - this specifically proves
        # the blank-building pair itself never contributes a shared key.
        real_a = self._row("Building A", "1st")
        real_b = self._row("Building A", "1st")

        dropped = app._rows_dropped_as_duplicate_sheet_extractions([[blank_a, real_a], [blank_b, real_b]])

        self.assertNotIn(id(blank_a), dropped)
        self.assertNotIn(id(blank_b), dropped)
        # The real pair still dedupes normally.
        self.assertEqual(len(dropped), 1)

    def test_matching_is_case_and_whitespace_tolerant(self):
        row_a = self._row(" 8 Laurence Pountney Hill ", "G")
        row_b = self._row("8 LAURENCE POUNTNEY HILL", "g")

        dropped = app._rows_dropped_as_duplicate_sheet_extractions([[row_a], [row_b]])

        self.assertEqual(len(dropped), 1)


def _kitt_sheet(ws, video_link=False):
    # "Key Features" is a real header-mapping synonym for special_features
    # (see extract_spreadsheet.FIELD_SYNONYMS) - this is the header-mapped
    # counterpart to the real Kitt's file's own "Space video" cell (there,
    # a genuine embedded hyperlink one sheet's extraction picked up and the
    # other didn't; here, a plain differing column value achieves the same
    # "one sheet has real text the other lacks" shape without needing to
    # replicate hyperlink-reading mechanics this fix doesn't touch).
    ws.append(["Building", "Floor/Unit", "Size (sq ft)", "Monthly Rate", "Key Features"])
    video_note = "Space video: https://youtu.be/abc123" if video_link else None
    ws.append(["8 Laurence Pountney Hill", "Ground", 1200, 5000, video_note])
    ws.append(["8 Laurence Pountney Hill", "2nd", 900, 4000, None])
    ws.append(["27-29 Some Other Street", "1st", 1500, 6000, None])


def _build_kitts_shaped_duplicate_sheets_workbook() -> bytes:
    wb = Workbook()
    ws_source = wb.active
    ws_source.title = "Source_Availability"
    _kitt_sheet(ws_source, video_link=False)

    ws_live = wb.create_sheet("Live Availability")
    _kitt_sheet(ws_live, video_link=True)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _copthall_shaped_distinct_sheets_workbook() -> bytes:
    wb = Workbook()
    ws_city = wb.active
    ws_city.title = "City"
    ws_city.append(["Building", "Floor/Unit", "Size (sq ft)", "Monthly Rate"])
    ws_city.append(["50 Gresham Street", "2nd Floor", 973, 14600])

    ws_midtown = wb.create_sheet("Mid Town")
    ws_midtown.append(["Building", "Floor/Unit", "Size (sq ft)", "Monthly Rate"])
    ws_midtown.append(["Holborn Gate", "1st Floor", 1100, 12000])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class DuplicateSheetExtractionEndToEndTests(unittest.TestCase):
    """Full pipeline via the real app.py Extract loop (AppTest), no Gemini
    call involved - both sheets here are header-mapped."""

    def setUp(self):
        _clear_pending()

    def tearDown(self):
        _clear_pending()

    def test_kitts_shaped_duplicate_sheets_stage_one_row_per_shared_unit(self):
        with patch("geocode.geocode_rows"):
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()

            at.file_uploader[0].upload(
                "Kitts_Availability_External.xlsx", _build_kitts_shaped_duplicate_sheets_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            at.run()

            extract_buttons = [b for b in at.button if b.label == "Extract"]
            extract_buttons[0].click().run()
            self.assertFalse(at.exception)

        pending = list_pending_staging_files()
        self.assertEqual(len(pending), 1)
        df = load_staging_as_dataframe(pending[0])

        # 3 real units total (2 at 8 Laurence Pountney Hill + 1 at 27-29
        # Some Other Street), never 6 - each sheet's own duplicate of the
        # SAME unit collapsed to exactly one row.
        self.assertEqual(len(df), 3)
        laurence_rows = df[df["building"].str.contains("Laurence Pountney", na=False)]
        self.assertEqual(len(laurence_rows), 2)
        # The richer row (the one with the real video-link special_features
        # text, from "Live Availability") is the one kept for Ground floor.
        ground_row = laurence_rows[laurence_rows["floor_unit"] == "Ground"].iloc[0]
        self.assertIn("Space video", str(ground_row.get("special_features") or ""))

    def test_copthall_shaped_distinct_sheets_keep_every_row(self):
        # Genuinely different-area sheets with no real overlap must never
        # be affected by this - both buildings survive independently.
        with patch("geocode.geocode_rows"):
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()

            at.file_uploader[0].upload(
                "Copthall.xlsx", _copthall_shaped_distinct_sheets_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            at.run()

            extract_buttons = [b for b in at.button if b.label == "Extract"]
            extract_buttons[0].click().run()
            self.assertFalse(at.exception)

        pending = list_pending_staging_files()
        self.assertEqual(len(pending), 1)
        df = load_staging_as_dataframe(pending[0])
        self.assertEqual(len(df), 2)


if __name__ == "__main__":
    unittest.main()
