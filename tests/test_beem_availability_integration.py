"""
Real-data regression tests for a production report: uploading the real
"beem Live Flex Availability.xlsx" left provider blank and prompted an
unnecessary "Possible summary or old sheet detected" decision, even though
the sheet is a completely ordinary, current, flat per-unit availability
table (43 rows, one clean header row: Property/Floor/Size/Min Desks/Max
Desks/Managed rent per sqft/Managed rent per month/Managed rent per year).

Traced to three independent, generic gaps:
1. extract_spreadsheet.suggest_mapping had no synonym for "Property" (a
   common, generic commercial-property header for the building/property
   identifier) - building is CRITICAL_FIELDS' own only entry, so this alone
   sent an otherwise cleanly header-mappable sheet into the Gemini fallback
   path entirely unnecessarily.
2. extract_spreadsheet_gemini.classify_sheet_for_extraction treated "a flat
   data table with no download/brochure line" as ambiguity evidence on its
   own - too broad: a provider is allowed to have a current availability
   spreadsheet with no brochures at all.
3. app.py's fill_missing_provider was only ever applied to header_mapped_
   rows, never gemini_rows - conflating SOURCE TYPE (spreadsheet vs PDF/
   email, which fill_missing_provider is correctly scoped by) with
   EXTRACTION METHOD (header-mapping vs Gemini fallback, which it must not
   be scoped by) - a spreadsheet that happens to need the Gemini fallback
   is still a spreadsheet source, with no column stating a provider either
   way, so the filename fallback is exactly as reliable for it.

This specific real file no longer needs the Gemini fallback at all once (1)
is fixed - "Property" maps directly to building - so these tests are fully
deterministic, no mocking or real Gemini/network calls needed.

The real workbook is NOT checked into this repo (this project's own
.gitignore excludes *.xlsx outright, the same "no real provider/customer
data in version control" convention test_email_upload_integration.py's own
external REAL_EMAIL fixture already follows) - every test here is skipped
automatically wherever BEEM_PATH doesn't exist, so this file never breaks
the full suite for anyone without a local copy of it.

Run with:
    .venv\\Scripts\\python.exe -m unittest tests.test_beem_availability_integration -v
"""

import sys
import unittest
from pathlib import Path

from openpyxl import load_workbook
from streamlit.testing.v1 import AppTest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from unittest.mock import patch

import extract_spreadsheet
import extract_spreadsheet_gemini
import geocode
from schema import ListingRow
from storage.file_store import list_pending_staging_files, load_staging_as_dataframe

BASE = Path(__file__).resolve().parent.parent
BEEM_PATH = Path(r"C:\Users\julie\OneDrive\Documents\beem Live Flex Availability.xlsx")
_HAS_FIXTURE = BEEM_PATH.exists()


def _beem_bytes() -> bytes:
    return BEEM_PATH.read_bytes()


def _clear_pending():
    # AppTest(app.py)'s own imports rely on running from the real repo's
    # own working directory (never an isolated cwd) - same convention as
    # test_app_upload_brochure_enrichment.py - so these tests share the
    # real repo's staging/ directory and must clean up after themselves.
    from storage import file_store as _file_store
    _file_store._list_pending_staging_files_cached.clear()
    _file_store._find_previous_upload_by_hash_cached.clear()
    _file_store._load_staging_as_dataframe_cached.clear()
    for p in list_pending_staging_files():
        (BASE / p).unlink(missing_ok=True)
        (BASE / p).with_suffix(".meta.json").unlink(missing_ok=True)


@unittest.skipUnless(_HAS_FIXTURE, "real Beem fixture not present in this environment")
class BeemHeaderMappingTests(unittest.TestCase):
    """suggest_mapping against the real Beem header row - the exact mapping
    for every header, computed once and reused across assertions."""

    @classmethod
    def setUpClass(cls):
        wb = load_workbook(BEEM_PATH, read_only=True)
        ws = wb["Sheet1"]
        cls.headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        cls.mapping = extract_spreadsheet.suggest_mapping(cls.headers)

    def test_property_maps_to_building(self):
        self.assertEqual(self.mapping["Property"], "building")

    def test_floor_maps_to_floor_unit(self):
        self.assertEqual(self.mapping["Floor"], "floor_unit")

    def test_size_maps_to_size_sqft(self):
        self.assertEqual(self.mapping["Size"], "size_sqft")

    def test_min_desks_maps_to_desks_min(self):
        self.assertEqual(self.mapping["Min Desks"], "desks_min")

    def test_max_desks_maps_to_desks_max(self):
        self.assertEqual(self.mapping["Max Desks"], "desks_max")

    def test_managed_rent_per_sqft_maps_to_rent_psf(self):
        header = next(h for h in self.headers if "per sqft" in h.lower())
        self.assertEqual(self.mapping[header], "rent_psf")

    def test_managed_rent_per_month_maps_to_rent_pcm(self):
        header = next(h for h in self.headers if "per month" in h.lower())
        self.assertEqual(self.mapping[header], "rent_pcm")

    def test_no_critical_fields_unmapped(self):
        self.assertEqual(extract_spreadsheet.unmapped_critical_fields(self.mapping), [])

    def test_unrelated_headers_never_map_to_building(self):
        # "Property" is a safe, generic synonym - it must never widen what
        # else maps to building.
        guess = extract_spreadsheet.suggest_mapping(["Property Id", "Property Address", "Managed By"])
        self.assertIsNone(guess["Property Id"])
        self.assertIsNone(guess["Property Address"])
        self.assertIsNone(guess["Managed By"])


@unittest.skipUnless(_HAS_FIXTURE, "real Beem fixture not present in this environment")
class BeemSheetClassificationTests(unittest.TestCase):
    """classify_sheet_for_extraction against the real Beem sheet - must be
    authoritative, never ambiguous, purely for lacking brochure links."""

    def test_sheet1_is_authoritative(self):
        wb = load_workbook(BEEM_PATH)
        ws = wb["Sheet1"]
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        result = extract_spreadsheet_gemini.classify_sheet_for_extraction(ws, text)

        self.assertEqual(result["outcome"], "authoritative")
        self.assertEqual(result["reasons"], [])

    def test_a_genuinely_hidden_hyperlink_less_sheet_is_still_caught(self):
        # The existing safety logic (is_non_authoritative_rollup_sheet) must
        # remain intact - a hidden sheet with no download lines at all is
        # still auto_skip, same as before this fix.
        wb = load_workbook(BEEM_PATH)
        ws = wb["Sheet1"]
        ws.sheet_state = "hidden"
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        result = extract_spreadsheet_gemini.classify_sheet_for_extraction(ws, text)

        self.assertEqual(result["outcome"], "auto_skip")

    def test_a_suspiciously_named_copy_of_the_same_sheet_is_still_ambiguous(self):
        wb = load_workbook(BEEM_PATH)
        ws = wb["Sheet1"]
        ws.title = "Archive"
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        result = extract_spreadsheet_gemini.classify_sheet_for_extraction(ws, text)

        self.assertEqual(result["outcome"], "ambiguous")
        self.assertIn("suspicious_sheet_name", result["reasons"])


class BeemProviderGuessTests(unittest.TestCase):
    def test_filename_guess_is_beem(self):
        self.assertEqual(extract_spreadsheet.guess_provider_name("beem Live Flex Availability.xlsx"), "beem")

    def test_flex_stopword_never_splits_a_compound_provider_name(self):
        # "flex" is only ever stripped as its own standalone word - a
        # provider whose OWN name happens to be one word containing "flex"
        # must be completely unaffected.
        self.assertEqual(extract_spreadsheet.guess_provider_name("Flexspace Availability.xlsx"), "Flexspace")


@unittest.skipUnless(_HAS_FIXTURE, "real Beem fixture not present in this environment")
class BeemFullUploadIntegrationTests(unittest.TestCase):
    """End to end through the real app.py upload flow - no mocking needed,
    since the fix means this file never reaches Gemini at all."""

    def setUp(self):
        _clear_pending()

    def tearDown(self):
        _clear_pending()

    def _upload_and_extract(self):
        at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
        at.run()
        at.file_uploader[0].upload(
            "beem Live Flex Availability.xlsx", _beem_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        at.run()
        extract_buttons = [b for b in at.button if b.label == "Extract"]
        extract_buttons[0].click().run()
        return at

    def test_no_old_or_summary_sheet_prompt_appears(self):
        at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
        at.run()
        at.file_uploader[0].upload(
            "beem Live Flex Availability.xlsx", _beem_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        at.run()

        markdown_text = "".join(m.value for m in at.markdown)
        self.assertNotIn("Possible summary or old sheet detected", markdown_text)

    def test_extraction_succeeds_with_no_exception(self):
        at = self._upload_and_extract()
        self.assertFalse(at.exception)

    def test_provider_becomes_beem(self):
        at = self._upload_and_extract()
        pending = list_pending_staging_files()
        df = load_staging_as_dataframe(pending[0])

        self.assertTrue((df["provider"] == "beem").all())

    def test_rows_with_an_explicit_building_are_all_extracted(self):
        # NOTE: the real sheet has 42 data rows, but most floors after the
        # first one for a given building leave the Property cell blank
        # (a real, common "same building as the row above" spreadsheet
        # convention) - extract_spreadsheet.py's header-mapped path has no
        # building-inheritance step at all (unlike extract.py's PDF path),
        # so a blank-building row is correctly (if incompletely) skipped
        # rather than mis-attributed to the wrong building - see storage.
        # file_store.dataframe_to_listing_rows' own "skip rows with no
        # building" docstring. This is a real, separate, pre-existing gap
        # (building-inheritance for header-mapped spreadsheets), not one of
        # the 3 reported bugs this test file targets - flagged, not fixed,
        # here. 14 is the correct count for THIS scope: one row per
        # distinct building actually stated in the sheet.
        at = self._upload_and_extract()
        pending = list_pending_staging_files()
        df = load_staging_as_dataframe(pending[0])

        self.assertEqual(len(df), 14)

    def test_a_real_rows_fields_are_correctly_mapped(self):
        at = self._upload_and_extract()
        pending = list_pending_staging_files()
        df = load_staging_as_dataframe(pending[0])

        first = df.iloc[0]
        self.assertEqual(first["building"], "New Derwent House WC1")
        self.assertEqual(first["floor_unit"], "4th Floor")
        self.assertEqual(first["size_sqft"], 2202.0)
        self.assertEqual(first["desks_min"], 32)
        self.assertEqual(first["desks_max"], 34)
        self.assertEqual(first["rent_psf"], 130.0)


@unittest.skipUnless(_HAS_FIXTURE, "real Beem fixture not present in this environment")
class BeemGeocodingAndSubmarketTests(unittest.TestCase):
    """
    Regression tests against the real Beem file's own Property column text -
    a real, confirmed report: several rows have valid postcode/lat/lng but a
    blank submarket, and two rows (New Derwent House WC1, Clove London
    Bridge SE1) never resolve a postcode/lat/lng at all after a conflicting
    Google Places candidate is correctly rejected. call_places_text_search/
    call_reverse_geocoding_api are mocked (no live network/API-key
    dependency, matching every other geocode test in this suite) - only the
    BUILDING TEXT itself comes from the real file, read directly rather than
    retyped, so a future real-file edit can never let this drift out of sync
    with what the source actually says.
    """

    @classmethod
    def setUpClass(cls):
        wb = load_workbook(BEEM_PATH, data_only=True)
        ws = wb["Sheet1"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        cls.buildings = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = dict(zip(headers, row))
            building = record.get("Property")
            if building:
                cls.buildings.setdefault(building.strip(), building)

    def _building(self, needle: str) -> str:
        match = next((b for b in self.buildings if needle in b), None)
        self.assertIsNotNone(match, f"expected a real Beem building containing {needle!r}")
        return self.buildings[match]

    def setUp(self):
        geocode.FAILURES.clear()

    def test_new_derwent_house_rejects_the_conflicting_top_candidate_then_finds_a_safe_one(self):
        building = self._building("New Derwent House")
        row = ListingRow(building=building, provider="beem")
        wrong_result = {
            "status": "OK", "lat": 51.5118097, "lng": -0.1414146,
            "address_components": [{"longText": "W1S 2ER", "types": ["postal_code"]}],
        }
        right_result = {
            "status": "OK", "lat": 51.5223, "lng": -0.1214,
            "address_components": [{"longText": "WC1X 8NP", "types": ["postal_code"]}],
        }

        def fake_places(query):
            return wrong_result if query == f"{building}, London, UK" else right_result

        with patch("geocode.call_places_text_search", side_effect=fake_places):
            geocode.geocode_row(row)

        self.assertEqual(row.postcode, "WC1X 8NP")
        self.assertEqual(geocode.FAILURES, [])

    def test_all_conflicting_candidates_leave_new_derwent_house_unresolved(self):
        # The pre-fix real symptom - and still the CORRECT outcome when
        # every candidate genuinely conflicts (never silently accept a
        # wrong location just to avoid a blank).
        building = self._building("New Derwent House")
        row = ListingRow(building=building, provider="beem")

        with patch(
            "geocode.call_places_text_search",
            return_value={
                "status": "OK", "lat": 51.5118097, "lng": -0.1414146,
                "address_components": [{"longText": "W1S 2ER", "types": ["postal_code"]}],
            },
        ):
            geocode.geocode_row(row)

        self.assertIsNone(row.lat)
        self.assertIsNone(row.postcode)
        self.assertEqual(len(geocode.FAILURES), 1)

    def test_source_locality_hints_backfill_blank_submarkets_with_no_google_call(self):
        # Every one of these real rows was reported with valid postcode/
        # lat/lng but blank submarket - source text already states the
        # locality, so it must be used instead of depending on Google's
        # own (confirmed patchy) reverse-geocode neighbourhood coverage.
        cases = [
            ("Orange Street", "Covent Garden"),
            ("Bayswater Road", "Paddington"),
            ("Nutmeg", "London Bridge"),
            ("Southwark Street", "London Bridge"),
        ]
        for needle, expected_area in cases:
            with self.subTest(building=needle):
                building = self._building(needle)
                row = ListingRow(building=building, provider="beem", lat=51.51, lng=-0.1)

                with patch("geocode.call_reverse_geocoding_api") as mock_reverse:
                    geocode.geocode_row(row)

                mock_reverse.assert_not_called()
                self.assertEqual(row.submarket, expected_area)

    def test_working_controls_still_resolve_the_same_submarket_as_before(self):
        # These two were already correct pre-fix (via Google's own
        # neighbourhood coverage) - confirms the new source-hint-first
        # priority produces the SAME answer, not a regression, for a
        # locality Google already handled well.
        cases = [("Red Lion Studios", "Clerkenwell"), ("Fashion Street", "Spitalfields")]
        for needle, expected_area in cases:
            with self.subTest(building=needle):
                building = self._building(needle)
                row = ListingRow(building=building, provider="beem", lat=51.52, lng=-0.1)

                with patch("geocode.call_reverse_geocoding_api") as mock_reverse:
                    geocode.geocode_row(row)

                mock_reverse.assert_not_called()  # now resolved from source text alone
                self.assertEqual(row.submarket, expected_area)


if __name__ == "__main__":
    unittest.main()
