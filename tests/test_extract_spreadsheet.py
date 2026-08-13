"""
Regression tests for extract_spreadsheet.py - the header-mapping path for
uploaded provider .xlsx/.csv spreadsheets (see the module docstring for the
overall flow: read_spreadsheet -> suggest_mapping (exact match, then a
conservative fuzzy fallback) -> unresolved_critical_fields (a targeted
rescue prompt, only if a CRITICAL field didn't map) -> build_rows).

Header strings used here for the "real UNION format" cases are taken
verbatim from the actual current-format UNION export files (checked
directly, not guessed) - not committed into this repo (real provider data
kept out of git-tracked fixtures, same principle as never using the real
data/master.xlsx in tests), just reproduced as plain header strings so the
synonym table is tested against the real thing it needs to handle.

Run with:
    .venv\\Scripts\\python.exe -m unittest tests.test_extract_spreadsheet -v
"""

import sys
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import MagicMock

import pandas as pd
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import extract_spreadsheet

REAL_UNION_HEADERS = [
    "External Ref", "Assigned Agents", "Property Address 1", "Property Postcode",
    "Lat", "Lng", "For Sale", "To Let", "Area", "Building", "Floor/Unit",
    "Size (sq ft)", "Desks (max)", "Marketing Price (Based on Min Term) PCM",
    "Marketing Price (Based on Min Term) PSF", "Brochure PDF", "Min. Term",
    "Special Features", "State of Space", "Legal Structure", "Broker Fee",
    "Contacts", "Floor Plan", "High Res Images",
]

# The OTHER real UNION export format actually seen in production - a much
# simpler "by-area" sheet (checked directly against the real Clerkenwell &
# Farringdon / Fitzrovia & Marylebone files, not guessed) whose building-
# name column is headered with the area's own name rather than any generic
# label - this is genuinely a different real format from REAL_UNION_HEADERS
# above, not a variant of it. Blank/nan-header columns present in the real
# files are omitted here since they carry no data and aren't relevant to
# what this constant tests.
REAL_UNION_BY_AREA_HEADERS = [
    "Clerkenwell & Farringdon", "Floor", "Current spec", "Size sq.ft",
    "Minimum Term", "Monthly Rate", "Price p/sq.ft", "Brochure",
]


class HeaderHashTests(unittest.TestCase):
    def test_same_headers_same_order_hash_identically(self):
        a = extract_spreadsheet.header_hash(["Building", "Floor/Unit"])
        b = extract_spreadsheet.header_hash(["Building", "Floor/Unit"])
        self.assertEqual(a, b)

    def test_different_order_hashes_differently(self):
        a = extract_spreadsheet.header_hash(["Building", "Floor/Unit"])
        b = extract_spreadsheet.header_hash(["Floor/Unit", "Building"])
        self.assertNotEqual(a, b)

    def test_different_headers_hash_differently(self):
        a = extract_spreadsheet.header_hash(["Building"])
        b = extract_spreadsheet.header_hash(["Building "])
        self.assertNotEqual(a, b)


class UnmappedCriticalFieldsTests(unittest.TestCase):
    def test_no_missing_fields_when_building_maps(self):
        mapping = {"Building": "building", "Floor": "floor_unit"}
        self.assertEqual(extract_spreadsheet.unmapped_critical_fields(mapping), [])

    def test_building_missing_when_unmapped(self):
        mapping = {"Floor": "floor_unit"}
        self.assertEqual(extract_spreadsheet.unmapped_critical_fields(mapping), ["building"])

    def test_address_1_is_deliberately_not_critical(self):
        # Considered (no fallback fills it in, unlike provider's filename-
        # based guess), but ruled out - see CRITICAL_FIELDS' own comment -
        # because the real Kitt's Availability file has NO address column
        # at all and must still be fully automatic; making address_1
        # critical would prompt for a rescue on Kitt's every single time.
        mapping = {"Building": "building"}  # no column maps to address_1
        self.assertEqual(extract_spreadsheet.unmapped_critical_fields(mapping), [])

    def test_real_union_by_area_headers_leave_building_unmapped(self):
        # The real UNION "by-area" export format - its building-name
        # column is headered with the area's own name, not any generic
        # label, so it never exact- or fuzzy-matches "building" (see
        # REAL_UNION_BY_AREA_HEADERS).
        mapping = extract_spreadsheet.suggest_mapping(REAL_UNION_BY_AREA_HEADERS)
        self.assertEqual(extract_spreadsheet.unmapped_critical_fields(mapping), ["building"])

    def test_real_kitts_and_full_union_headers_leave_nothing_critical_missing(self):
        # Both real formats map "building" on their own already - the
        # rescue mechanism must never interrupt either of them.
        kitts_headers = [
            "Area", "Building", "Floor/Unit", "Link to Brochure", "Key Features ", "State of Space",
        ]
        for headers in (REAL_UNION_HEADERS, kitts_headers):
            with self.subTest(headers=headers):
                mapping = extract_spreadsheet.suggest_mapping(headers)
                self.assertEqual(extract_spreadsheet.unmapped_critical_fields(mapping), [])


class ApplyCriticalFieldRescueTests(unittest.TestCase):
    def test_assigns_the_rescued_header_to_the_field(self):
        mapping = {"Clerkenwell & Farringdon": None, "Floor": "floor_unit"}
        rescued = extract_spreadsheet.apply_critical_field_rescue(
            mapping, {"building": "Clerkenwell & Farringdon"}
        )
        self.assertEqual(rescued["Clerkenwell & Farringdon"], "building")

    def test_a_none_assignment_leaves_the_field_unmapped(self):
        mapping = {"Floor": "floor_unit"}
        rescued = extract_spreadsheet.apply_critical_field_rescue(mapping, {"building": None})
        self.assertNotIn("building", rescued.values())

    def test_never_mutates_the_mapping_passed_in(self):
        mapping = {"Floor": "floor_unit"}
        extract_spreadsheet.apply_critical_field_rescue(mapping, {"building": "Floor"})
        self.assertEqual(mapping["Floor"], "floor_unit")

    def test_a_humans_choice_overrides_whatever_suggest_mapping_guessed(self):
        mapping = {"Floor": "floor_unit"}
        rescued = extract_spreadsheet.apply_critical_field_rescue(mapping, {"building": "Floor"})
        self.assertEqual(rescued["Floor"], "building")

    def test_rescuing_the_real_union_by_area_format_produces_real_rows(self):
        # End-to-end: the exact rescue a human would make in the app for
        # this real format (assign the area-name-headered column to
        # building) - confirms it actually unblocks extraction, not just
        # that the mapping dict looks right in isolation.
        df = pd.DataFrame([
            {h: None for h in REAL_UNION_BY_AREA_HEADERS} | {
                "Clerkenwell & Farringdon": "55 Goswell Road", "Floor": "3rd (South)", "Size sq.ft": 624,
            },
        ])
        mapping = extract_spreadsheet.suggest_mapping(REAL_UNION_BY_AREA_HEADERS)
        mapping = extract_spreadsheet.apply_critical_field_rescue(
            mapping, {"building": "Clerkenwell & Farringdon"}
        )
        self.assertEqual(extract_spreadsheet.unmapped_critical_fields(mapping), [])

        rows = extract_spreadsheet.build_rows(df, mapping, source_file="union.xlsx")

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].building, "55 Goswell Road")


class ApplyProviderStructuralFallbackTests(unittest.TestCase):
    """
    UNION's "by-area" format headers the Building column with the area's
    own name - different per file, so no synonym/fuzzy match can ever catch
    it (see REAL_UNION_BY_AREA_HEADERS/ApplyCriticalFieldRescueTests above).
    apply_provider_structural_fallback resolves this automatically, from
    the first upload of a brand new area onward, once the filename is
    recognized as UNION - no per-area rescue ever needed.
    """

    def test_real_clerkenwell_and_farringdon_file_resolves_automatically(self):
        mapping = extract_spreadsheet.suggest_mapping(REAL_UNION_BY_AREA_HEADERS)
        mapping = extract_spreadsheet.apply_provider_structural_fallback(
            mapping, REAL_UNION_BY_AREA_HEADERS,
            "UNION - Availability - June 26 - Clerkenwell & Farringdon.xlsx",
        )
        self.assertEqual(mapping["Clerkenwell & Farringdon"], "building")
        self.assertEqual(extract_spreadsheet.unresolved_critical_fields(mapping, {}), [])

    def test_every_real_by_area_variant_seen_so_far_resolves_automatically(self):
        # Each is a genuinely different area name/filename - including one
        # never seen when this fix was proposed (London Bridge & Southbank)
        # and one using "Category" rather than "Current spec" - proving
        # this generalizes rather than being another one-off patch.
        cases = [
            (["Fitzrovia & Marylebone", "Floor", "Current spec", "Size sq.ft",
              "Minimum Term", "Monthly Rate", "Price p/sq.ft", "Brochure"],
             "UNION - Availability - June 26 - Fitzrovia & Marylebone.xlsx", "Fitzrovia & Marylebone"),
            (["Mayfair / St James", "Floor", "Current spec", "Size sq.ft",
              "Minimum Term", "Monthly Rate", "Price p/sq.ft", "Brochure"],
             "UNION - Availability - June 26 - Mayfair & St James.xlsx", "Mayfair / St James"),
            (["Midtown", "Floor", "Current spec", "Size sq.ft",
              "Minimum Term", "Monthly Rate", "Price p/sq.ft", "Brochure"],
             "UNION - Availability - June 26 - Midtown.xlsx", "Midtown"),
            (["London Bridge / Southbank", "Floor", "Category", "Size sq.ft",
              "Minimum Term", "Monthly Rate", "Price p/sq.ft", "Brochure"],
             "UNION - Availabiliy - June 26 - London Bridge & Southbank.xlsx", "London Bridge / Southbank"),
            # The real leading blank column present in the actual files
            # (dropped from REAL_UNION_BY_AREA_HEADERS as irrelevant to what
            # that constant tests) - confirms this isn't sensitive to a
            # fixed absolute column index, only to the run of fixed headers.
            ([None, "Soho / Covent Garden", "Floor", "Current spec", "Size sq.ft",
              "Minimum Term", "Monthly Rate", "Price p/sq.ft", "Brochure"],
             "UNION - Availability - June 26 - Soho & Covent Garden.xlsx", "Soho / Covent Garden"),
        ]
        for headers, filename, expected_building_header in cases:
            with self.subTest(filename=filename):
                mapping = extract_spreadsheet.suggest_mapping(headers)
                mapping = extract_spreadsheet.apply_provider_structural_fallback(mapping, headers, filename)
                self.assertEqual(mapping[expected_building_header], "building")
                self.assertEqual(extract_spreadsheet.unresolved_critical_fields(mapping, {}), [])

    def test_real_kitts_and_full_union_headers_are_unaffected(self):
        # Both already map "building" on their own - the fallback must be
        # a complete no-op for them, not just harmless.
        kitts_headers = ["Area", "Building", "Floor/Unit", "Link to Brochure"]
        for headers, filename in (
            (REAL_UNION_HEADERS, "UNION - Shoreditch_2026-07-14.xlsx"),
            (kitts_headers, "Kitt's Availability (External).xlsx"),
        ):
            with self.subTest(filename=filename):
                mapping = extract_spreadsheet.suggest_mapping(headers)
                fallback_mapping = extract_spreadsheet.apply_provider_structural_fallback(
                    mapping, headers, filename
                )
                self.assertEqual(fallback_mapping, mapping)

    def test_never_overrides_an_already_mapped_building(self):
        mapping = {"Building": "building", "Clerkenwell & Farringdon": None}
        fallback_mapping = extract_spreadsheet.apply_provider_structural_fallback(
            mapping, ["Building", "Clerkenwell & Farringdon"],
            "UNION - Availability - June 26 - Clerkenwell & Farringdon.xlsx",
        )
        self.assertEqual(fallback_mapping["Building"], "building")
        self.assertIsNone(fallback_mapping["Clerkenwell & Farringdon"])

    def test_does_not_fire_for_an_unrecognized_provider_even_with_the_same_layout(self):
        # The exact real UNION by-area fingerprint, but from a filename that
        # isn't UNION - a coincidental structural match on a genuinely
        # different provider must never misfire.
        mapping = extract_spreadsheet.suggest_mapping(REAL_UNION_BY_AREA_HEADERS)
        mapping = extract_spreadsheet.apply_provider_structural_fallback(
            mapping, REAL_UNION_BY_AREA_HEADERS, "SomeOtherAgent - Availability.xlsx",
        )
        self.assertEqual(extract_spreadsheet.unresolved_critical_fields(mapping, {}), ["building"])

    def test_never_mutates_the_mapping_passed_in(self):
        mapping = extract_spreadsheet.suggest_mapping(REAL_UNION_BY_AREA_HEADERS)
        original = dict(mapping)
        extract_spreadsheet.apply_provider_structural_fallback(
            mapping, REAL_UNION_BY_AREA_HEADERS,
            "UNION - Availability - June 26 - Clerkenwell & Farringdon.xlsx",
        )
        self.assertEqual(mapping, original)

    def test_automatic_fallback_produces_real_rows_for_every_real_missing_submarket_address(self):
        # End-to-end, no human rescue involved at all - the same real
        # Clerkenwell & Farringdon rows currently missing submarket
        # (see geocode.py's SubmarketBackfillTests) actually extract with
        # the right building value straight away.
        addresses = ["55 Goswell Road", "13 Northburgh Street", "50 Great Sutton Street"]
        df = pd.DataFrame([
            {h: None for h in REAL_UNION_BY_AREA_HEADERS} | {"Clerkenwell & Farringdon": address}
            for address in addresses
        ])
        mapping = extract_spreadsheet.suggest_mapping(REAL_UNION_BY_AREA_HEADERS)
        mapping = extract_spreadsheet.apply_provider_structural_fallback(
            mapping, REAL_UNION_BY_AREA_HEADERS,
            "UNION - Availability - June 26 - Clerkenwell & Farringdon.xlsx",
        )
        self.assertEqual(extract_spreadsheet.unresolved_critical_fields(mapping, {}), [])

        rows = extract_spreadsheet.build_rows(df, mapping, source_file="union.xlsx")

        self.assertEqual([r.building for r in rows], addresses)


class UnresolvedCriticalFieldsTests(unittest.TestCase):
    def test_a_field_with_no_rescue_entry_at_all_is_unresolved(self):
        mapping = {"Floor": "floor_unit"}
        self.assertEqual(extract_spreadsheet.unresolved_critical_fields(mapping, {}), ["building"])

    def test_a_confirmed_blank_field_is_not_unresolved(self):
        # A rescue entry of None means "a human already confirmed this
        # format genuinely has no such column" - must not keep prompting
        # for it on every future upload of the same format.
        mapping = {"Floor": "floor_unit"}
        rescue = {"building": None}
        self.assertEqual(extract_spreadsheet.unresolved_critical_fields(mapping, rescue), [])

    def test_a_real_rescue_assignment_resolves_the_field(self):
        rescue = {"building": "Clerkenwell & Farringdon"}
        mapping = extract_spreadsheet.apply_critical_field_rescue(
            {"Clerkenwell & Farringdon": None}, rescue
        )
        self.assertEqual(extract_spreadsheet.unresolved_critical_fields(mapping, rescue), [])


class SuggestMappingTests(unittest.TestCase):
    def test_maps_the_real_union_header_format(self):
        guess = extract_spreadsheet.suggest_mapping(REAL_UNION_HEADERS)

        self.assertEqual(guess["Building"], "building")
        self.assertEqual(guess["Floor/Unit"], "floor_unit")
        self.assertEqual(guess["Size (sq ft)"], "size_sqft")
        self.assertEqual(guess["Desks (max)"], "desks_max")
        self.assertEqual(guess["Marketing Price (Based on Min Term) PCM"], "rent_pcm")
        self.assertEqual(guess["Marketing Price (Based on Min Term) PSF"], "rent_psf")
        self.assertEqual(guess["Brochure PDF"], "brochure_link")
        self.assertEqual(guess["Special Features"], "special_features")
        self.assertEqual(guess["State of Space"], "state_of_space")
        self.assertEqual(guess["Contacts"], "contacts")
        self.assertEqual(guess["Property Address 1"], "address_1")
        self.assertEqual(guess["Property Postcode"], "postcode")
        self.assertEqual(guess["Lat"], "lat")
        self.assertEqual(guess["Lng"], "lng")
        self.assertEqual(guess["Area"], "submarket")
        self.assertEqual(guess["External Ref"], "internal_ref")
        self.assertEqual(guess["Assigned Agents"], "provider")

    def test_columns_with_no_known_synonym_are_left_unmapped(self):
        guess = extract_spreadsheet.suggest_mapping(REAL_UNION_HEADERS)

        # These have no ListingRow equivalent at all - must never be forced
        # onto some unrelated field.
        for header in ("For Sale", "To Let", "Min. Term", "Legal Structure", "Broker Fee", "High Res Images"):
            self.assertIsNone(guess[header], f"{header!r} should have no guess")

        # "Floor Plan" DOES have a real equivalent now - floorplan_link (see
        # schema.py) - and must map to it, not be left unmapped.
        self.assertEqual(guess["Floor Plan"], "floorplan_link")

    def test_our_own_staging_xlsx_header_format_round_trips(self):
        # staging_writer.write_rows_to_xlsx's own Title-Case header labels
        # (e.g. "Size Sqft" for the size_sqft field) should map straight
        # back onto themselves via the field-name/title-case auto-synonym,
        # with no extra synonym table entry needed.
        guess = extract_spreadsheet.suggest_mapping(["Building", "Floor Unit", "Size Sqft", "Contacts"])
        self.assertEqual(guess["Building"], "building")
        self.assertEqual(guess["Size Sqft"], "size_sqft")
        self.assertEqual(guess["Contacts"], "contacts")

    def test_never_maps_two_headers_onto_the_same_field(self):
        # Both "Brochure PDF" and "Link to File" are known brochure_link
        # synonyms - only the first (in header order) should win; the
        # second must be left for a human to resolve rather than silently
        # dropped or double-mapped.
        guess = extract_spreadsheet.suggest_mapping(["Brochure PDF", "Link to File"])
        self.assertEqual(guess["Brochure PDF"], "brochure_link")
        self.assertIsNone(guess["Link to File"])

    def test_maps_the_real_kitts_brochure_link_header_and_close_variants(self):
        # "Link to Brochure" is the real header used by the actual Kitt's
        # Availability (External).xlsx export - previously unmapped,
        # defaulting to "(ignore)" in the confirm-mapping UI on first
        # upload of that format.
        guess = extract_spreadsheet.suggest_mapping(["Link to Brochure", "Brochure Link", "Link to Brochure PDF"])
        self.assertEqual(guess["Link to Brochure"], "brochure_link")
        self.assertEqual(guess["Brochure Link"], None)  # already used by "Link to Brochure" above
        self.assertEqual(guess["Link to Brochure PDF"], None)  # same

    def test_brochure_link_close_variant_maps_alone(self):
        for header in ("Link to Brochure", "Brochure Link", "Link to Brochure PDF"):
            with self.subTest(header=header):
                guess = extract_spreadsheet.suggest_mapping([header])
                self.assertEqual(guess[header], "brochure_link")

    def test_maps_the_real_kitts_key_features_header(self):
        # "Key Features" (note: trailing space in the real header cell,
        # stripped by normalize_key) is the real column used by the actual
        # Kitt's Availability (External).xlsx export for descriptive
        # amenity/notes text - the same concept as special_features, just
        # different wording from "Special Features".
        guess = extract_spreadsheet.suggest_mapping(["Key Features "])
        self.assertEqual(guess["Key Features "], "special_features")

    def test_real_kitts_header_set_maps_every_genuinely_mappable_column_and_nothing_else(self):
        # The full real header set from Kitt's Availability (External).xlsx
        # - confirms the confirm-mapping UI would show a fully correct
        # mapping on first upload (10 real fields, nothing guessed wrong)
        # with every genuinely unmappable column still correctly left for
        # a human to see as "(ignore)", not silently guessed.
        headers = [
            "Area", "Building", "Floor/Unit", "Size \n(sq ft)", "Desks \n(max)",
            "Marketing Price \n(Based on Min Term)\nPCM", "Marketing Price \n(Based on Min Term)\nPSF",
            "Link to Brochure", "Min. term", "Key Features ", "State of Space", "Legal Structure",
            "Broker Fee", "Marketing Permission", "Commercial Model", "Patch?",
            "Unit Lead - Viewings go to this person initially", "Unit Support - back up cover for viewings",
            "Who Onboarded?", "Landlord/Agent Onboarded", "Other Info", "Access Information",
            "Link to Floorplan", "Link to High Res Images", "Matterport Link",
        ]
        guess = extract_spreadsheet.suggest_mapping(headers)

        expected_mapped = {
            "Area": "submarket",
            "Building": "building",
            "Floor/Unit": "floor_unit",
            "Size \n(sq ft)": "size_sqft",
            "Desks \n(max)": "desks_max",
            "Marketing Price \n(Based on Min Term)\nPCM": "rent_pcm",
            "Marketing Price \n(Based on Min Term)\nPSF": "rent_psf",
            "Link to Brochure": "brochure_link",
            "Key Features ": "special_features",
            "State of Space": "state_of_space",
            "Link to Floorplan": "floorplan_link",
        }
        for header, field in expected_mapped.items():
            self.assertEqual(guess[header], field, f"{header!r} should map to {field!r}")

        for header in set(headers) - set(expected_mapped):
            self.assertIsNone(guess[header], f"{header!r} has no real field and must not be guessed")

    def test_never_suggests_an_unmappable_field(self):
        guess = extract_spreadsheet.suggest_mapping(["Property Id", "Source File"])
        self.assertIsNone(guess["Property Id"])
        self.assertIsNone(guess["Source File"])

    def test_real_union_by_area_header_set_maps_every_genuinely_mappable_column(self):
        # The full real header set from the UNION "by-area" export format
        # (see REAL_UNION_BY_AREA_HEADERS) - "Current spec", "Monthly
        # Rate", "Price p/sq.ft", and "Brochure" previously left every one
        # of these 4 real columns unmapped even after the fuzzy fallback
        # (none cleared the 0.84 bidirectional-coverage gate - see
        # EXTRA_SYNONYMS' own comment for the actual scores computed), so
        # real rows from this format had rent_pcm/rent_psf/brochure_link/
        # state_of_space permanently blank. "Minimum Term" genuinely has
        # no ListingRow equivalent and must stay unmapped.
        guess = extract_spreadsheet.suggest_mapping(REAL_UNION_BY_AREA_HEADERS)

        expected_mapped = {
            "Floor": "floor_unit",
            "Current spec": "state_of_space",
            "Size sq.ft": "size_sqft",
            "Monthly Rate": "rent_pcm",
            "Price p/sq.ft": "rent_psf",
            "Brochure": "brochure_link",
        }
        for header, field in expected_mapped.items():
            self.assertEqual(guess[header], field, f"{header!r} should map to {field!r}")

        self.assertIsNone(guess["Minimum Term"])


class SuggestMappingFuzzyFallbackTests(unittest.TestCase):
    def test_catches_near_miss_variants_of_a_fields_own_name_or_synonym(self):
        # None of these exact-match FIELD_SYNONYMS - each is a plausible
        # real-world rewording (pluralization, added punctuation/
        # parentheses, an extra connective word) that a human confirming
        # the mapping would obviously accept, so the fuzzy fallback should
        # too, with no synonym-table entry needed.
        cases = {
            "Rent PCM": "rent_pcm",
            "Rent (PCM)": "rent_pcm",
            "Special Feature": "special_features",
            "State of the Space": "state_of_space",
            "Assigned Agent": "provider",
            "Brochure Link": "brochure_link",
        }
        for header, field in cases.items():
            with self.subTest(header=header):
                guess = extract_spreadsheet.suggest_mapping([header])
                self.assertEqual(guess[header], field, f"{header!r} should fuzzy-match {field!r}")

    def test_does_not_force_a_match_for_genuinely_unrelated_headers(self):
        # These are real UNION/Kitt's columns with no ListingRow equivalent
        # at all - confirmed during the fuzzy-match threshold investigation
        # to be the closest real-world near-misses seen (up to 0.667
        # bidirectional word similarity against some field, well under the
        # 0.84 threshold) - included here as a standing regression check
        # against the threshold ever being loosened past that headroom.
        headers = [
            "For Sale", "To Let", "Min. Term", "Legal Structure", "Broker Fee",
            "High Res Images", "Patch?", "Marketing Permission",
            "Commercial Model", "Who Onboarded?", "Landlord/Agent Onboarded",
            "Other Info", "Access Information",
            "Link to High Res Images", "Matterport Link",
        ]
        for header in headers:
            with self.subTest(header=header):
                guess = extract_spreadsheet.suggest_mapping([header])
                self.assertIsNone(guess[header], f"{header!r} should not be fuzzy-matched to anything")

        # "Floor Plan"/"Link to Floorplan" DO have a real equivalent now -
        # floorplan_link (see schema.py) - matched via the exact-synonym
        # pass, not even reaching the fuzzy fallback this test covers.
        for header in ("Floor Plan", "Link to Floorplan"):
            with self.subTest(header=header):
                guess = extract_spreadsheet.suggest_mapping([header])
                self.assertEqual(guess[header], "floorplan_link")

    def test_never_fuzzy_matches_an_unmappable_field(self):
        # "Property Id" is close enough in shape to trip a looser fuzzy
        # design onto an unrelated real field (e.g. "property postcode") -
        # property_id/source_file must stay unreachable even via the fuzzy
        # pass, same as the exact-match pass already guarantees.
        guess = extract_spreadsheet.suggest_mapping(["Property Id", "Source File"])
        self.assertIsNone(guess["Property Id"])
        self.assertIsNone(guess["Source File"])

    def test_fuzzy_pass_never_double_maps_a_field_already_used_by_an_exact_match(self):
        # "Special Features" exact-matches special_features first; "Key
        # Feature" (singular, a plausible fuzzy near-miss of the "Key
        # Features" synonym) must be left unmapped rather than stealing the
        # same field a real column already claimed.
        guess = extract_spreadsheet.suggest_mapping(["Special Features", "Key Feature"])
        self.assertEqual(guess["Special Features"], "special_features")
        self.assertIsNone(guess["Key Feature"])

    def test_fuzzy_pass_never_double_maps_a_field_between_two_fuzzy_candidates(self):
        # Both are plausible near-misses of the same field - only the first
        # (in header order) should win, exactly like the exact-match pass.
        guess = extract_spreadsheet.suggest_mapping(["Rent PCM", "Rent (PCM)"])
        self.assertEqual(guess["Rent PCM"], "rent_pcm")
        self.assertIsNone(guess["Rent (PCM)"])


class BuildRowsTests(unittest.TestCase):
    def test_applies_mapping_and_sets_source_file(self):
        df = pd.DataFrame([
            {"Building": "40 New Bond Street", "Floor/Unit": "3rd Floor", "Size (sq ft)": 5000, "Ignore Me": "x"},
        ])
        mapping = {"Building": "building", "Floor/Unit": "floor_unit", "Size (sq ft)": "size_sqft", "Ignore Me": None}

        rows = extract_spreadsheet.build_rows(df, mapping, source_file="union_export.xlsx")

        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row.building, "40 New Bond Street")
        self.assertEqual(row.floor_unit, "3rd Floor")
        self.assertEqual(row.size_sqft, 5000.0)
        self.assertEqual(row.source_file, "union_export.xlsx")

    def test_all_blank_row_is_skipped(self):
        df = pd.DataFrame([
            {"Building": "City Tower", "Floor/Unit": "5th Floor"},
            {"Building": None, "Floor/Unit": None},
        ])
        mapping = {"Building": "building", "Floor/Unit": "floor_unit"}

        rows = extract_spreadsheet.build_rows(df, mapping, source_file="a.xlsx")

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].building, "City Tower")

    def test_a_row_with_blank_building_but_a_non_blank_other_field_is_skipped_not_crashed(self):
        # Grounded in a real Kitt's Availability row: a spreadsheet-
        # author's own section-header/note ("COMING SOON / ADDITIONAL
        # OPTIONS TO SHARE" in the Area column), every other column -
        # including Building - blank. Not all-blank, so the old "skip only
        # if every field is None" check let it straight through to
        # ListingRow(building=None, ...), which raised and aborted the
        # WHOLE file's extraction over this one non-property row.
        df = pd.DataFrame([
            {"Building": "City Tower", "Floor/Unit": "5th Floor"},
            {"Building": None, "Floor/Unit": None, "Area": "COMING SOON / ADDITIONAL OPTIONS TO SHARE"},
        ])
        mapping = {"Building": "building", "Floor/Unit": "floor_unit", "Area": "submarket"}

        rows = extract_spreadsheet.build_rows(df, mapping, source_file="kitts.xlsx")

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].building, "City Tower")

    def test_a_fractional_value_in_an_int_field_is_rounded_not_a_crash(self):
        # Grounded in a real Kitt's Availability row: desks_max resolved
        # to 38.257143 (almost certainly a computed/averaged cell caught
        # up in the mapped column range) - desks_max is Optional[int], and
        # Pydantic's int validation rejects a fractional float outright,
        # which previously aborted the WHOLE file's extraction over this
        # one cell.
        df = pd.DataFrame([{"Building": "33 Cavendish Square", "Desks": 38.257143}])
        mapping = {"Building": "building", "Desks": "desks_max"}

        rows = extract_spreadsheet.build_rows(df, mapping, source_file="kitts.xlsx")

        self.assertEqual(rows[0].desks_max, 38)

    def test_a_whole_number_float_in_an_int_field_still_comes_through_correctly(self):
        df = pd.DataFrame([{"Building": "City Tower", "Desks": 12.0}])
        mapping = {"Building": "building", "Desks": "desks_max"}

        rows = extract_spreadsheet.build_rows(df, mapping, source_file="a.xlsx")

        self.assertEqual(rows[0].desks_max, 12)

    def test_a_genuinely_blank_cell_in_an_int_field_is_none_not_a_crash(self):
        # Grounded in a real Kitt's Availability file: most of its 949 rows
        # are blank padding rows below the real data (skipped already, via
        # the blank-building check) - but the real desks_max column has its
        # own genuinely blank cells among real rows too, which pandas reads
        # back as an actual float NaN, not None. isinstance(nan, float) is
        # True, so it slid straight through the "already a real number"
        # branch as if it were one - round(nan) then raised ValueError and
        # aborted the WHOLE file's extraction over this one blank cell.
        df = pd.DataFrame([{"Building": "City Tower", "Desks": float("nan")}])
        mapping = {"Building": "building", "Desks": "desks_max"}

        rows = extract_spreadsheet.build_rows(df, mapping, source_file="kitts.xlsx")

        self.assertIsNone(rows[0].desks_max)

    def test_non_numeric_placeholder_in_a_numeric_column_becomes_none_not_a_crash(self):
        # Grounded in a real current-format UNION file: its Lat/Lng columns
        # hold the literal text "Needs manual lookup" for rows the provider
        # hasn't geocoded yet themselves - must not blow up the entire
        # file's extraction over one placeholder cell in one column.
        df = pd.DataFrame([
            {"Building": "111 Wardour Street", "Lat": "Needs manual lookup", "Lng": "Needs manual lookup"},
        ])
        mapping = {"Building": "building", "Lat": "lat", "Lng": "lng"}

        rows = extract_spreadsheet.build_rows(df, mapping, source_file="union.xlsx")

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].building, "111 Wardour Street")
        self.assertIsNone(rows[0].lat)
        self.assertIsNone(rows[0].lng)

    def test_real_numeric_lat_lng_still_come_through_correctly(self):
        df = pd.DataFrame([{"Building": "City Tower", "Lat": 51.5, "Lng": -0.1}])
        mapping = {"Building": "building", "Lat": "lat", "Lng": "lng"}

        rows = extract_spreadsheet.build_rows(df, mapping, source_file="union.xlsx")

        self.assertEqual(rows[0].lat, 51.5)
        self.assertEqual(rows[0].lng, -0.1)

    def test_unmapped_columns_never_reach_listingrow(self):
        # A column mapped to None (or simply absent from the mapping dict)
        # must never surface as an unexpected kwarg - ListingRow's own
        # extra="ignore" would tolerate it anyway, but build_rows should
        # not even pass it through.
        df = pd.DataFrame([{"Building": "City Tower", "Legal Structure": "Freehold"}])
        mapping = {"Building": "building", "Legal Structure": None}

        rows = extract_spreadsheet.build_rows(df, mapping, source_file="a.xlsx")

        self.assertEqual(rows[0].building, "City Tower")

    def test_build_rows_never_fills_in_a_provider_on_its_own(self):
        # Provider fallback is app.py's fill_missing_provider's job now
        # (applied uniformly to every upload type, not just spreadsheets) -
        # build_rows itself must leave a genuinely unmapped provider as
        # None, not guess at one.
        df = pd.DataFrame([{"Building": "City Tower"}])
        mapping = {"Building": "building"}

        rows = extract_spreadsheet.build_rows(df, mapping, source_file="a.xlsx")

        self.assertIsNone(rows[0].provider)
        self.assertIsNone(rows[0].internal_ref)


class GuessProviderNameTests(unittest.TestCase):
    def test_strips_parenthetical_asides_and_boilerplate_words(self):
        # The real filename this feature was built for.
        self.assertEqual(
            extract_spreadsheet.guess_provider_name("Kitt's Availability (External).xlsx"), "Kitt's"
        )

    def test_strips_embedded_dates(self):
        guess = extract_spreadsheet.guess_provider_name("UNION - Soho & Covent Garden_2026-07-17.xlsx")
        self.assertNotIn("2026", guess)
        self.assertNotIn("07", guess)

    def test_falls_back_to_the_stem_when_nothing_is_left(self):
        # Every word here is boilerplate - stripping all of them would
        # leave an empty guess, which must never happen.
        self.assertEqual(extract_spreadsheet.guess_provider_name("Availability Export.xlsx"), "Availability Export")

    def test_csv_extension_is_stripped_too(self):
        self.assertEqual(extract_spreadsheet.guess_provider_name("Breezblok.csv"), "Breezblok")

    def test_strips_a_month_name_and_number_date_with_no_punctuation_between_them(self):
        # The real UNION by-area filename convention - "June 26" has
        # neither punctuation between the month and the number nor an
        # all-numeric shape, so _FILENAME_DATE_RE alone doesn't catch it;
        # it previously leaked straight into the guess as "UNION June 26
        # Fitzrovia Marylebone".
        self.assertEqual(
            extract_spreadsheet.guess_provider_name(
                "UNION - Availability - June 26 - Fitzrovia & Marylebone.xlsx"
            ),
            "UNION",
        )

    def test_strips_a_number_then_month_name_date_too(self):
        self.assertEqual(extract_spreadsheet.guess_provider_name("UNION - 26 June.xlsx"), "UNION")

    def test_strips_known_london_area_names(self):
        # Every real UNION "by-area" export filename seen so far - each is
        # a genuinely different header format (see CRITICAL_FIELDS'
        # comment), but should still guess the same underlying provider.
        # None of these areas is on any stopword list - see
        # _KNOWN_PROVIDER_NAMES - "UNION" is recognized as the whole answer
        # regardless of what area name follows it.
        cases = [
            "UNION - Availability - June 26 - Clerkenwell & Farringdon.xlsx",
            "UNION - Availability - June 26 - Fitzrovia & Marylebone.xlsx",
            "UNION - Availability - June 26 - Soho & Covent Garden.xlsx",
            "UNION - Shoreditch_2026-07-14.xlsx",
        ]
        for filename in cases:
            with self.subTest(filename=filename):
                self.assertEqual(extract_spreadsheet.guess_provider_name(filename), "UNION")

    def test_the_real_kitts_pdf_filename_still_guesses_correctly(self):
        self.assertEqual(
            extract_spreadsheet.guess_provider_name(
                "Kitt's Availability (External) - Live Availability.pdf"
            ),
            "Kitt's",
        )

    def test_a_brand_new_area_never_seen_before_still_guesses_union_correctly(self):
        # The real trigger for this design: a new UNION by-area export for
        # an area not covered by any previous fix (previously "London
        # Bridge"/"Southwark" would have leaked straight into the guess,
        # same as every other new area before it) - proves the fix is
        # general, not just another entry added reactively for this one.
        self.assertEqual(
            extract_spreadsheet.guess_provider_name(
                "UNION - Availability - June 26 - London Bridge & Southwark.xlsx"
            ),
            "UNION",
        )

    def test_an_entirely_novel_area_name_also_works_with_no_list_update_needed(self):
        # Not a real filename seen yet - demonstrates the design handles
        # ANY future area, not just ones added to a list after the fact.
        self.assertEqual(
            extract_spreadsheet.guess_provider_name(
                "UNION - Availability - September 26 - Bermondsey & Elephant and Castle.xlsx"
            ),
            "UNION",
        )

    def test_a_provider_name_that_merely_starts_with_the_same_letters_is_not_matched(self):
        # _leading_known_provider matches a whole word, not a prefix - a
        # hypothetical different provider whose name happens to start with
        # the same characters as a known one must never be misidentified.
        self.assertEqual(
            extract_spreadsheet.guess_provider_name("Unionville Estates Availability.xlsx"),
            "Unionville Estates",
        )


class ParseXludfFallbackTests(unittest.TestCase):
    """
    Formula strings here are taken verbatim from a real Google Sheets/
    IMPORTRANGE .xlsx export (Kitt's Availability (External).xlsx, not
    committed into this repo) - checked directly, not guessed.
    """

    def test_quoted_string_fallback_with_nested_importrange_call(self):
        formula = (
            '=IFERROR(__xludf.DUMMYFUNCTION("IMPORTRANGE(""https://docs.google.com/spreadsheets/d/abc/edit#gid=1"",'
            '""\'Availability\'!A:ab"")"),"Area")'
        )
        self.assertEqual(extract_spreadsheet._parse_xludf_fallback(formula), "Area")

    def test_quoted_string_fallback_with_embedded_newline(self):
        formula = '=IFERROR(__xludf.DUMMYFUNCTION("""COMPUTED_VALUE"""),"Size \n(sq ft)")'
        self.assertEqual(extract_spreadsheet._parse_xludf_fallback(formula), "Size \n(sq ft)")

    def test_numeric_fallback(self):
        formula = '=IFERROR(__xludf.DUMMYFUNCTION("""COMPUTED_VALUE"""),759.0)'
        self.assertEqual(extract_spreadsheet._parse_xludf_fallback(formula), 759.0)

    def test_quoted_string_fallback_with_a_pound_sign(self):
        formula = '=IFERROR(__xludf.DUMMYFUNCTION("""COMPUTED_VALUE"""),"£296")'
        self.assertEqual(extract_spreadsheet._parse_xludf_fallback(formula), "£296")

    def test_not_a_formula_at_all_returns_none(self):
        self.assertIsNone(extract_spreadsheet._parse_xludf_fallback("just plain text"))
        self.assertIsNone(extract_spreadsheet._parse_xludf_fallback(None))
        self.assertIsNone(extract_spreadsheet._parse_xludf_fallback(759.0))

    def test_a_different_kind_of_formula_returns_none_rather_than_guessing(self):
        self.assertIsNone(extract_spreadsheet._parse_xludf_fallback("=SUM(A1:A10)"))


class ResolveCellValueTests(unittest.TestCase):
    def _cell(self, value):
        cell = MagicMock()
        cell.value = value
        return cell

    def test_prefers_the_cached_value_when_present(self):
        value_cell = self._cell("Area")
        formula_cell = self._cell('=IFERROR(__xludf.DUMMYFUNCTION("..."),"Area")')
        self.assertEqual(extract_spreadsheet._resolve_cell_value(value_cell, formula_cell), "Area")

    def test_falls_back_to_the_formula_when_the_cache_is_missing(self):
        value_cell = self._cell(None)
        formula_cell = self._cell('=IFERROR(__xludf.DUMMYFUNCTION("""COMPUTED_VALUE"""),"Building")')
        self.assertEqual(extract_spreadsheet._resolve_cell_value(value_cell, formula_cell), "Building")

    def test_a_genuinely_blank_cell_stays_none(self):
        value_cell = self._cell(None)
        formula_cell = self._cell(None)
        self.assertIsNone(extract_spreadsheet._resolve_cell_value(value_cell, formula_cell))


class ReadSpreadsheetXludfIntegrationTests(unittest.TestCase):
    """
    Builds a workbook via openpyxl containing REAL cached-value-missing
    formula cells (assigning a formula string to a cell and saving it
    without ever letting a real spreadsheet engine compute/cache a result
    is exactly what "no cached value available" means) - end-to-end proof
    that read_spreadsheet resolves such a file's real header/data text
    correctly, not just the underlying helper functions in isolation.
    """

    def _build_xlsx_bytes(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws["A1"] = '=IFERROR(__xludf.DUMMYFUNCTION("IMPORTRANGE(""url"",""range"")"),"Building")'
        ws["B1"] = '=IFERROR(__xludf.DUMMYFUNCTION("""COMPUTED_VALUE"""),"Size \n(sq ft)")'
        ws["A2"] = '=IFERROR(__xludf.DUMMYFUNCTION("""COMPUTED_VALUE"""),"28 Bruton Street")'
        ws["B2"] = '=IFERROR(__xludf.DUMMYFUNCTION("""COMPUTED_VALUE"""),759.0)'
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def test_headers_resolve_to_real_text_not_formula_text(self):
        df = extract_spreadsheet.read_spreadsheet(self._build_xlsx_bytes(), ".xlsx")
        self.assertEqual(list(df.columns), ["Building", "Size \n(sq ft)"])

    def test_data_rows_resolve_to_real_values_not_formula_text(self):
        df = extract_spreadsheet.read_spreadsheet(self._build_xlsx_bytes(), ".xlsx")
        self.assertEqual(df.iloc[0]["Building"], "28 Bruton Street")
        self.assertEqual(df.iloc[0]["Size \n(sq ft)"], 759.0)

    def test_a_plain_non_formula_workbook_is_completely_unaffected(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Building"
        ws["B1"] = "Size (sq ft)"
        ws["A2"] = "28 Bruton Street"
        ws["B2"] = 759
        buffer = BytesIO()
        wb.save(buffer)

        df = extract_spreadsheet.read_spreadsheet(buffer.getvalue(), ".xlsx")

        self.assertEqual(list(df.columns), ["Building", "Size (sq ft)"])
        self.assertEqual(df.iloc[0]["Building"], "28 Bruton Street")
        self.assertEqual(df.iloc[0]["Size (sq ft)"], 759)


class ListSheetNamesTests(unittest.TestCase):
    def test_csv_has_no_sheet_concept(self):
        self.assertEqual(extract_spreadsheet.list_sheet_names(b"a,b\n1,2\n", ".csv"), [None])

    def test_xlsx_returns_every_sheet_in_file_order(self):
        wb = Workbook()
        wb.active.title = "City"
        wb.create_sheet("Mid Town")
        wb.create_sheet("Portfolio")
        buffer = BytesIO()
        wb.save(buffer)

        names = extract_spreadsheet.list_sheet_names(buffer.getvalue(), ".xlsx")

        self.assertEqual(names, ["City", "Mid Town", "Portfolio"])


class ReadSpreadsheetSheetNameTests(unittest.TestCase):
    def _build_two_sheet_xlsx(self) -> bytes:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "City"
        ws1["A1"] = "Building"
        ws1["A2"] = "28 Lime Street"

        ws2 = wb.create_sheet("Mid Town")
        ws2["A1"] = "Building"
        ws2["A2"] = "89 Charterhouse Street"

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def test_no_sheet_name_defaults_to_active_sheet(self):
        # Regression check: every pre-existing caller omits sheet_name and
        # must keep reading the same (active) sheet it always did.
        df = extract_spreadsheet.read_spreadsheet(self._build_two_sheet_xlsx(), ".xlsx")
        self.assertEqual(df.iloc[0]["Building"], "28 Lime Street")

    def test_explicit_sheet_name_reads_that_sheet(self):
        df = extract_spreadsheet.read_spreadsheet(self._build_two_sheet_xlsx(), ".xlsx", sheet_name="Mid Town")
        self.assertEqual(df.iloc[0]["Building"], "89 Charterhouse Street")


class ClassifyLinkLabelTests(unittest.TestCase):
    """classify_link_label - the pure text->type classifier, no I/O."""

    def test_floor_plan_label_variants_classify_as_floorplan(self):
        for text in ("Floor Plan", "Floorplan", "FLOOR PLAN", "View Floor Plan", "Download Floorplan"):
            self.assertEqual(extract_spreadsheet.classify_link_label(text), "floorplan", text)

    def test_brochure_label_variants_classify_as_brochure(self):
        for text in ("Brochure", "BROCHURE", "View Brochure", "Download Brochure"):
            self.assertEqual(extract_spreadsheet.classify_link_label(text), "brochure", text)

    def test_combined_document_label_classifies_as_brochure_only(self):
        # A single document that's both stays brochure - never floorplan,
        # never both (see the module's own docstring).
        self.assertEqual(
            extract_spreadsheet.classify_link_label("Download Brochure and Floorplans"), "brochure",
        )

    def test_uninformative_label_is_unclassified(self):
        for text in ("Click Here", "View", "Open", "", None):
            self.assertIsNone(extract_spreadsheet.classify_link_label(text))


class BrochureFloorplanLinkClassificationTests(unittest.TestCase):
    """
    End-to-end regression tests for the confirmed real gap: extract_
    spreadsheet.py's header-mapped fast path (build_rows) previously copied
    every row's link straight through to whatever field the COLUMN header
    mapped to, discarding the per-row hyperlink label's own, stronger
    evidence of that specific cell's real document type - e.g. a column
    headered "Brochure" (or another brochure_link synonym) whose individual
    rows are actually a mix of real brochure and floor-plan links,
    distinguished only by each cell's own visible hyperlink text (a real,
    confirmed shape: floor plans delivered as a Box.com link visibly
    labeled "FLOOR PLAN" sitting in an otherwise brochure-headed column).
    """

    def _xlsx_bytes(self, headers: list, rows: list) -> bytes:
        """
        rows: list of lists, one per data row, each entry either a plain
        value or a (display_text, url) tuple - the tuple becomes a real
        openpyxl hyperlink whose OWN cell text is display_text (never the
        URL itself), exactly the shape read_spreadsheet's own hyperlink_
        display_text capture is for.
        """
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append([v[0] if isinstance(v, tuple) else v for v in row])
            row_idx = ws.max_row
            for col_idx, value in enumerate(row, start=1):
                if isinstance(value, tuple):
                    display_text, url = value
                    ws.cell(row=row_idx, column=col_idx).hyperlink = url
                    ws.cell(row=row_idx, column=col_idx).value = display_text
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _build_rows(self, headers: list, rows: list) -> list:
        data = self._xlsx_bytes(headers, rows)
        df = extract_spreadsheet.read_spreadsheet(data, ".xlsx")
        mapping = extract_spreadsheet.suggest_mapping(list(df.columns))
        return extract_spreadsheet.build_rows(df, mapping, source_file="test.xlsx")

    def test_floor_plan_label_with_box_url_lands_in_floorplan_link(self):
        rows = self._build_rows(
            ["Building", "Brochure"],
            [["Nutmeg House", ("FLOOR PLAN", "https://app.box.com/s/floorplan123")]],
        )
        self.assertEqual(rows[0].floorplan_link, "https://app.box.com/s/floorplan123")
        self.assertIsNone(rows[0].brochure_link)

    def test_brochure_label_with_box_url_stays_brochure_link(self):
        rows = self._build_rows(
            ["Building", "Brochure"],
            [["Kent House", ("BROCHURE", "https://app.box.com/s/brochure456")]],
        )
        self.assertEqual(rows[0].brochure_link, "https://app.box.com/s/brochure456")
        self.assertIsNone(rows[0].floorplan_link)

    def test_hyperlink_display_text_is_preserved_alongside_target(self):
        data = self._xlsx_bytes(
            ["Building", "Brochure"],
            [["Nutmeg House", ("FLOOR PLAN", "https://app.box.com/s/floorplan123")]],
        )
        df = extract_spreadsheet.read_spreadsheet(data, ".xlsx")
        self.assertEqual(df.iloc[0]["Brochure"], "https://app.box.com/s/floorplan123")
        self.assertEqual(df.attrs["hyperlink_display_text"]["Brochure"][0], "FLOOR PLAN")

    def test_explicit_floorplan_label_outranks_a_generic_synonym_column(self):
        # "Link to File" is a real brochure_link synonym (see EXTRA_
        # SYNONYMS) with no wording of its own suggesting either document
        # type - a genuinely generic column header, unlike "Brochure"
        # itself. The per-row label must still win.
        rows = self._build_rows(
            ["Building", "Link to File"],
            [["Nutmeg House", ("Download Floorplan", "https://app.box.com/s/fp789")]],
        )
        self.assertEqual(rows[0].floorplan_link, "https://app.box.com/s/fp789")
        self.assertIsNone(rows[0].brochure_link)

    def test_brochure_containing_floorplan_pages_remains_brochure(self):
        rows = self._build_rows(
            ["Building", "Brochure"],
            [["Nutmeg House", ("Download Brochure and Floorplans", "https://app.box.com/s/combined")]],
        )
        self.assertEqual(rows[0].brochure_link, "https://app.box.com/s/combined")
        self.assertIsNone(rows[0].floorplan_link)

    def test_separate_brochure_and_floorplan_columns_both_survive(self):
        rows = self._build_rows(
            ["Building", "Brochure", "Floor Plan"],
            [[
                "Nutmeg House",
                ("BROCHURE", "https://app.box.com/s/brochure1"),
                ("FLOOR PLAN", "https://app.box.com/s/floorplan1"),
            ]],
        )
        self.assertEqual(rows[0].brochure_link, "https://app.box.com/s/brochure1")
        self.assertEqual(rows[0].floorplan_link, "https://app.box.com/s/floorplan1")

    def test_pdf_floorplan_link_classifies_as_floorplan_not_brochure(self):
        # The real, confirmed VERSE Building case: a floor plan can itself
        # be a PDF, so extension alone is never a safe signal either way.
        rows = self._build_rows(
            ["Building", "Brochure"],
            [["Verse Building", ("FLOOR PLAN", "https://app.box.com/s/verse-1st-floor.pdf")]],
        )
        self.assertEqual(rows[0].floorplan_link, "https://app.box.com/s/verse-1st-floor.pdf")
        self.assertIsNone(rows[0].brochure_link)

    def test_png_floorplan_link_classifies_as_floorplan_not_brochure(self):
        rows = self._build_rows(
            ["Building", "Brochure"],
            [["Nutmeg House", ("FLOOR PLAN", "https://app.box.com/s/floorplan.png")]],
        )
        self.assertEqual(rows[0].floorplan_link, "https://app.box.com/s/floorplan.png")
        self.assertIsNone(rows[0].brochure_link)

    def test_placeholder_value_produces_neither_link(self):
        rows = self._build_rows(["Building", "Brochure"], [["Nutmeg House", "TBC"]])
        self.assertIsNone(rows[0].brochure_link)
        self.assertIsNone(rows[0].floorplan_link)

    def test_provider_name_does_not_affect_classification(self):
        rows = self._build_rows(
            ["Building", "Provider", "Brochure"],
            [["Nutmeg House", "UNION", ("FLOOR PLAN", "https://app.box.com/s/floorplan123")]],
        )
        self.assertEqual(rows[0].floorplan_link, "https://app.box.com/s/floorplan123")
        self.assertIsNone(rows[0].brochure_link)

        rows_other_provider = self._build_rows(
            ["Building", "Provider", "Brochure"],
            [["Nutmeg House", "Some Other Agent", ("FLOOR PLAN", "https://app.box.com/s/floorplan123")]],
        )
        self.assertEqual(rows_other_provider[0].floorplan_link, "https://app.box.com/s/floorplan123")


if __name__ == "__main__":
    unittest.main()
