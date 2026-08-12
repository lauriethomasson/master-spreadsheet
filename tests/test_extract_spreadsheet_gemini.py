"""
Regression tests for extract_spreadsheet_gemini.py - the Gemini text-
extraction fallback for spreadsheet sheets with no single consistent header
row (see the module docstring). Never calls the real Gemini API - get_client/
call_gemini are mocked throughout, same principle as test_brochure_link.py
mocking GCS: this dev environment has no live-model determinism to test
against, only the deterministic code around it (rendering, row construction,
brochure-link resolution, building inheritance).

Run with:
    .venv\\Scripts\\python.exe -m unittest tests.test_extract_spreadsheet_gemini -v
"""

import contextlib
import io
import sys
import unittest
from datetime import date
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import extract_spreadsheet_gemini


class RenderSheetAsTextTests(unittest.TestCase):
    def _sheet(self, rows):
        wb = Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        return ws

    def test_blank_rows_are_skipped_entirely(self):
        ws = self._sheet([["Building"], [None], ["28 Lime Street"]])
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)
        self.assertEqual(text, "Row 1: Building\nRow 3: 28 Lime Street")

    def test_merged_continuation_blank_cells_are_preserved_not_dropped(self):
        # Real Copthall Portfolio-sheet shape: Area/Building blank on a
        # continuation row must render as an empty slot between pipes, not
        # be silently compacted away - column position is the only signal
        # Gemini has that this row inherits the value above it.
        ws = self._sheet([
            ["Area", "Building", "Floor"],
            ["City", "Copthall Avenue", "LG Floor"],
            [None, None, "2nd Floor"],
        ])
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)
        lines = text.splitlines()
        self.assertEqual(lines[2], "Row 3:  |  | 2nd Floor")

    def test_trailing_blank_cells_are_trimmed(self):
        ws = self._sheet([["Building", None, None]])
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)
        self.assertEqual(text, "Row 1: Building")

    def test_hyperlink_target_is_inlined_after_display_text(self):
        ws = self._sheet([["Download Brochure"]])
        ws["A1"].hyperlink = "https://example.com/brochure.pdf"
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)
        self.assertEqual(text, "Row 1: Download Brochure (https://example.com/brochure.pdf)")

    def test_hyperlink_with_no_display_text_uses_bare_url(self):
        ws = self._sheet([[None]])
        ws["A1"].hyperlink = "https://example.com/brochure.pdf"
        # openpyxl's hyperlink setter auto-populates .value to match when the
        # cell had none - reset it to actually construct "a hyperlink with no
        # separate display text" rather than "display text that happens to
        # equal the URL".
        ws["A1"].value = None
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)
        self.assertEqual(text, "Row 1: https://example.com/brochure.pdf")

    def test_all_blank_sheet_renders_to_empty_string(self):
        ws = self._sheet([[None, None], [None]])
        self.assertEqual(extract_spreadsheet_gemini.render_sheet_as_text(ws), "")


class ExtractSheetTests(unittest.TestCase):
    def _sheet(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Office"
        ws["A2"] = "4th Floor"
        return ws

    def test_empty_sheet_never_calls_gemini(self):
        wb = Workbook()
        ws = wb.active  # genuinely blank
        with patch("extract_spreadsheet_gemini.get_client") as mock_get_client, \
             patch("extract_spreadsheet_gemini.call_gemini") as mock_call_gemini:
            rows = extract_spreadsheet_gemini.extract_sheet(ws, "file.xlsx — Sheet1", "file.xlsx")

        self.assertEqual(rows, [])
        mock_get_client.assert_not_called()
        mock_call_gemini.assert_not_called()

    def test_builds_listing_rows_from_gemini_output(self):
        raw = {
            "provider": "Copthall Estates",
            "contacts": "Kiri Norton-Brennan, 0795 811 8382",
            "units": [
                {"building": "28 Lime Street", "floor_unit": "4th Floor", "size_sqft": 1358,
                 "rent_pcm": 19805, "rent_psf": 175, "special_features": "22 desks"},
            ],
        }
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            rows = extract_spreadsheet_gemini.extract_sheet(self._sheet(), "file.xlsx — City", "file.xlsx")

        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row.building, "28 Lime Street")
        self.assertEqual(row.provider, "Copthall Estates")
        self.assertEqual(row.contacts, "Kiri Norton-Brennan, 0795 811 8382")
        self.assertEqual(row.size_sqft, 1358)
        self.assertEqual(row.source_file, "file.xlsx — City")

    def test_unit_with_blank_building_inherits_from_prior_unit(self):
        raw = {
            "provider": None, "contacts": None,
            "units": [
                {"building": "111 Wardour Street", "floor_unit": "4th"},
                {"building": None, "floor_unit": "5th"},
            ],
        }
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            rows = extract_spreadsheet_gemini.extract_sheet(self._sheet(), "file.xlsx — Sheet1", "file.xlsx")

        self.assertEqual([r.building for r in rows], ["111 Wardour Street", "111 Wardour Street"])

    def test_unit_with_blank_building_and_no_prior_is_skipped(self):
        raw = {"provider": None, "contacts": None, "units": [{"building": None, "floor_unit": "4th"}]}
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            rows = extract_spreadsheet_gemini.extract_sheet(self._sheet(), "file.xlsx — Sheet1", "file.xlsx")

        self.assertEqual(rows, [])

    def test_no_units_returns_empty_list(self):
        raw = {"provider": "Copthall Estates", "contacts": None, "units": []}
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            rows = extract_spreadsheet_gemini.extract_sheet(self._sheet(), "file.xlsx — Incentives", "file.xlsx")

        self.assertEqual(rows, [])


class AddressHouseNumberVerificationTests(unittest.TestCase):
    """The real reported failure: Gemini transcribed a raw sheet cell of
    "14-18 Copthall Avenue, EC2R 7DJ" as address_1 "18 Copthall Avenue",
    silently dropping the first half of a hyphenated house-number range.
    extract_sheet must catch and correct this against the raw sheet text
    itself - deterministic regex parsing of a cell can't drop characters the
    way an LLM transcription can - rather than relying solely on
    master_merge.py's house_number_changed to catch it later (and only if
    something else ever re-uploads the correct value to compare against)."""

    def _copthall_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["28 Lime Street - City"])
        ws.append(["Modern refurbished offices with excellent natural light."])
        ws.append(["14-18 Copthall Avenue, EC2R 7DJ", "Download Floorplans"])
        ws.append(["Floor", "Size (sq ft)", "Rent"])
        ws.append(["4th Floor", 1358, 19805])
        return ws

    def test_dropped_range_is_restored_from_the_raw_postcode_line(self):
        raw = {
            "provider": "Copthall Estates", "contacts": None,
            "units": [{
                "building": "28 Lime Street", "address_1": "18 Copthall Avenue",
                "postcode": "EC2R 7DJ", "floor_unit": "4th Floor", "size_sqft": 1358,
            }],
        }
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            rows = extract_spreadsheet_gemini.extract_sheet(self._copthall_sheet(), "file.xlsx — City", "file.xlsx")

        self.assertEqual(rows[0].address_1, "14-18 Copthall Avenue")

    def test_correct_extraction_is_left_untouched(self):
        # Gemini got it right this time - the raw line agrees with address_1,
        # so nothing should be rewritten at all.
        raw = {
            "provider": "Copthall Estates", "contacts": None,
            "units": [{
                "building": "28 Lime Street", "address_1": "14-18 Copthall Avenue",
                "postcode": "EC2R 7DJ", "floor_unit": "4th Floor", "size_sqft": 1358,
            }],
        }
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            rows = extract_spreadsheet_gemini.extract_sheet(self._copthall_sheet(), "file.xlsx — City", "file.xlsx")

        self.assertEqual(rows[0].address_1, "14-18 Copthall Avenue")

    def test_no_confident_raw_match_warns_but_does_not_override_or_crash(self):
        # No postcode given, and the unit's building never appears as its
        # own heading line in this sheet at all - nothing to confidently
        # verify against, so address_1 must survive unchanged and the call
        # must still complete normally, just with a warning on stderr.
        raw = {
            "provider": None, "contacts": None,
            "units": [{
                "building": "Some Other Building Not In This Sheet",
                "address_1": "18 Copthall Avenue", "floor_unit": "4th Floor",
            }],
        }
        wb = Workbook()
        ws = wb.active
        ws.append(["Unrelated content with no address or postcode at all"])

        stderr = io.StringIO()
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw), \
             contextlib.redirect_stderr(stderr):
            rows = extract_spreadsheet_gemini.extract_sheet(ws, "file.xlsx — City", "file.xlsx")

        self.assertEqual(rows[0].address_1, "18 Copthall Avenue")
        self.assertIn("WARNING", stderr.getvalue())

    def test_units_sharing_one_building_block_dont_bleed_into_each_others_address(self):
        # Two floors of the same building (no postcode on either) - the
        # block-boundary logic must still land on THIS building's own address
        # line, not wander into a neighboring building's block.
        wb = Workbook()
        ws = wb.active
        ws.append(["28 Lime Street - City"])
        ws.append(["14-18 Copthall Avenue"])
        ws.append(["Floor", "Size (sq ft)"])
        ws.append(["4th Floor", 1358])
        ws.append(["5th Floor", 2000])
        ws.append(["40 New Bond Street - Mayfair"])
        ws.append(["1 New Bond Street"])
        ws.append(["Floor", "Size (sq ft)"])
        ws.append(["Ground Floor", 900])

        raw = {
            "provider": None, "contacts": None,
            "units": [
                {"building": "28 Lime Street", "address_1": "18 Copthall Avenue", "floor_unit": "4th Floor"},
                {"building": "28 Lime Street", "address_1": "18 Copthall Avenue", "floor_unit": "5th Floor"},
                {"building": "40 New Bond Street", "address_1": "1 New Bond Street", "floor_unit": "Ground Floor"},
            ],
        }
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            rows = extract_spreadsheet_gemini.extract_sheet(ws, "file.xlsx — City", "file.xlsx")

        self.assertEqual(rows[0].address_1, "14-18 Copthall Avenue")
        self.assertEqual(rows[1].address_1, "14-18 Copthall Avenue")
        self.assertEqual(rows[2].address_1, "1 New Bond Street")


class UndercountedBuildingsTests(unittest.TestCase):
    """Real production case: a live Gemini call against the real Copthall
    Estates "Mid Town" sheet's Cursitor Street mini-table (G/LG East, 1st
    Floor, 4th Floor) returned only 1 unit - a silent, validly-parsed but
    short response no existing check catches, since the one row that DID
    survive (G/LG East) had a completely normal size_sqft/rent_pcm. This
    reproduces that exact sheet's shape - confirmed against the real file:
    a blank leading column on every row, including each building's own
    heading line, two building blocks separated by a prose amenities
    paragraph and an address/download-link line."""

    def _mid_town_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws.append([None, "Charterhouse Street - Farringdon / Barbican"])
        ws.append([None, "Nestled in the vibrant Farringdon/Smithfield quarter, this stylish workspace features high-speed fibre."])
        ws.append([None, "89 Charterhouse Street, EC1M 6PE", None, None, None, "Download Floorplans"])
        ws.append([
            None, "Office", "Sq.Ft", "Price Per Sq.Ft", "Monthly List Price",
            "Office Description", "Minimum Term", "Available From", "Commission",
        ])
        ws.append([
            None, "3rd Floor", 1960.0, 175.0, 28583.0,
            "Currently set up with 24 desks, 1 large boardroom, 1 meeting room, 1 chat room, 1 phone booths "
            "and separate breakout/kitchen",
            "24 Months", "Now", "Up to 15% - See incentives tab",
        ])
        ws.append([None, "Cursitor Street - Chancery Lane"])
        ws.append([
            None, "Cursitor Street offers fully furnished, self-contained office spaces with meeting rooms, "
            "kitchens, high-speed internet, air-conditioning, showers, bike storage, a passenger lift, and "
            "24/7 access.",
        ])
        ws.append([None, "11 Cursitor Street, EC4A 1LL", None, None, None, "Download Floorplans"])
        ws.append([
            None, "Office", "Sq.Ft", "Price Per Sq.Ft", "Monthly List Price",
            "Office Description", "Minimum Term", "Available From", "Commission",
        ])
        ws.append([
            None, "G/LG East", 1800.0, 143.0, 21379.0,
            "24+ desks on the ground floor and a dedicated kitchen, and a breakout area in the LG floor. "
            "The layout also provides the option for a private meeting room on the ground floor, with its "
            "location to be determined by the tenant",
            "24 Months", "Now",
            "10% on the first 12 months, 2% commission on months 13-24 for any deals with a lease term of "
            "36 months or longer. Payable when the deal enters year 3.  Bonus Available for 4th and G/LG "
            "East Floors **See incentives tab",
        ])
        ws.append([
            None, "1st Floor", 1696.0, 160.0, 22614.0,
            "18 desks, 1 exec office, 10 person boardroom, dedicated kitchen and breakout space",
            "24 Months", "1st October 2026",
        ])
        ws.append([
            None, "4th Floor", 706.0, 160.0, 9413.0,
            "12 desks, dedicated kitchen, dedicated 8 person boardroom",
            "24 Months", "Now",
        ])
        return ws

    def test_partial_response_is_flagged(self):
        # The exact real failure: only G/LG East (the first, longest-
        # described row) survived out of Cursitor Street's 3 floors. Also
        # the true-positive re-check for the heading-anchor/boundary rework
        # below - this must still be caught, not just the 3 false positives
        # it fixes.
        text = extract_spreadsheet_gemini.render_sheet_as_text(self._mid_town_sheet())
        units = [
            {"building": "Charterhouse Street"},
            {"building": "Cursitor Street"},
        ]

        mismatches = extract_spreadsheet_gemini.find_undercounted_buildings(text, units)

        self.assertEqual(mismatches, [("Cursitor Street", 3, 1)])

    def test_correct_extraction_has_no_mismatch(self):
        text = extract_spreadsheet_gemini.render_sheet_as_text(self._mid_town_sheet())
        units = [
            {"building": "Charterhouse Street"},
            {"building": "Cursitor Street"},
            {"building": "Cursitor Street"},
            {"building": "Cursitor Street"},
        ]

        self.assertEqual(extract_spreadsheet_gemini.find_undercounted_buildings(text, units), [])

    def test_single_floor_building_has_no_false_positive(self):
        # Charterhouse's own 1-floor block must never be flagged just
        # because its address/header lines are also multi-column - checked
        # alongside Cursitor Street's full, correct 3-unit extraction (the
        # real calling convention: app.py always passes every unit
        # extracted from the WHOLE sheet at once, never a single
        # building's units in isolation - block boundaries are only
        # anchored against buildings that actually appear in `units`, so a
        # building missing entirely from that list can't bound anything
        # after it).
        text = extract_spreadsheet_gemini.render_sheet_as_text(self._mid_town_sheet())
        units = [
            {"building": "Charterhouse Street"},
            {"building": "Cursitor Street"},
            {"building": "Cursitor Street"},
            {"building": "Cursitor Street"},
        ]

        self.assertEqual(extract_spreadsheet_gemini.find_undercounted_buildings(text, units), [])

    def test_building_not_found_in_text_returns_no_mismatch(self):
        mismatches = extract_spreadsheet_gemini.find_undercounted_buildings(
            "Row 1: unrelated content", [{"building": "Nowhere"}],
        )
        self.assertEqual(mismatches, [])


class UndercountedBuildingsFalsePositiveTests(unittest.TestCase):
    """The first version of find_undercounted_buildings/_apparent_data_row_
    count had 3 confirmed real false positives against the actual Copthall
    Estates Availability file's City, Westend Soho, and Portfolio sheets -
    each reproduced here against the exact shape that triggered it, plus the
    Mid Town true positive re-checked above in UndercountedBuildingsTests."""

    def _sheet(self, rows):
        wb = Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        return ws

    def test_building_name_mentioned_in_another_buildings_description_is_ignored(self):
        # Real false positive: a plain substring search for a building's
        # name landed inside a COMPLETELY DIFFERENT building's own prose
        # description paragraph (which happened to mention it in passing),
        # rather than that building's own real heading line further down -
        # here "The Scalpel" is name-dropped inside Moor House's description,
        # before The Scalpel's own heading and block even appear.
        text = extract_spreadsheet_gemini.render_sheet_as_text(self._sheet([
            [None, "Moor House - Moorgate"],
            [None, "Just five minutes from The Scalpel and Bank station, this striking building offers "
                   "24/7 access and a rooftop terrace."],
            [None, "1 London Wall, EC2Y 5EA", None, None, None, "Download Floorplans"],
            [None, "Office", "Sq.Ft", "Rent PCM", "Available From"],
            [None, "2nd Floor", 2200.0, 15000.0, "Now"],
            [None, "The Scalpel - Leadenhall"],
            [None, "An instantly recognisable landmark tower with panoramic City views."],
            [None, "52 Lime Street, EC3M 7AF", None, None, None, "Download Floorplans"],
            [None, "Office", "Sq.Ft", "Rent PCM", "Available From"],
            [None, "5th Floor", 3100.0, 21000.0, "Now"],
            [None, "9th Floor", 2900.0, 19500.0, "1st October 2026"],
        ]))
        units = [
            {"building": "Moor House"},
            {"building": "The Scalpel"},
            {"building": "The Scalpel"},
        ]

        mismatches = extract_spreadsheet_gemini.find_undercounted_buildings(text, units)

        self.assertEqual(mismatches, [])

    def test_fully_occupied_building_with_zero_units_does_not_inflate_neighbor(self):
        # Real false positive: a building with NO surviving units at all
        # (correctly, per PROMPT, since its own mini-table just says "Fully
        # Occupied") never appears in `units`, so it can't bound the
        # PRECEDING building's own block boundary by its heading line -
        # the old substring-based boundary ran straight past it to the next
        # SURVIVING building's heading, silently counting the fully-occupied
        # building's own mini-table header row as an extra data row for the
        # building before it.
        text = extract_spreadsheet_gemini.render_sheet_as_text(self._sheet([
            [None, "Bankside House - Southwark"],
            [None, "A refurbished period building moments from Southwark station."],
            [None, "10 Bankside, SE1 9EY", None, None, None, "Download Floorplans"],
            [None, "Office", "Sq.Ft", "Rent PCM", "Available From"],
            [None, "1st Floor", 1500.0, 12000.0, "Now"],
            [None, "Riverside Building - Southwark"],
            [None, "A modern glass-fronted office building overlooking the Thames."],
            [None, "15 Riverside Walk, SE1 9EZ", None, None, None, "Download Floorplans"],
            [None, "Office", "Sq.Ft", "Rent PCM", "Available From"],
            [None, "Fully Occupied"],
            [None, "Skyline Tower - Southwark"],
            [None, "A striking new-build tower with flexible floorplates."],
            [None, "20 Skyline Way, SE1 9FA", None, None, None, "Download Floorplans"],
            [None, "Office", "Sq.Ft", "Rent PCM", "Available From"],
            [None, "3rd Floor", 1800.0, 14000.0, "Now"],
            [None, "4th Floor", 1750.0, 13500.0, "Now"],
        ]))
        units = [
            {"building": "Bankside House"},
            {"building": "Skyline Tower"},
            {"building": "Skyline Tower"},
        ]

        mismatches = extract_spreadsheet_gemini.find_undercounted_buildings(text, units)

        self.assertEqual(mismatches, [])

    def test_single_consistent_table_sheet_is_never_flagged(self):
        # Real false positive: a single-consistent-table sheet (PROMPT shape
        # (a) - one sheet-wide header row, building repeated per data row,
        # no per-building heading/address/download lines at all) isn't the
        # repeating-blocks shape this heuristic was designed for. With no
        # per-building end-of-block anchor available, the old code ran the
        # last building's own block all the way to the end of the sheet,
        # picking up trailing boilerplate rows as if they were extra units.
        text = extract_spreadsheet_gemini.render_sheet_as_text(self._sheet([
            [None, "Building", "Floor", "Size (sq ft)", "Rent PCM"],
            [None, "Portfolio Building A", "1st Floor", 1000, 20000],
            [None, "Portfolio Building A", "2nd Floor", 1200, 24000],
            [None, "Portfolio Building B", "Ground Floor", 900, 18000],
            [None, "Terms and Conditions apply", "See website for details", "Not for reliance"],
            [None, "This document does not constitute an offer or contract", "E&OE", "v2"],
        ]))
        units = [
            {"building": "Portfolio Building A"},
            {"building": "Portfolio Building A"},
            {"building": "Portfolio Building B"},
        ]

        mismatches = extract_spreadsheet_gemini.find_undercounted_buildings(text, units)

        self.assertEqual(mismatches, [])


class SheetShowsFullyOccupiedBuildingTests(unittest.TestCase):
    def test_true_when_a_fully_occupied_row_is_present(self):
        text = (
            "Row 1: | Riverside Building - Southwark\n"
            "Row 2: | A modern glass-fronted office building.\n"
            "Row 3: | 15 Riverside Walk, SE1 9EZ | | | | Download Floorplans\n"
            "Row 4: | Office | Sq.Ft | Rent PCM | Available From\n"
            "Row 5: | Fully Occupied"
        )
        self.assertTrue(extract_spreadsheet_gemini.sheet_shows_fully_occupied_building(text))

    def test_false_when_no_fully_occupied_row_is_present(self):
        text = "Row 1: unrelated content\nRow 2: nothing here either"
        self.assertFalse(extract_spreadsheet_gemini.sheet_shows_fully_occupied_building(text))


class BrochureLinkDisambiguationTests(unittest.TestCase):
    """PROMPT's brochure_link field note - real, confirmed against the
    actual Copthall Estates Availability file: every building's own
    address/link line in shape (b) carries TWO separate hyperlinks side by
    side (e.g. the real City sheet's row 11, 27-30 Lime Street) - a
    "Download Floorplans" link and a separate "Download Brochure" link,
    never the same asset twice. Only the "Brochure"-labeled one belongs in
    brochure_link. Never calls the real Gemini API (see this module's own
    docstring) - reproduces that real block's raw-text shape for context,
    and mocks Gemini's own return value as if it had already correctly
    picked the Brochure link per the strengthened prompt, proving that
    choice survives end to end into the final ListingRow untouched (not
    silently swapped for the Floorplans link, or dropped, by anything
    downstream - e.g. finalize_brochure_link)."""

    def _lime_street_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws.append([None, "27-30 Lime Street - Fenchurch St / Bank"])
        ws.append([None, "A striking City tower with a full range of on-site amenities."])
        ws.append([
            None, "27-30 Lime Street, EC3M 7AF", None, None, None,
            "Download Floorplans", None, None, "Download Brochure",
        ])
        ws["F3"].hyperlink = "https://lmstern.sharepoint.com/floorplans/27-30-lime-street"
        ws["I3"].hyperlink = "https://copthallestates.com/brochures/27-30-lime-street.pdf"
        ws.append([None, "Office", "Sq.Ft", "Rent PCM", "Available From"])
        ws.append([None, "5th Floor", 2400.0, 18000.0, "Now"])
        return ws

    def test_brochure_labeled_link_is_used_not_the_floorplans_one(self):
        raw = {
            "provider": "Copthall Estates",
            "contacts": None,
            "units": [{
                "building": "27-30 Lime Street", "floor_unit": "5th Floor",
                "size_sqft": 2400, "rent_pcm": 18000,
                "brochure_link": "https://copthallestates.com/brochures/27-30-lime-street.pdf",
            }],
        }
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            rows = extract_spreadsheet_gemini.extract_sheet(
                self._lime_street_sheet(), "file.xlsx — City", "file.xlsx"
            )

        self.assertEqual(rows[0].brochure_link, "https://copthallestates.com/brochures/27-30-lime-street.pdf")
        self.assertNotEqual(rows[0].brochure_link, "https://lmstern.sharepoint.com/floorplans/27-30-lime-street")


class FindBuildingsMissingBrochureLinkTests(unittest.TestCase):
    """find_buildings_missing_brochure_link - the structural safety net for
    a different real, confirmed failure: every building checked in the real
    Copthall Estates Availability file has a genuine, working Download
    Brochure link in the source, so a building whose own block clearly has
    one but whose extracted unit(s) came back with brochure_link=None is
    worth a warning, exactly like find_undercounted_buildings already does
    for row counts."""

    def _sheet(self, rows):
        wb = Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        return ws

    def test_flags_a_building_with_a_source_brochure_link_but_no_extracted_one(self):
        text = extract_spreadsheet_gemini.render_sheet_as_text(self._sheet([
            [None, "Riverside Building - Southwark"],
            [None, "A modern glass-fronted office building overlooking the Thames."],
            [
                None, "15 Riverside Walk, SE1 9EZ", None, None, None,
                "Download Floorplans (https://a.example.com/floorplans)", None, None,
                "Download Brochure (https://b.example.com/brochure.pdf)",
            ],
            [None, "Office", "Sq.Ft", "Rent PCM", "Available From"],
            [None, "3rd Floor", 1800.0, 14000.0, "Now"],
        ]))
        units = [{"building": "Riverside Building", "brochure_link": None}]

        missing = extract_spreadsheet_gemini.find_buildings_missing_brochure_link(text, units)

        self.assertEqual(missing, ["Riverside Building"])

    def test_not_flagged_when_a_unit_has_the_brochure_link(self):
        text = extract_spreadsheet_gemini.render_sheet_as_text(self._sheet([
            [None, "Riverside Building - Southwark"],
            [None, "A modern glass-fronted office building overlooking the Thames."],
            [
                None, "15 Riverside Walk, SE1 9EZ", None, None, None,
                "Download Brochure (https://b.example.com/brochure.pdf)",
            ],
            [None, "Office", "Sq.Ft", "Rent PCM", "Available From"],
            [None, "3rd Floor", 1800.0, 14000.0, "Now"],
        ]))
        units = [{"building": "Riverside Building", "brochure_link": "https://b.example.com/brochure.pdf"}]

        missing = extract_spreadsheet_gemini.find_buildings_missing_brochure_link(text, units)

        self.assertEqual(missing, [])

    def test_not_flagged_when_source_never_had_a_brochure_link_at_all(self):
        # Not every provider supplies one - a building with none in the
        # source at all is not itself suspicious.
        text = extract_spreadsheet_gemini.render_sheet_as_text(self._sheet([
            [None, "Riverside Building - Southwark"],
            [None, "A modern glass-fronted office building overlooking the Thames."],
            [None, "15 Riverside Walk, SE1 9EZ", None, None, None, "Download Floorplans"],
            [None, "Office", "Sq.Ft", "Rent PCM", "Available From"],
            [None, "3rd Floor", 1800.0, 14000.0, "Now"],
        ]))
        units = [{"building": "Riverside Building", "brochure_link": None}]

        missing = extract_spreadsheet_gemini.find_buildings_missing_brochure_link(text, units)

        self.assertEqual(missing, [])

    def test_bare_word_brochure_with_no_real_link_is_not_a_false_alarm(self):
        # "brochure" appearing in running prose with no actual hyperlink
        # target next to it is not the same thing as a real Download
        # Brochure link - must not be mistaken for one.
        text = extract_spreadsheet_gemini.render_sheet_as_text(self._sheet([
            [None, "Riverside Building - Southwark"],
            [None, "See our brochure tab for full incentive details."],
            [None, "15 Riverside Walk, SE1 9EZ", None, None, None, "Download Floorplans"],
            [None, "Office", "Sq.Ft", "Rent PCM", "Available From"],
            [None, "3rd Floor", 1800.0, 14000.0, "Now"],
        ]))
        units = [{"building": "Riverside Building", "brochure_link": None}]

        missing = extract_spreadsheet_gemini.find_buildings_missing_brochure_link(text, units)

        self.assertEqual(missing, [])

    def test_building_block_not_found_returns_no_mismatch(self):
        missing = extract_spreadsheet_gemini.find_buildings_missing_brochure_link(
            "Row 1: unrelated content", [{"building": "Nowhere", "brochure_link": None}],
        )
        self.assertEqual(missing, [])


class DeterministicBrochureLinkForBuildingTests(unittest.TestCase):
    """deterministic_brochure_link_for_building - reads the real Excel
    hyperlink behind a building's own "Download Brochure" cell directly,
    bypassing Gemini's own free-form reading of the same rendered text
    entirely. Confirmed against the real Copthall Estates Availability.xlsx:
    every building's address/link row carries Floorplans and Brochure as two
    separate cells, each with its own distinct real URL."""

    def _sheet(self, rows):
        wb = Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        return ws

    def _riverside_block(self, ws):
        ws.append([None, "Riverside Building - Southwark"])
        ws.append([None, "A modern glass-fronted office building overlooking the Thames."])
        ws.append([
            None, "15 Riverside Walk, SE1 9EZ", None, None, None,
            "Download Floorplans", None, None, "Download Brochure",
        ])
        ws["F3"].hyperlink = "https://a.example.com/floorplans"
        ws["I3"].hyperlink = "https://b.example.com/brochure.pdf"

    def test_recovers_the_single_unambiguous_brochure_link(self):
        ws = self._sheet([])
        self._riverside_block(ws)
        ws.append([None, "Office", "Sq.Ft", "Rent PCM", "Available From"])
        ws.append([None, "3rd Floor", 1800.0, 14000.0, "Now"])
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        url = extract_spreadsheet_gemini.deterministic_brochure_link_for_building(text, "Riverside Building")

        self.assertEqual(url, "https://b.example.com/brochure.pdf")

    def test_never_returns_the_floorplans_link(self):
        ws = self._sheet([])
        self._riverside_block(ws)
        ws.append([None, "Office", "Sq.Ft", "Rent PCM", "Available From"])
        ws.append([None, "3rd Floor", 1800.0, 14000.0, "Now"])
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        url = extract_spreadsheet_gemini.deterministic_brochure_link_for_building(text, "Riverside Building")

        self.assertNotEqual(url, "https://a.example.com/floorplans")

    def test_no_brochure_link_in_source_returns_none(self):
        text = extract_spreadsheet_gemini.render_sheet_as_text(self._sheet([
            [None, "Riverside Building - Southwark"],
            [None, "A modern glass-fronted office building overlooking the Thames."],
            [None, "15 Riverside Walk, SE1 9EZ", None, None, None, "Download Floorplans"],
            [None, "Office", "Sq.Ft", "Rent PCM", "Available From"],
            [None, "3rd Floor", 1800.0, 14000.0, "Now"],
        ]))

        url = extract_spreadsheet_gemini.deterministic_brochure_link_for_building(text, "Riverside Building")

        self.assertIsNone(url)

    def test_two_conflicting_brochure_links_returns_none(self):
        ws = self._sheet([])
        self._riverside_block(ws)
        ws.append([None, "Old Brochure (superseded)"])
        ws["B4"].hyperlink = "https://c.example.com/old-brochure.pdf"
        ws.append([None, "Office", "Sq.Ft", "Rent PCM", "Available From"])
        ws.append([None, "3rd Floor", 1800.0, 14000.0, "Now"])
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        url = extract_spreadsheet_gemini.deterministic_brochure_link_for_building(text, "Riverside Building")

        self.assertIsNone(url)

    def test_building_block_not_found_returns_none(self):
        url = extract_spreadsheet_gemini.deterministic_brochure_link_for_building(
            "Row 1: unrelated content", "Nowhere",
        )
        self.assertIsNone(url)


class DeterministicFloorplanLinkForBuildingTests(unittest.TestCase):
    """deterministic_floorplan_link_for_building - the exact mirror of
    DeterministicBrochureLinkForBuildingTests above, for a building's own
    "Download Floorplans"-labeled cell."""

    def _sheet(self, rows):
        wb = Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        return ws

    def _riverside_block(self, ws):
        ws.append([None, "Riverside Building - Southwark"])
        ws.append([None, "A modern glass-fronted office building overlooking the Thames."])
        ws.append([
            None, "15 Riverside Walk, SE1 9EZ", None, None, None,
            "Download Floorplans", None, None, "Download Brochure",
        ])
        ws["F3"].hyperlink = "https://a.example.com/floorplans"
        ws["I3"].hyperlink = "https://b.example.com/brochure.pdf"

    def test_recovers_the_single_unambiguous_floorplan_link(self):
        ws = self._sheet([])
        self._riverside_block(ws)
        ws.append([None, "Office", "Sq.Ft", "Rent PCM", "Available From"])
        ws.append([None, "3rd Floor", 1800.0, 14000.0, "Now"])
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        url = extract_spreadsheet_gemini.deterministic_floorplan_link_for_building(text, "Riverside Building")

        self.assertEqual(url, "https://a.example.com/floorplans")

    def test_never_returns_the_brochure_link(self):
        ws = self._sheet([])
        self._riverside_block(ws)
        ws.append([None, "Office", "Sq.Ft", "Rent PCM", "Available From"])
        ws.append([None, "3rd Floor", 1800.0, 14000.0, "Now"])
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        url = extract_spreadsheet_gemini.deterministic_floorplan_link_for_building(text, "Riverside Building")

        self.assertNotEqual(url, "https://b.example.com/brochure.pdf")

    def test_no_floorplan_link_in_source_returns_none(self):
        text = extract_spreadsheet_gemini.render_sheet_as_text(self._sheet([
            [None, "Riverside Building - Southwark"],
            [None, "A modern glass-fronted office building overlooking the Thames."],
            [None, "15 Riverside Walk, SE1 9EZ", None, None, None, "Download Brochure"],
            [None, "Office", "Sq.Ft", "Rent PCM", "Available From"],
            [None, "3rd Floor", 1800.0, 14000.0, "Now"],
        ]))

        url = extract_spreadsheet_gemini.deterministic_floorplan_link_for_building(text, "Riverside Building")

        self.assertIsNone(url)

    def test_building_block_not_found_returns_none(self):
        url = extract_spreadsheet_gemini.deterministic_floorplan_link_for_building(
            "Row 1: unrelated content", "Nowhere",
        )
        self.assertIsNone(url)


class ClassificationIsGenericNotProviderSpecificTests(unittest.TestCase):
    """deterministic_brochure_link_for_building/deterministic_floorplan_
    link_for_building take only raw sheet text and a building name - no
    provider argument exists anywhere in this classification path, so it
    cannot special-case a provider even in principle. Proven here with a
    deliberately fictional, generic building/provider/host - the exact
    same code path a real UNION or any other provider's file goes through."""

    def test_a_completely_fictional_generic_provider_classifies_correctly(self):
        wb = Workbook()
        ws = wb.active
        ws.append([None, "Zzyzx Business Park - Nowhereville"])
        ws.append([None, "A fictional test building, not a real provider or address."])
        ws.append([
            None, "1 Fictional Way, ZZ1 1ZZ", None, None, None,
            "Download Floorplans", None, None, "Download Brochure",
        ])
        ws["F3"].hyperlink = "https://totally-generic-host.example/s/floorplan-xyz"
        ws["I3"].hyperlink = "https://totally-generic-host.example/s/brochure-xyz"
        ws.append([None, "Office", "Sq.Ft", "Rent PCM", "Available From"])
        ws.append([None, "1st Floor", 1000.0, 5000.0, "Now"])
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        brochure_url = extract_spreadsheet_gemini.deterministic_brochure_link_for_building(
            text, "Zzyzx Business Park",
        )
        floorplan_url = extract_spreadsheet_gemini.deterministic_floorplan_link_for_building(
            text, "Zzyzx Business Park",
        )

        self.assertEqual(brochure_url, "https://totally-generic-host.example/s/brochure-xyz")
        self.assertEqual(floorplan_url, "https://totally-generic-host.example/s/floorplan-xyz")


class CombinedBrochureAndFloorplanCellTests(unittest.TestCase):
    """A single cell whose own text mentions BOTH "brochure" and
    "floorplan" (a real, common combined-document naming pattern, e.g.
    "Download Brochure and Floorplans") is one document that stays
    classified as a brochure only - it must never ALSO populate
    floorplan_link with that same URL, an unnecessary duplication."""

    def test_floorplan_cell_text_check_backs_off_when_brochure_is_also_present(self):
        self.assertIsNone(
            extract_spreadsheet_gemini._floorplan_url_from_cell_text(
                "Download Brochure and Floorplans (https://example.com/combined.pdf)",
            ),
        )

    def test_brochure_cell_text_check_is_unaffected(self):
        self.assertEqual(
            extract_spreadsheet_gemini._brochure_url_from_cell_text(
                "Download Brochure and Floorplans (https://example.com/combined.pdf)",
            ),
            "https://example.com/combined.pdf",
        )

    def test_a_genuinely_floorplan_only_cell_still_classifies_as_floorplan(self):
        self.assertEqual(
            extract_spreadsheet_gemini._floorplan_url_from_cell_text(
                "Download Floorplans (https://example.com/floorplan.pdf)",
            ),
            "https://example.com/floorplan.pdf",
        )

    def test_combined_cell_populates_only_brochure_link_end_to_end(self):
        wb = Workbook()
        ws = wb.active
        ws.append([None, "Riverside Building - Southwark"])
        ws.append([None, "A modern glass-fronted office building overlooking the Thames."])
        ws.append([None, "15 Riverside Walk, SE1 9EZ", None, None, None, "Download Brochure and Floorplans"])
        ws["F3"].hyperlink = "https://example.com/combined.pdf"
        ws.append([None, "Office", "Sq.Ft", "Rent PCM", "Available From"])
        ws.append([None, "3rd Floor", 1800.0, 14000.0, "Now"])
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        brochure_url = extract_spreadsheet_gemini.deterministic_brochure_link_for_building(text, "Riverside Building")
        floorplan_url = extract_spreadsheet_gemini.deterministic_floorplan_link_for_building(
            text, "Riverside Building",
        )

        self.assertEqual(brochure_url, "https://example.com/combined.pdf")
        self.assertIsNone(floorplan_url)


class ApplyDeterministicBrochureLinksEndToEndTests(unittest.TestCase):
    """Same real, confirmed shape as DeterministicBrochureLinkForBuildingTests,
    exercised through extract_sheet end to end (Gemini mocked) - proves the
    override actually reaches the final ListingRow.brochure_link, survives
    finalize_brochure_link's own downstream handling, and correctly leaves
    every OTHER case (multiple floors, no source link, conflicting links)
    exactly as Gemini itself extracted."""

    def _riverside_block(self, ws):
        ws.append([None, "Riverside Building - Southwark"])
        ws.append([None, "A modern glass-fronted office building overlooking the Thames."])
        ws.append([
            None, "15 Riverside Walk, SE1 9EZ", None, None, None,
            "Download Floorplans", None, None, "Download Brochure",
        ])
        ws["F3"].hyperlink = "https://a.example.com/floorplans"
        ws["I3"].hyperlink = "https://b.example.com/brochure.pdf"

    def _sheet_with_floors(self, floor_rows):
        wb = Workbook()
        ws = wb.active
        self._riverside_block(ws)
        ws.append([None, "Office", "Sq.Ft", "Rent PCM", "Available From"])
        for floor in floor_rows:
            ws.append([None, floor, 1800.0, 14000.0, "Now"])
        return ws

    def test_one_building_one_brochure_one_floor(self):
        raw = {
            "provider": "Copthall Estates", "contacts": None, "fully_occupied_buildings": [],
            "units": [
                {"building": "Riverside Building", "floor_unit": "3rd Floor",
                 "size_sqft": 1800, "rent_pcm": 14000, "brochure_link": None},
            ],
        }
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            rows = extract_spreadsheet_gemini.extract_sheet(
                self._sheet_with_floors(["3rd Floor"]), "file.xlsx — City", "file.xlsx"
            )

        self.assertEqual(rows[0].brochure_link, "https://b.example.com/brochure.pdf")

    def test_one_building_one_brochure_multiple_floors_all_inherit_it(self):
        raw = {
            "provider": "Copthall Estates", "contacts": None, "fully_occupied_buildings": [],
            "units": [
                {"building": "Riverside Building", "floor_unit": "3rd Floor",
                 "size_sqft": 1800, "rent_pcm": 14000, "brochure_link": None},
                {"building": "Riverside Building", "floor_unit": "4th Floor",
                 "size_sqft": 1900, "rent_pcm": 15000, "brochure_link": None},
                {"building": "Riverside Building", "floor_unit": "5th Floor",
                 "size_sqft": 2000, "rent_pcm": 16000, "brochure_link": None},
            ],
        }
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            rows = extract_spreadsheet_gemini.extract_sheet(
                self._sheet_with_floors(["3rd Floor", "4th Floor", "5th Floor"]), "file.xlsx — City", "file.xlsx"
            )

        self.assertEqual(len(rows), 3)
        for row in rows:
            self.assertEqual(row.brochure_link, "https://b.example.com/brochure.pdf")

    def test_gemini_returns_none_but_excel_has_explicit_brochure_link(self):
        # The real, confirmed failure mode this whole mechanism exists for:
        # Gemini drops brochure_link entirely despite the source having one.
        raw = {
            "provider": "Copthall Estates", "contacts": None, "fully_occupied_buildings": [],
            "units": [
                {"building": "Riverside Building", "floor_unit": "3rd Floor",
                 "size_sqft": 1800, "rent_pcm": 14000, "brochure_link": None},
            ],
        }
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            rows = extract_spreadsheet_gemini.extract_sheet(
                self._sheet_with_floors(["3rd Floor"]), "file.xlsx — City", "file.xlsx"
            )

        self.assertEqual(rows[0].brochure_link, "https://b.example.com/brochure.pdf")

    def test_gemini_returns_the_floorplan_url_but_excel_clearly_differs(self):
        # Gemini picked the wrong (Floorplans) link - the deterministic
        # value must win outright, not merely fill a gap.
        raw = {
            "provider": "Copthall Estates", "contacts": None, "fully_occupied_buildings": [],
            "units": [
                {"building": "Riverside Building", "floor_unit": "3rd Floor",
                 "size_sqft": 1800, "rent_pcm": 14000,
                 "brochure_link": "https://a.example.com/floorplans"},
            ],
        }
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            rows = extract_spreadsheet_gemini.extract_sheet(
                self._sheet_with_floors(["3rd Floor"]), "file.xlsx — City", "file.xlsx"
            )

        self.assertEqual(rows[0].brochure_link, "https://b.example.com/brochure.pdf")
        self.assertNotEqual(rows[0].brochure_link, "https://a.example.com/floorplans")

    def test_no_source_brochure_link_never_fabricates_one(self):
        wb = Workbook()
        ws = wb.active
        ws.append([None, "Riverside Building - Southwark"])
        ws.append([None, "A modern glass-fronted office building overlooking the Thames."])
        ws.append([None, "15 Riverside Walk, SE1 9EZ", None, None, None, "Download Floorplans"])
        ws["F3"].hyperlink = "https://a.example.com/floorplans"
        ws.append([None, "Office", "Sq.Ft", "Rent PCM", "Available From"])
        ws.append([None, "3rd Floor", 1800.0, 14000.0, "Now"])
        raw = {
            "provider": "Copthall Estates", "contacts": None, "fully_occupied_buildings": [],
            "units": [
                {"building": "Riverside Building", "floor_unit": "3rd Floor",
                 "size_sqft": 1800, "rent_pcm": 14000, "brochure_link": None},
            ],
        }
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            rows = extract_spreadsheet_gemini.extract_sheet(ws, "file.xlsx — City", "file.xlsx")

        self.assertIsNone(rows[0].brochure_link)

    def test_conflicting_brochure_links_are_never_guessed(self):
        wb = Workbook()
        ws = wb.active
        self._riverside_block(ws)
        ws.append([None, "Old Brochure (superseded)"])
        ws["B4"].hyperlink = "https://c.example.com/old-brochure.pdf"
        ws.append([None, "Office", "Sq.Ft", "Rent PCM", "Available From"])
        ws.append([None, "3rd Floor", 1800.0, 14000.0, "Now"])
        raw = {
            "provider": "Copthall Estates", "contacts": None, "fully_occupied_buildings": [],
            "units": [
                {"building": "Riverside Building", "floor_unit": "3rd Floor",
                 "size_sqft": 1800, "rent_pcm": 14000, "brochure_link": None},
            ],
        }
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            rows = extract_spreadsheet_gemini.extract_sheet(ws, "file.xlsx — City", "file.xlsx")

        # Never guessed at from the conflicting pair - left exactly as Gemini
        # (the best available signal here) extracted it, which in this case
        # was nothing.
        self.assertIsNone(rows[0].brochure_link)

    def test_does_not_replace_an_already_valid_link_with_a_wrong_neighbor(self):
        # A single-consistent-table sheet (shape (a)) has no "download" line
        # at all, so _building_block_bounds returns None and the override is
        # a complete no-op - Gemini's own extraction is untouched either way.
        wb = Workbook()
        ws = wb.active
        ws.append([None, "Building", "Floor", "Size (sq ft)", "Rent PCM"])
        ws.append([None, "Portfolio Building A", "1st Floor", 1000, 20000])
        raw = {
            "provider": "Some Provider", "contacts": None, "fully_occupied_buildings": [],
            "units": [
                {"building": "Portfolio Building A", "floor_unit": "1st Floor",
                 "size_sqft": 1000, "rent_pcm": 20000,
                 "brochure_link": "https://provider.example.com/brochure.pdf"},
            ],
        }
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            rows = extract_spreadsheet_gemini.extract_sheet(ws, "file.xlsx — Sheet1", "file.xlsx")

        self.assertEqual(rows[0].brochure_link, "https://provider.example.com/brochure.pdf")


class ExtractSheetWithMetadataTests(unittest.TestCase):
    """extract_sheet_with_metadata's second return value - the {"provider",
    "building"} dicts Gemini names as explicitly fully-occupied (see PROMPT's
    own fully_occupied_buildings field) - and extract_sheet's own backward-
    compatible wrapper around it."""

    def _sheet(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Office"
        ws["A2"] = "4th Floor"
        return ws

    def test_returns_fully_occupied_buildings_alongside_rows(self):
        raw = {
            "provider": "Copthall Estates", "contacts": None,
            "fully_occupied_buildings": ["50 Wells Street"],
            "units": [
                {"building": "28 Lime Street", "floor_unit": "4th Floor",
                 "size_sqft": 1358, "rent_pcm": 19805},
            ],
        }
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            rows, fully_occupied = extract_spreadsheet_gemini.extract_sheet_with_metadata(
                self._sheet(), "file.xlsx — City", "file.xlsx"
            )

        self.assertEqual(len(rows), 1)
        self.assertEqual(fully_occupied, [{"provider": "Copthall Estates", "building": "50 Wells Street"}])

    def test_no_fully_occupied_buildings_returns_empty_list(self):
        raw = {
            "provider": "Copthall Estates", "contacts": None, "fully_occupied_buildings": [],
            "units": [{"building": "28 Lime Street", "floor_unit": "4th Floor"}],
        }
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            _rows, fully_occupied = extract_spreadsheet_gemini.extract_sheet_with_metadata(
                self._sheet(), "file.xlsx — City", "file.xlsx"
            )

        self.assertEqual(fully_occupied, [])

    def test_missing_field_in_raw_response_defaults_to_empty_list(self):
        # Older/looser Gemini responses without this new field at all must
        # never crash - .get() with a default, not a required key.
        raw = {"provider": "Copthall Estates", "contacts": None, "units": []}
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            rows, fully_occupied = extract_spreadsheet_gemini.extract_sheet_with_metadata(
                self._sheet(), "file.xlsx — City", "file.xlsx"
            )

        self.assertEqual(rows, [])
        self.assertEqual(fully_occupied, [])

    def test_extract_sheet_wrapper_drops_the_metadata_and_returns_rows_only(self):
        raw = {
            "provider": "Copthall Estates", "contacts": None,
            "fully_occupied_buildings": ["Some Building"],
            "units": [{"building": "28 Lime Street", "floor_unit": "4th Floor"}],
        }
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            rows = extract_spreadsheet_gemini.extract_sheet(self._sheet(), "file.xlsx — City", "file.xlsx")

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].building, "28 Lime Street")


class IsNonAuthoritativeRollupSheetTests(unittest.TestCase):
    """is_non_authoritative_rollup_sheet - confirmed against the real
    Copthall Estates Availability.xlsx: its "Portfolio" sheet is a hidden,
    flat Area/Station/Building/Office/... table with zero real hyperlinks
    anywhere, roughly a year stale compared to its visible sibling sheets.
    Requires BOTH hidden-state AND "no download line anywhere" - see the
    function's own docstring for why either alone is unsafe."""

    def _sheet(self, rows, state="visible"):
        wb = Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        ws.sheet_state = state
        return ws

    def _portfolio_like_rows(self):
        return [
            [None, "Area", "Station", "Building", "Office", "Sq.Ft"],
            [None, "City", "Liverpool St", "50 Gresham", "3rd Floor", 973],
        ]

    def _riverside_block_rows(self):
        # A genuine shape (b) repeating block WITH a real "download" line -
        # what a genuinely authoritative sheet always has somewhere.
        return [
            [None, "Riverside Building - Southwark"],
            [None, "A modern glass-fronted office building overlooking the Thames."],
            [None, "15 Riverside Walk, SE1 9EZ", None, None, None, "Download Floorplans"],
            [None, "Office", "Sq.Ft", "Rent PCM", "Available From"],
            [None, "3rd Floor", 1800.0, 14000.0, "Now"],
        ]

    def test_hidden_with_no_download_lines_is_non_authoritative(self):
        # The real, confirmed Copthall Portfolio shape.
        ws = self._sheet(self._portfolio_like_rows(), state="hidden")
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        self.assertTrue(extract_spreadsheet_gemini.is_non_authoritative_rollup_sheet(ws, text))

    def test_very_hidden_with_no_download_lines_is_non_authoritative(self):
        ws = self._sheet(self._portfolio_like_rows(), state="veryHidden")
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        self.assertTrue(extract_spreadsheet_gemini.is_non_authoritative_rollup_sheet(ws, text))

    def test_visible_with_no_download_lines_is_not_flagged(self):
        # A different provider's genuinely current, but hyperlink-less,
        # free-form sheet - visible, so never touched by this rule at all.
        ws = self._sheet(self._portfolio_like_rows(), state="visible")
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        self.assertFalse(extract_spreadsheet_gemini.is_non_authoritative_rollup_sheet(ws, text))

    def test_hidden_but_with_real_download_lines_is_not_flagged(self):
        # A hidden sheet that DOES carry a genuine per-building hyperlinked
        # block must never be skipped just because it's hidden - "skip
        # every hidden sheet" is exactly the global rule this must avoid.
        ws = self._sheet(self._riverside_block_rows(), state="hidden")
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        self.assertFalse(extract_spreadsheet_gemini.is_non_authoritative_rollup_sheet(ws, text))

    def test_visible_with_real_download_lines_is_not_flagged(self):
        ws = self._sheet(self._riverside_block_rows(), state="visible")
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        self.assertFalse(extract_spreadsheet_gemini.is_non_authoritative_rollup_sheet(ws, text))


class ExtractUpdateDateTests(unittest.TestCase):
    def test_parses_four_digit_year(self):
        d = extract_spreadsheet_gemini.extract_update_date(
            "Row 1: Copthall Estates City Availability - Updated 03/08/2026"
        )
        self.assertEqual(d, date(2026, 8, 3))

    def test_parses_two_digit_year_as_20xx(self):
        d = extract_spreadsheet_gemini.extract_update_date(
            "Row 1: Copthall Estates Availibility - Updated 05/08/25"
        )
        self.assertEqual(d, date(2025, 8, 5))

    def test_no_updated_phrase_returns_none(self):
        self.assertIsNone(extract_spreadsheet_gemini.extract_update_date("Row 1: Some Provider Availability"))

    def test_invalid_date_is_ignored_not_raised(self):
        self.assertIsNone(extract_spreadsheet_gemini.extract_update_date("Row 1: Updated 32/13/2026"))

    def test_only_checks_the_first_few_lines(self):
        text = "\n".join([f"Row {i}: filler" for i in range(1, 20)] + ["Row 20: Updated 01/01/2020"])
        self.assertIsNone(extract_spreadsheet_gemini.extract_update_date(text))


class ClassifySheetForExtractionTests(unittest.TestCase):
    """classify_sheet_for_extraction - the three-outcome classification
    (auto_skip / ambiguous / authoritative) app.py's upload-time decision UI
    and Extract-time sheet loop both rely on. Confirmed against the real
    Copthall Estates Availability.xlsx's own real shapes throughout."""

    def _sheet(self, rows, title="Sheet1", state="visible"):
        wb = Workbook()
        ws = wb.active
        ws.title = title
        for row in rows:
            ws.append(row)
        ws.sheet_state = state
        return ws

    def _portfolio_rows(self):
        rows = [[None, "Copthall Estates Availibility - Updated 05/08/25"],
                [None, "Area", "Station", "Building", "Office", "Sq.Ft", "Price Per Sq.Ft"]]
        for i in range(10):
            rows.append([None, "City", "Bank", f"Building {i}", "3rd Floor", 1000 + i, 150])
        return rows

    def _city_block_rows(self, updated="03/08/2026"):
        return [
            [None, f"Copthall Estates City Availability - Updated {updated}"],
            [None, "28 Lime Street - Fenchurch St / Bank"],
            [None, "A striking City tower with amenities."],
            [None, "27-30 Lime Street, EC3M 7AF", None, None, None, "Download Floorplans"],
            [None, "Office", "Sq.Ft", "Rent PCM", "Available From"],
            [None, "5th Floor", 2400.0, 18000.0, "Now"],
        ]

    def _incentives_rows(self):
        return [
            [None, "Copthall Estates Incentives - Updated 03/08/26", None, None, None, None, None, None, None,
             None, None, "To register a lead or discuss an enquiry, contact:"],
            [None, "Standard Commission", None, None, None, None, None, None, None, None, None,
             "Enquiry@copthallestates.com  0203 002 2503"],
            [None, "10% on the first 12 months, 2% on months 13-24.", None, None, None, None, None, None, None,
             None, None, "Kiri Norton-Brennan - 0795 811 8382"],
            [None, "89 Charterhouse Street - Enhanced Incentive"],
        ]

    def test_hidden_and_no_download_lines_is_auto_skipped_unchanged(self):
        # The real, confirmed Copthall Portfolio shape - auto_skip must
        # survive exactly as it did before this classification existed.
        ws = self._sheet(self._portfolio_rows(), title="Portfolio", state="hidden")
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        result = extract_spreadsheet_gemini.classify_sheet_for_extraction(ws, text)

        self.assertEqual(result["outcome"], "auto_skip")

    def test_ordinary_authoritative_sheet_has_no_reasons(self):
        ws = self._sheet(self._city_block_rows(), title="City", state="visible")
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        result = extract_spreadsheet_gemini.classify_sheet_for_extraction(ws, text)

        self.assertEqual(result["outcome"], "authoritative")
        self.assertEqual(result["reasons"], [])

    def test_incentives_shaped_sheet_is_authoritative_not_ambiguous(self):
        # Must never be flagged just for lacking hyperlinks - a flat table
        # with no download/brochure line carries no staleness signal on its
        # own (see classify_sheet_for_extraction's own docstring).
        ws = self._sheet(self._incentives_rows(), title="Incentives", state="visible")
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        result = extract_spreadsheet_gemini.classify_sheet_for_extraction(ws, text)

        self.assertEqual(result["outcome"], "authoritative")

    def test_visible_portfolio_named_flat_table_is_ambiguous(self):
        ws = self._sheet(self._portfolio_rows(), title="Portfolio", state="visible")
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        result = extract_spreadsheet_gemini.classify_sheet_for_extraction(ws, text)

        self.assertEqual(result["outcome"], "ambiguous")
        self.assertIn("suspicious_sheet_name", result["reasons"])

    def test_hidden_sheet_with_real_download_lines_is_ambiguous_not_authoritative(self):
        ws = self._sheet(self._city_block_rows(), title="City", state="hidden")
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        result = extract_spreadsheet_gemini.classify_sheet_for_extraction(ws, text)

        self.assertEqual(result["outcome"], "ambiguous")
        self.assertIn("hidden", result["reasons"])

    def test_suspicious_name_alone_never_reaches_auto_skip(self):
        ws = self._sheet(self._city_block_rows(), title="Summary", state="visible")
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        result = extract_spreadsheet_gemini.classify_sheet_for_extraction(ws, text)

        self.assertNotEqual(result["outcome"], "auto_skip")

    def test_hidden_state_alone_never_reaches_auto_skip(self):
        ws = self._sheet(self._city_block_rows(), title="City", state="hidden")
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        result = extract_spreadsheet_gemini.classify_sheet_for_extraction(ws, text)

        self.assertNotEqual(result["outcome"], "auto_skip")

    def test_no_download_links_alone_never_reaches_auto_skip(self):
        # Visible + no download lines - still must never auto-skip without
        # the hidden/veryHidden half of the confident pair (is_non_
        # authoritative_rollup_sheet's own two-signal intersection).
        ws = self._sheet(self._portfolio_rows(), title="Portfolio", state="visible")
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        result = extract_spreadsheet_gemini.classify_sheet_for_extraction(ws, text)

        self.assertNotEqual(result["outcome"], "auto_skip")

    def test_stale_date_versus_siblings_is_ambiguous(self):
        ws = self._sheet(self._city_block_rows(updated="05/08/25"), title="City", state="visible")
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        result = extract_spreadsheet_gemini.classify_sheet_for_extraction(
            ws, text, sibling_dates={"Mid Town": date(2026, 8, 3)},
        )

        self.assertEqual(result["outcome"], "ambiguous")
        self.assertIn("stale_update_date_vs_siblings", result["reasons"])

    def test_small_date_gap_versus_siblings_is_not_flagged(self):
        # Real Blackfriars vs City gap (~1 month) - a normal per-area
        # update cadence difference, never staleness.
        ws = self._sheet(self._city_block_rows(updated="01/07/2026"), title="Blackfriars", state="visible")
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        result = extract_spreadsheet_gemini.classify_sheet_for_extraction(
            ws, text, sibling_dates={"City": date(2026, 8, 3)},
        )

        self.assertEqual(result["outcome"], "authoritative")

    def test_no_sibling_dates_never_guesses_staleness(self):
        ws = self._sheet(self._city_block_rows(updated="05/08/25"), title="City", state="visible")
        text = extract_spreadsheet_gemini.render_sheet_as_text(ws)

        result = extract_spreadsheet_gemini.classify_sheet_for_extraction(ws, text, sibling_dates={})

        self.assertEqual(result["outcome"], "authoritative")


if __name__ == "__main__":
    unittest.main()
