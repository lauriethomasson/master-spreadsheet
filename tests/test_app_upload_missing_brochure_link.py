"""
Regression test for app.py's _warn_if_brochure_link_missing AND for
extract_spreadsheet_gemini's deterministic Excel-hyperlink recovery that now
runs ahead of it (see _apply_deterministic_brochure_links) - a real,
confirmed production case: every building checked in the real Copthall
Estates Availability file has a genuine, working "Download Brochure" link
in the source sheet, so a building whose own block clearly has one but
whose extracted unit(s) came back with no brochure_link at all is now
usually SILENTLY RECOVERED straight from the source hyperlink rather than
merely warned about - the warning is now only the remaining safety net for
the genuinely unresolvable case (a real conflict between two different
"Brochure"-labeled links in the same block, deliberately never guessed).

Runs the real app.py end-to-end via Streamlit's AppTest, with extract_
spreadsheet_gemini's Gemini call mocked to reproduce that exact failure -
never calls the real API (same principle as test_app_upload_rescue_removal.py).

Run with:
    .venv\\Scripts\\python.exe -m unittest tests.test_app_upload_missing_brochure_link -v
"""

import sys
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook
from streamlit.testing.v1 import AppTest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from storage.file_store import list_pending_staging_files

BASE = Path(__file__).resolve().parent.parent


def _clear_pending():
    for p in list_pending_staging_files():
        (BASE / p).unlink(missing_ok=True)
        (BASE / p).with_suffix(".meta.json").unlink(missing_ok=True)


def _build_riverside_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append([None, "Riverside Building - Southwark"])
    ws.append([None, "A modern glass-fronted office building overlooking the Thames."])
    ws.append([
        None, "15 Riverside Walk, SE1 9EZ", None, None, None,
        "Download Floorplans (https://a.example.com/floorplans)", None, None,
        "Download Brochure (https://b.example.com/brochure.pdf)",
    ])
    ws.append([None, "Office", "Sq.Ft", "Rent PCM", "Available From"])
    ws.append([None, "3rd Floor", 1800.0, 14000.0, "Now"])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_riverside_xlsx_with_conflicting_brochure_links() -> bytes:
    # Two DIFFERENT "Brochure"-labeled hyperlinks in the same building's own
    # block - genuinely ambiguous, the one shape deterministic recovery must
    # never guess between (see extract_spreadsheet_gemini.deterministic_
    # brochure_link_for_building). Synthetic (not seen verbatim in a real
    # Copthall file) but exercises the exact safety rule the real file's
    # single-link-per-building shape never itself triggers.
    wb = Workbook()
    ws = wb.active
    ws.append([None, "Riverside Building - Southwark"])
    ws.append([None, "A modern glass-fronted office building overlooking the Thames."])
    ws.append([
        None, "15 Riverside Walk, SE1 9EZ", None, None, None,
        "Download Floorplans", None, None, "Download Brochure",
    ])
    ws["F3"].hyperlink = "https://a.example.com/floorplans"
    ws["I3"].hyperlink = "https://b.example.com/brochure.pdf"
    # A second, DIFFERENT real hyperlink whose own display text also mentions
    # "Brochure" - the genuine conflict (see _brochure_urls_in_block) that
    # must never be guessed between.
    ws.append([None, "Old Brochure (superseded)"])
    ws["B4"].hyperlink = "https://c.example.com/old-brochure.pdf"
    ws.append([None, "Office", "Sq.Ft", "Rent PCM", "Available From"])
    ws.append([None, "3rd Floor", 1800.0, 14000.0, "Now"])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class MissingBrochureLinkWarningTests(unittest.TestCase):
    def setUp(self):
        _clear_pending()

    def tearDown(self):
        _clear_pending()

    def test_deterministic_recovery_silently_fills_a_gemini_dropped_link_no_warning(self):
        raw = {
            "provider": "Copthall Estates",
            "contacts": None,
            "units": [
                {
                    "building": "Riverside Building", "floor_unit": "3rd Floor",
                    "size_sqft": 1800, "rent_pcm": 14000, "brochure_link": None,
                },
            ],
        }
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()
            self.assertFalse(at.exception)

            at.file_uploader[0].upload(
                "riverside.xlsx", _build_riverside_xlsx(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            at.run()

            extract_buttons = [b for b in at.button if b.label == "Extract"]
            extract_buttons[0].click().run()
            self.assertFalse(at.exception)

        warnings = "".join(w.value for w in at.warning)
        self.assertNotIn("Download Brochure link", warnings)

    def test_conflicting_brochure_links_are_never_guessed_and_still_warn(self):
        raw = {
            "provider": "Copthall Estates",
            "contacts": None,
            "units": [
                {
                    "building": "Riverside Building", "floor_unit": "3rd Floor",
                    "size_sqft": 1800, "rent_pcm": 14000, "brochure_link": None,
                },
            ],
        }
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()
            self.assertFalse(at.exception)

            at.file_uploader[0].upload(
                "riverside.xlsx", _build_riverside_xlsx_with_conflicting_brochure_links(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            at.run()

            extract_buttons = [b for b in at.button if b.label == "Extract"]
            extract_buttons[0].click().run()
            self.assertFalse(at.exception)

        warnings = "".join(w.value for w in at.warning)
        self.assertIn("Riverside Building", warnings)
        self.assertIn("Download Brochure link", warnings)

    def test_present_brochure_link_shows_no_warning(self):
        raw = {
            "provider": "Copthall Estates",
            "contacts": None,
            "units": [
                {
                    "building": "Riverside Building", "floor_unit": "3rd Floor",
                    "size_sqft": 1800, "rent_pcm": 14000,
                    "brochure_link": "https://b.example.com/brochure.pdf",
                },
            ],
        }
        with patch("extract_spreadsheet_gemini.get_client"), \
             patch("extract_spreadsheet_gemini.call_gemini", return_value=raw):
            at = AppTest.from_file(str(BASE / "app.py"), default_timeout=30)
            at.run()

            at.file_uploader[0].upload(
                "riverside.xlsx", _build_riverside_xlsx(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            at.run()

            extract_buttons = [b for b in at.button if b.label == "Extract"]
            extract_buttons[0].click().run()
            self.assertFalse(at.exception)

        warnings = "".join(w.value for w in at.warning)
        self.assertNotIn("Download Brochure link", warnings)


if __name__ == "__main__":
    unittest.main()
