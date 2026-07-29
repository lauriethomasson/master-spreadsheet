"""
storage/file_store.py

Manages staging .xlsx files created by the Upload page and consumed/edited
by the Review page. Each staging upload gets its own file plus a sidecar
.meta.json tracking {filename, timestamp, status, n_rows}, so multiple
pending uploads can coexist before any of them is approved.
"""

import json
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font

from schema import ListingRow
from staging_writer import write_rows_to_xlsx

STAGING_DIR = Path("staging")


def _meta_path(xlsx_path: Path) -> Path:
    return xlsx_path.with_suffix(".meta.json")


def _read_meta(xlsx_path: Path) -> dict:
    with open(_meta_path(xlsx_path), encoding="utf-8") as f:
        return json.load(f)


def _write_meta(xlsx_path: Path, meta: dict) -> None:
    with open(_meta_path(xlsx_path), "w", encoding="utf-8") as f:
        json.dump(meta, f, indent=2)


def save_staging_file(rows: list[ListingRow], original_filename: str) -> str:
    STAGING_DIR.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    stem = Path(original_filename).stem
    staging_path = STAGING_DIR / f"{timestamp}_{stem}.xlsx"

    write_rows_to_xlsx(rows, staging_path)
    _write_meta(
        staging_path,
        {
            "filename": original_filename,
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "status": "pending_review",
            "n_rows": len(rows),
        },
    )
    return str(staging_path)


def _staging_signature() -> tuple:
    """One (name, mtime) pair per sidecar file — changes whenever a file is
    added, removed, or edited in place (e.g. mark_as_approved rewriting an
    existing meta.json), so it's a reliable cache key for the directory's
    current state without needing an explicit .clear() on every write path.
    Every upload/approval produces a brand new signature that's never looked
    up again, so the cached functions below bound entries/ttl to keep that
    unreachable history from growing without limit over a long-running process.
    """
    return tuple(sorted((p.name, p.stat().st_mtime) for p in STAGING_DIR.glob("*.meta.json")))


def list_pending_staging_files() -> list[str]:
    if not STAGING_DIR.exists():
        return []
    return _list_pending_staging_files_cached(_staging_signature())


@st.cache_data(max_entries=4, ttl=3600)
def _list_pending_staging_files_cached(signature: tuple) -> list[str]:
    pending = []
    for xlsx_path in STAGING_DIR.glob("*.xlsx"):
        try:
            meta = _read_meta(xlsx_path)
        except FileNotFoundError:
            continue
        if meta.get("status") == "pending_review":
            pending.append((meta.get("timestamp", ""), str(xlsx_path)))

    pending.sort(reverse=True)
    return [path for _, path in pending]


def load_staging_as_dataframe(path: str) -> pd.DataFrame:
    return _load_staging_as_dataframe_cached(path, Path(path).stat().st_mtime)


@st.cache_data(max_entries=8, ttl=3600)
def _load_staging_as_dataframe_cached(path: str, mtime: float) -> pd.DataFrame:
    return pd.read_excel(path)


def save_staging_dataframe(path: str, df: pd.DataFrame) -> None:
    df.to_excel(path, index=False)
    wb = load_workbook(path)
    ws = wb.active
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    wb.save(path)


def mark_as_approved(path: str) -> None:
    meta = _read_meta(Path(path))
    meta["status"] = "approved"
    _write_meta(Path(path), meta)


def _clean_value(value):
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        return value.item()
    return value


def dataframe_to_listing_rows(df: pd.DataFrame) -> list[ListingRow]:
    rows = []
    for record in df.to_dict(orient="records"):
        cleaned = {key: _clean_value(value) for key, value in record.items()}
        rows.append(ListingRow(**cleaned))
    return rows
