"""
master_writer.py

Owns every write to the cumulative master spreadsheet, and the versioned-
backup/restore mechanism that rides along with every such write.

The write itself stays a full atomic rewrite (build the complete desired
xlsx in memory, one blob_store.write_bytes() call), never a cell-level
patch of the existing file - master_merge.py computes the complete new
content (existing master rows, minus the fields a user approved changing,
plus brand-new rows) before this module ever touches storage. That keeps
the exact same crash-safety guarantees the original wipe-and-replace
design had, just fed a cumulative row list instead of one batch's rows.

Every successful write to master.xlsx - whether from a normal approval or
a restore - also writes a timestamped snapshot to versions/, then prunes
versions/ down to the most recent MAX_VERSIONS. Versioning is best-effort
bookkeeping riding on top of an already-successful live write, not part of
the same all-or-nothing operation: if the live master.xlsx write succeeds,
that's success as far as the caller/log/UI are concerned, even if the
version snapshot or pruning step then has a problem - the live data never
gets held hostage by that secondary bookkeeping. See _snapshot_version.
"""

import json
import sys
from datetime import datetime, timezone
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from schema import ListingRow
from staging_writer import read_xlsx_with_hyperlinks, write_rows_to_xlsx
from storage import blob_store

DEFAULT_MASTER_PATH = "data/master.xlsx"
LOG_PATH = "data/master_write_log.jsonl"
VERSIONS_PREFIX = "versions"
MAX_VERSIONS = 30


def master_exists(master_path: str = DEFAULT_MASTER_PATH) -> bool:
    return blob_store.exists(master_path)


def load_master_as_dataframe(master_path: str = DEFAULT_MASTER_PATH) -> pd.DataFrame:
    # mtime is part of the cache key (not an underscore-prefixed arg), so a
    # fresh write_master() — which always changes the file's mtime — is
    # enough to invalidate this on its own, with no explicit .clear() needed.
    # A superseded (path, mtime) entry is never looked up again, so max_entries/ttl
    # bound how much of that unreachable history st.cache_data keeps around.
    return _load_master_as_dataframe_cached(master_path, blob_store.get_mtime(master_path))


@st.cache_data(max_entries=4, ttl=3600)
def _load_master_as_dataframe_cached(master_path: str, mtime: float) -> pd.DataFrame:
    return read_xlsx_with_hyperlinks(blob_store.read_bytes(master_path))


def get_master_write_log(log_path: str = LOG_PATH) -> list:
    if not blob_store.exists(log_path):
        return []
    return _get_master_write_log_cached(log_path, blob_store.get_mtime(log_path))


@st.cache_data(max_entries=4, ttl=3600)
def _get_master_write_log_cached(log_path: str, mtime: float) -> list:
    text = blob_store.read_bytes(log_path).decode("utf-8")
    return [json.loads(line) for line in text.splitlines() if line.strip()]


def log_master_write(
    success: bool,
    row_count: int = None,
    error: str = None,
    source: str = "approve",
    new_count: int = None,
    updated_count: int = None,
    version_path: str = None,
    timestamp: str = None,
    fields_changed: int = None,
    removed_count: int = None,
):
    entry = {
        "timestamp": timestamp or datetime.now(timezone.utc).isoformat(),
        "success": success,
        "row_count": row_count,
        "error": error,
        "source": source,
        "new_count": new_count,
        "updated_count": updated_count,
        "version_path": version_path,
        "fields_changed": fields_changed,
        "removed_count": removed_count,
    }
    blob_store.append_text(LOG_PATH, json.dumps(entry) + "\n")
    print(f"[master_writer] {entry}", file=sys.stderr)


def _version_path(stamp: str) -> str:
    return f"{VERSIONS_PREFIX}/master_{stamp}.xlsx"


def _snapshot_version(data: bytes, stamp: str):
    """
    Best-effort: writes a timestamped copy of `data` to versions/ and prunes
    old ones beyond MAX_VERSIONS. Never raises - a hiccup here must not turn
    an already-successful live master.xlsx write into a reported failure.
    Returns the version's path on success, None if the snapshot itself
    failed (a pruning failure is swallowed too, but doesn't change the
    return value - the snapshot existing is what matters for restore).
    """
    path = _version_path(stamp)
    try:
        blob_store.write_bytes(path, data)
    except Exception as e:
        print(f"[master_writer] WARNING: version snapshot failed: {e}", file=sys.stderr)
        return None

    try:
        versions = sorted(blob_store.list_with_mtimes(VERSIONS_PREFIX, ".xlsx"), key=lambda pair: pair[0])
        excess = len(versions) - MAX_VERSIONS
        if excess > 0:
            for old_path, _mtime in versions[:excess]:
                blob_store.delete(old_path)
    except Exception as e:
        print(f"[master_writer] WARNING: version pruning failed: {e}", file=sys.stderr)

    return path


def list_versions(limit: int = None) -> list:
    """
    Most-recent-first list of {path, timestamp, label} for the version
    history UI, joined against master_write_log.jsonl (matched by the
    version_path shared between a write's log entry and its version
    snapshot) for a human-readable change summary when available.
    """
    versions = sorted(blob_store.list_with_mtimes(VERSIONS_PREFIX, ".xlsx"), key=lambda pair: pair[0], reverse=True)
    if limit:
        versions = versions[:limit]

    log_by_version_path = {}
    for entry in get_master_write_log():
        vp = entry.get("version_path")
        if vp:
            log_by_version_path[vp] = entry

    result = []
    for path, _mtime in versions:
        entry = log_by_version_path.get(path)
        if entry is None:
            result.append({"path": path, "timestamp": None, "label": "—"})
            continue

        source = entry.get("source", "approve")
        if source.startswith("restore:"):
            label = f"Restored from {source.split(':', 1)[1]}"
        elif source == "manual_edit":
            n = entry.get("fields_changed") or 0
            label = f"Manual edit: {n} field{'s' if n != 1 else ''} changed"
        else:
            new_c, upd_c, rem_c = entry.get("new_count"), entry.get("updated_count"), entry.get("removed_count")
            if new_c is not None or upd_c is not None or rem_c is not None:
                label = f"{upd_c or 0} updated, {new_c or 0} new"
                if rem_c:
                    label += f", {rem_c} removed"
            else:
                label = "Approved upload"
        result.append({"path": path, "timestamp": entry["timestamp"], "label": label})
    return result


def write_master(
    approved_rows: list[ListingRow],
    master_path: str = DEFAULT_MASTER_PATH,
    new_count: int = None,
    updated_count: int = None,
    source: str = "approve",
    fields_changed: int = None,
    removed_count: int = None,
):
    # Grouped by provider regardless of what order the merge upstream
    # produced (matched rows keep their prior on-disk position, new rows
    # are appended at the end) - sorting here, at the single write choke
    # point, guarantees the invariant holds after every write rather than
    # depending on every call site remembering to sort. Stable and
    # case-insensitive; rows with no provider sort last.
    approved_rows = sorted(
        approved_rows, key=lambda r: (r.provider is None, (r.provider or "").lower())
    )
    try:
        buffer = BytesIO()
        write_rows_to_xlsx(approved_rows, buffer)

        buffer.seek(0)
        wb = load_workbook(buffer)
        ws = wb.active
        actual_row_count = ws.max_row - 1
        if actual_row_count != len(approved_rows):
            raise ValueError(
                f"Validation failed: expected {len(approved_rows)} rows, "
                f"got {actual_row_count} in written file"
            )

        data = buffer.getvalue()
        blob_store.write_bytes(master_path, data)
    except Exception as e:
        log_master_write(success=False, error=str(e), source=source)
        raise

    # Milliseconds, not just seconds - two writes landing in the same second
    # (a fast successive approve+restore, or just an unlucky click) would
    # otherwise produce the same version filename and silently overwrite
    # each other's snapshot instead of keeping both.
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H-%M-%S-%f")[:-3]
    version_path = _snapshot_version(data, stamp)
    log_master_write(
        success=True,
        row_count=len(approved_rows),
        source=source,
        new_count=new_count,
        updated_count=updated_count,
        version_path=version_path,
        timestamp=stamp,
        fields_changed=fields_changed,
        removed_count=removed_count,
    )


def restore_version(version_path: str, master_path: str = DEFAULT_MASTER_PATH):
    """
    Restoring is just another write, not a separate code path: it replaces
    the live master.xlsx with a historical version's exact bytes (no
    re-serialization through ListingRow/write_rows_to_xlsx, which would risk
    subtly altering a file that was already valid) and goes through the same
    versioning/pruning tail as write_master(), so restoring itself produces
    a brand new version - nothing is ever a one-way trip.
    """
    try:
        data = blob_store.read_bytes(version_path)
        wb = load_workbook(BytesIO(data))  # sanity check it's a readable workbook before promoting it to live
        row_count = wb.active.max_row - 1
        blob_store.write_bytes(master_path, data)
    except Exception as e:
        log_master_write(success=False, error=str(e), source=f"restore:{version_path}")
        raise

    # Milliseconds, not just seconds - two writes landing in the same second
    # (a fast successive approve+restore, or just an unlucky click) would
    # otherwise produce the same version filename and silently overwrite
    # each other's snapshot instead of keeping both.
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H-%M-%S-%f")[:-3]
    new_version_path = _snapshot_version(data, stamp)
    log_master_write(
        success=True,
        row_count=row_count,
        source=f"restore:{version_path}",
        version_path=new_version_path,
        timestamp=stamp,
    )


def main():
    if len(sys.argv) not in (2, 3):
        print("Usage: python master_writer.py <rows.json> [master_path]", file=sys.stderr)
        raise SystemExit(1)

    with open(sys.argv[1], encoding="utf-8") as f:
        raw_rows = json.load(f)

    rows = [ListingRow(**r) for r in raw_rows]
    master_path = sys.argv[2] if len(sys.argv) == 3 else DEFAULT_MASTER_PATH
    write_master(rows, master_path)
    print(f"Wrote {len(rows)} row(s) to {master_path}")


if __name__ == "__main__":
    main()
