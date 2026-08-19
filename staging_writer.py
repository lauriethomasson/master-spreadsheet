import json
import sys
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from schema import ListingRow

MIN_COLUMN_WIDTH = 10
MAX_COLUMN_WIDTH = 40

# Columns holding a URL - given a real openpyxl hyperlink (not just plain text)
# so the cell is directly clickable when the .xlsx is opened in Excel/Google
# Sheets, not just styled to look like a link. The cell's displayed text is
# its own LINK_DISPLAY_TEXT entry, not the (often long, unreadable) raw URL -
# the real URL only ever lives in the hyperlink target from that point on,
# which is exactly why every read-back of a written file (read_xlsx_with_
# hyperlinks below) has to recover it from there instead of the cell's plain
# value.
HYPERLINK_COLUMNS = ["brochure_link", "floorplan_link"]
HYPERLINK_FONT = Font(color="0563C1", underline="single")
HYPERLINK_DISPLAY_TEXT = "Open brochure"  # kept for brochure_link/back-compat single-column callers
LINK_DISPLAY_TEXT = {"brochure_link": "Open brochure", "floorplan_link": "Open floor plan"}

# brochure_link specifically (never floorplan_link - ListingRow.brochure_
# link_broken's own docstring scopes it to the brochure link only), for a
# row whose last known render attempt CONFIRMED the link dead (see that
# field's own docstring for exactly what qualifies). The cell keeps its
# real hyperlink target underneath (Option A of the two considered: the
# alternative - dropping the hyperlink entirely - would make read_xlsx_
# with_hyperlinks below read this cell's own display text, "Broken link",
# back as brochure_link's actual value on the next load, corrupting a
# real data field with UI text) - only the DISPLAYED text/font changes,
# so read_xlsx_with_hyperlinks needs no change at all: it already
# recovers the real URL from cell.hyperlink.target regardless of what
# the cell displays. The cell is technically still one click away from
# the dead link (Excel has no way to attach a hyperlink yet disable
# clicking it) - a deliberate, accepted tradeoff for never losing the
# underlying URL.
BROKEN_LINK_DISPLAY_TEXT = "Broken link"
BROKEN_LINK_FONT = Font(color="808080", italic=True)  # grey, italic, no underline - reads as inert, not urgent

# Columns that regularly hold long text (a multi-sentence description, several
# contacts, or a long URL) - wrapped so the full value is visible on several
# lines within the row instead of overflowing past the column's fixed width or
# being clipped. Deliberately a separate list from display_utils.WIDE_TEXT_COLUMNS:
# that one governs the in-app editable grid's column width and never includes
# brochure_link/floorplan_link (the grid shows a fixed "Open ..." label there
# regardless of the underlying URL's length), whereas here they're exactly the
# columns most likely to overflow a written .xlsx's column width.
WRAP_COLUMNS = ["special_features", "contacts", "brochure_link", "floorplan_link"]
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

# Points, matching Excel's own row-height unit - tall enough for ~3-4 wrapped
# lines. openpyxl can't compute a true auto-fit height (that's an application-
# side text-layout calculation, not derivable from raw XML), and row height is
# a per-row property in Excel - a row can't be tall for one column's wrapped
# text and short for the rest - so every data row gets this same fixed height.
DATA_ROW_HEIGHT = 60

# Present in every row for traceability (geocode failure logs, the
# brochure_link PDF-fallback default) but not something Mark/Laurie need to
# see day-to-day - hidden as an Excel column rather than dropped from the
# file, so it's still there to unhide/read directly (e.g. via pandas) if a
# listing ever needs tracing back to its source upload. property_id is
# master_merge.py's internal identity for a property across uploads - never
# meaningful to a person, but needed in the file so re-loading the master
# preserves it. brochure_link_broken is pipeline diagnostics (see that
# field's own schema.py docstring and build_merge_plan's own routing of it
# into silent_updates, never the reviewable diff) - a raw True/False/blank
# column would otherwise expose that internal bookkeeping directly next to
# brochure_link, which is exactly what BROKEN_LINK_DISPLAY_TEXT's own label
# swap is meant to communicate instead. brochure_link_is_floorplan is the
# same idea for a different fact (see its own schema.py docstring) - its
# own label swap below communicates it instead of a raw column. floorplan_
# link is a pure visibility change, not a data/logic one - the field, its
# data, and its own enrichment (FLOORPLAN_PROMPT etc. in brochure_
# enrichment.py) are all completely untouched, still written into this
# file exactly as before and still hyperlinked (see HYPERLINK_COLUMNS/
# LINK_DISPLAY_TEXT below, both unaffected by this) - it just no longer
# shows as its own visible column.
HIDDEN_COLUMNS = [
    "source_file", "property_id", "brochure_link_broken", "brochure_link_is_floorplan", "floorplan_link",
]


def title_case_label(field_name: str) -> str:
    """Display-only label for a snake_case field name - "internal_ref" ->
    "Internal Ref", "rent_pcm" -> "Rent Pcm" - used for the written
    .xlsx header row here and for on-screen column_config labels
    (display_utils.py). Purely cosmetic: the underlying field/column name
    (what schema.py, matching logic, etc. all reference) never changes -
    see _label_to_field_name/read_xlsx_with_hyperlinks, which recovers the
    real name from this text on every read rather than assuming a fixed
    position, for exactly that reason."""
    return field_name.replace("_", " ").title()


def _label_to_field_name(label: str) -> str:
    """Exact inverse of title_case_label - "Internal Ref" -> "internal_ref".
    Lossless for every real field name (all pure lowercase_with_underscores):
    title_case_label only replaces "_" with " " and capitalizes each word's
    first letter, so lowercasing and putting the underscores back recovers
    the original exactly, regardless of whether that field still exists in
    the CURRENT schema - see read_xlsx_with_hyperlinks."""
    return label.lower().replace(" ", "_")


def write_rows_to_xlsx(rows: list[ListingRow], output) -> None:
    """
    output is a filesystem path (str/Path) OR a writable binary stream (e.g.
    io.BytesIO) - openpyxl's Workbook.save() accepts either natively. A
    stream is what master_writer.py/storage/file_store.py pass so the result
    can go straight to blob_store (local disk or a GCS upload) without ever
    touching a local temp file.
    """
    fields = list(ListingRow.model_fields.keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "Listings"

    ws.append([title_case_label(f) for f in fields])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for row in rows:
        data = row.model_dump()
        ws.append([data[field] for field in fields])

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = DATA_ROW_HEIGHT

    for col_idx, field in enumerate(fields, start=1):
        max_len = len(field)
        for row in rows:
            value = getattr(row, field)
            if value is not None:
                max_len = max(max_len, len(str(value)))
        width = min(max(max_len + 2, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH)
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = width
        if field in HIDDEN_COLUMNS:
            ws.column_dimensions[column_letter].hidden = True

        if field in WRAP_COLUMNS:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = WRAP_ALIGNMENT

        if field in HYPERLINK_COLUMNS:
            display_text = LINK_DISPLAY_TEXT.get(field, HYPERLINK_DISPLAY_TEXT)
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    url = cell.value
                    cell.hyperlink = url
                    if field == "brochure_link" and rows[row_idx - 2].brochure_link_broken:
                        cell.value = BROKEN_LINK_DISPLAY_TEXT
                        cell.font = BROKEN_LINK_FONT
                    elif field == "brochure_link" and rows[row_idx - 2].brochure_link_is_floorplan:
                        cell.value = LINK_DISPLAY_TEXT["floorplan_link"]
                        cell.font = HYPERLINK_FONT
                    else:
                        cell.value = display_text
                        cell.font = HYPERLINK_FONT

    if isinstance(output, (str, Path)):
        Path(output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def read_xlsx_with_hyperlinks(data: bytes) -> pd.DataFrame:
    """
    Reads a workbook written by write_rows_to_xlsx() back into a DataFrame -
    like pd.read_excel(), except for HYPERLINK_COLUMNS, where the cell's
    displayed text is its own LINK_DISPLAY_TEXT entry rather than the real
    URL (see write_rows_to_xlsx). pd.read_excel only ever sees that
    displayed text, which would silently replace every brochure_link with
    the literal string "Open brochure" on every read-back after a write -
    reads the cell's actual hyperlink target instead for those columns;
    every other column is read exactly as pd.read_excel would.

    Column names are recovered from row 1's own cell text via
    _label_to_field_name, NOT assumed from ListingRow.model_fields by fixed
    position. Position-based reading was tried and reverted: it silently
    breaks the moment the on-disk column order doesn't exactly match the
    CURRENT schema's order - which real files hit in two different ways
    (confirmed against the actual data/master.xlsx): a field inserted
    mid-schema after older files were already written (property_id sits
    right after source_file today, but older files predate it entirely, so
    every later column is off by one), and a field removed from the
    schema's MIDDLE rather than its end (this module's own history - the
    size_sqft_min/max/rent_psf_min/max/rent_pcm_min/max fields used to sit
    between rent_psf and brochure_link, not at the tail). Reversing each
    header cell's own text instead is immune to both: every column's
    identity is read from itself, independent of position, so an old file's
    different order/extra now-removed columns are simply read correctly (or,
    for a field no longer in ListingRow, read as a same-named dict key that
    ListingRow(**cleaned)'s default extra="ignore" then drops harmlessly -
    see storage/file_store.dataframe_to_listing_rows).
    """
    wb = load_workbook(BytesIO(data))
    ws = wb.active
    headers = [_label_to_field_name(cell.value) for cell in ws[1]]
    hyperlink_col_indices = {i for i, h in enumerate(headers) if h in HYPERLINK_COLUMNS}

    records = []
    for row in ws.iter_rows(min_row=2):
        record = {}
        for col_idx, cell in enumerate(row):
            header = headers[col_idx]
            if col_idx in hyperlink_col_indices and cell.hyperlink is not None:
                record[header] = cell.hyperlink.target
            else:
                record[header] = cell.value
        records.append(record)
    return pd.DataFrame(records, columns=headers)


def main():
    if len(sys.argv) != 3:
        print("Usage: python staging_writer.py <rows.json> <output.xlsx>", file=sys.stderr)
        raise SystemExit(1)

    with open(sys.argv[1], encoding="utf-8") as f:
        raw_rows = json.load(f)

    rows = [ListingRow(**r) for r in raw_rows]
    write_rows_to_xlsx(rows, Path(sys.argv[2]))
    print(f"Wrote {len(rows)} row(s) to {sys.argv[2]}")


if __name__ == "__main__":
    main()
