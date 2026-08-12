"""
extract_spreadsheet.py

Extraction path for uploaded provider .xlsx/.csv spreadsheets - unlike
extract.py/extract_email.py (which turn unstructured PDF/email content into
ListingRows via Gemini vision), a provider spreadsheet is already one row
per property in tabular form. The only real unknown is which of its own
column headers corresponds to which ListingRow field, so this is a header-
mapping problem, not an extraction problem - no Gemini call is involved.

Flow: read_spreadsheet() gets the raw DataFrame with the provider's own
headers untouched -> suggest_mapping() guesses a {header: field_name_or_
None} mapping automatically (exact synonym match, then a conservative
fuzzy fallback - no per-column human confirmation step) -> if every
CRITICAL_FIELDS entry mapped successfully, build_rows() applies the
mapping straight away. If a critical field didn't map (e.g. a provider
whose export uses the area name itself as the building-column header, or a
sheet with no single consistent header row at all - see
unmapped_critical_fields), the caller (app.py) falls back to Gemini text
extraction (extract_spreadsheet_gemini.extract_sheet) for that sheet,
automatically and without asking a human - a per-column confirmation
prompt used to exist here but was removed: a sheet with no real header row
has no legitimate column for a person to pick in the first place, only
nonsense pseudo-header options. get_saved_critical_field_rescue/
apply_critical_field_rescue (storage.file_store, keyed by header_hash)
still apply an ALREADY-saved rescue answer from before that prompt was
removed, so an old answer keeps working, but save_critical_field_rescue
itself (which created a new one) is unused now - kept in place rather than
deleted, dormant, in case a future caller wants to set one
programmatically. Rows then flow through the same geocode_rows()/
save_staging_file() steps as every other source type.
"""

import difflib
import hashlib
import math
import re
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from master_merge import field_kind, normalize_key
from schema import ListingRow
from staging_writer import title_case_label
from storage.file_store import dataframe_to_listing_rows

# Fields no header should ever be mapped to - property_id is assigned only
# once a row lands in the master (master_merge.py), never from a source
# file, and source_file is set programmatically from the upload itself
# (see build_rows), not read from any column.
UNMAPPABLE_FIELDS = ("property_id", "source_file")

# Fields important enough that leaving them unmapped shouldn't happen
# silently - see unmapped_critical_fields. building is schema-required
# (ListingRow.building: str, not Optional) - a spreadsheet whose building
# column goes unmapped doesn't just have one blank field per row, it
# produces ZERO rows outright (dataframe_to_listing_rows skips every row
# with no building at all) - confirmed against two real UNION "by-area"
# export files, whose building-name column is headered with the area's
# own name ("Clerkenwell & Farringdon", "Fitzrovia & Marylebone" - text
# that's different per file and shares no vocabulary with "building" in
# any file-independent way a synonym or fuzzy match could ever catch).
#
# address_1 was considered too (no fallback fills it in either, unlike
# provider's filename-based guess_provider_name) but deliberately left
# out: confirmed against the real Kitt's Availability file - a format
# this rescue mechanism must NOT interrupt, since every one of its
# genuinely critical fields already maps on its own - that Kitt's has NO
# address column at all (only Building, used as the property identifier
# the same way several other real formats also do). Including address_1
# here would prompt for a rescue on Kitt's every single time, for a field
# that format simply doesn't have and was never expected to need. building
# earns the special treatment because losing it is catastrophic (zero
# rows, not just one blank field); address_1 going unmapped just leaves
# that one field blank per row, same as any other non-critical column.
CRITICAL_FIELDS = ("building",)

# Extra known header variants (already in normalize_key's output form - see
# _build_field_synonyms) beyond a field's own name/title-case label, drawn
# from real provider spreadsheets actually seen (not speculative guesses) -
# e.g. UNION's current (full) export format uses "Floor/Unit", "Size (sq
# ft)", "Marketing Price (Based on Min Term) PCM", etc.; UNION's OTHER real
# "by-area" export format (a separate, genuinely different header set - no
# Area/Building/Address columns at all, see CRITICAL_FIELDS above) uses its
# own distinct wording instead - "Current spec", "Monthly Rate", "Price
# p/sq.ft", "Brochure" - added below once confirmed against that real file.
# These 4 specifically were tried against the fuzzy fallback first and
# rejected deliberately, not by oversight - none cleared the 0.84
# bidirectional-coverage gate (best real score was "Brochure" at 0.762
# against "brochure link", still failing because "link" has literally no
# counterpart in a single-word header - a structural gap, not a close
# miss), and the two next-closest ("Price p/sq.ft" at 0.631, "Current
# spec" at 0.528) sit well inside the range of already-confirmed false
# positives (e.g. "For Sale"/"To Let" false-matched at 0.667 during the
# fuzzy-threshold investigation) - loosening the threshold to reach these
# would reopen exactly those false positives, not just add safe coverage.
# An unrecognized header simply gets no guess (mapped to None, dropped)
# rather than a forced fuzzy pick.
#
# "size" and "price per sq ft" (bare, no punctuation) were added once
# confirmed against a real UNION by-area file's OTHER own sheet naming
# convention ("Aldgate & Whitechapel") - the same by-area format's City-
# style sheets spell these "Size sq.ft"/"Price p/sq.ft" (already covered:
# "Size sq.ft" normalizes to "size sqft", matching size_sqft's own title-
# case-label synonym; "Price p/sq.ft" normalizes to "price psqft", already
# present below) - confirming UNION genuinely varies this wording sheet-to-
# sheet within the SAME workbook, not just file-to-file.
EXTRA_SYNONYMS = {
    "floor_unit": ("floorunit", "floor"),
    "size_sqft": ("size sq ft", "sq ft", "size"),
    "desks_max": ("desks",),
    "rent_pcm": ("marketing price based on min term pcm", "monthly rate"),
    "rent_psf": ("marketing price based on min term psf", "price psqft", "price per sq ft"),
    "brochure_link": (
        "brochure pdf", "link to file", "link to brochure", "brochure link", "link to brochure pdf", "brochure",
    ),
    "floorplan_link": (
        "floorplan", "floor plan", "floorplans", "floor plans", "link to floorplan", "download floorplan",
        "floorplan link", "floor plan link",
    ),
    "state_of_space": ("current spec",),
    "address_1": ("property address 1", "address"),
    "postcode": ("property postcode",),
    "lat": ("latitude",),
    "lng": ("longitude",),
    "submarket": ("area",),
    "internal_ref": ("external ref",),
    "provider": ("assigned agents",),
    "special_features": ("key features",),
}


def _build_field_synonyms() -> dict:
    synonyms = {}
    for field_name in ListingRow.model_fields:
        if field_name in UNMAPPABLE_FIELDS:
            continue
        options = {normalize_key(field_name), normalize_key(title_case_label(field_name))}
        options.update(EXTRA_SYNONYMS.get(field_name, ()))
        synonyms[field_name] = options
    return synonyms


FIELD_SYNONYMS = _build_field_synonyms()


# The fields whose header text is virtually never anything BUT a genuine
# availability-table column label, together - "Floor", a Size column, a
# Monthly Rate column, a Price-per-sq.ft column, and "Brochure" are never
# something a spreadsheet's own DATA cells (or a title/prose line) would
# independently happen to say in the same row, so requiring a recognizable
# synonym (see FIELD_SYNONYMS) for every one of these five, together, in a
# single row is strong enough structural evidence to trust unconditionally
# for ANY provider - see _find_header_row. Deliberately five fields, not
# fewer: "building" is excluded on purpose - UNION's own by-area export
# convention headers that column with the area's own name (see
# _PROVIDER_BUILDING_FALLBACK_RUNS), which this detector must recognize as
# a genuine header row too, before that column's real meaning is even
# resolved.
_HEADER_ROW_EVIDENCE_FIELDS = ("floor_unit", "size_sqft", "rent_pcm", "rent_psf", "brochure_link")

# How many rows from the top to consider as a candidate header row at all -
# generous enough for a real UNION by-area sheet (confirmed against a real
# file: rows 1 and 3 blank, row 2 introductory prose, the real header at
# row 4) without risking mistaking an early DATA row on some other,
# unrelated sheet shape for a header this deep into the sheet.
_MAX_HEADER_SCAN_ROWS = 10


def _looks_like_header_row(cells: list) -> bool:
    """
    True when `cells` (one row's raw values, in column order) carries
    _HEADER_ROW_EVIDENCE_FIELDS' own strong structural evidence of being a
    genuine availability-table header row - a recognizable synonym for
    EVERY one of those five fields, each in its own cell (not necessarily
    adjacent, and never more than one field claiming the same cell - a
    header row's columns are already distinct by construction, so no
    used-fields bookkeeping is needed the way suggest_mapping's own two
    passes need it against reused headers).
    """
    normalized = {normalize_key(c) for c in cells if c is not None and str(c).strip()}
    return all(
        any(option in normalized for option in FIELD_SYNONYMS[field])
        for field in _HEADER_ROW_EVIDENCE_FIELDS
    )


def _find_header_row(ws) -> int:
    """
    The 1-based row number to treat as `ws`'s real header row - almost
    always row 1 (the overwhelming majority of real provider exports:
    returned immediately the first time it qualifies, so this changes
    nothing for any format already working today), but scans up to
    _MAX_HEADER_SCAN_ROWS for the first row that clears _looks_like_
    header_row's own strong-evidence bar when row 1 itself doesn't.

    Confirmed necessary against a real UNION by-area export ("City"): its
    row 1 is entirely blank and row 2 is introductory prose, with the real
    header at row 4.

    Falls back to row 1 (the previous, unconditional assumption) if no row
    within the scan window clears the bar at all - never guesses a row with
    only partial evidence, so any format whose header doesn't happen to
    carry this exact five-field combination (most of them - a sheet with no
    Brochure column, say) is completely unaffected by this function
    existing at all.
    """
    last_row = min(ws.max_row, _MAX_HEADER_SCAN_ROWS)
    for row_num in range(1, last_row + 1):
        cells = [cell.value for cell in ws[row_num]]
        if _looks_like_header_row(cells):
            return row_num
    return 1


# Matches the exact shape of a Google Sheets/IMPORTRANGE .xlsx export's
# formula cells - e.g. '=IFERROR(__xludf.DUMMYFUNCTION("..."),"Area")' or
# '=IFERROR(__xludf.DUMMYFUNCTION("""COMPUTED_VALUE"""),759.0)'. Google's
# export wraps every formula-derived cell this way; __xludf.DUMMYFUNCTION
# is a Google-only placeholder that can never actually execute in any real
# spreadsheet engine, so IFERROR always falls through to its second
# argument - which is exactly the real value a spreadsheet program would
# show for that cell. Greedy .* for the DUMMYFUNCTION argument naturally
# backtracks to the RIGHTMOST "), <fallback>)" split, correctly separating
# the (possibly comma/paren-containing) inner call from the real fallback
# even when the inner argument itself contains nested parens/commas (e.g.
# an IMPORTRANGE(...) call). re.DOTALL since the fallback itself can
# contain a real embedded newline (confirmed against a real export, e.g.
# a "Size \n(sq ft)" header).
_XLUDF_FORMULA_RE = re.compile(r"^=IFERROR\(__xludf\.DUMMYFUNCTION\(.*\),\s*(.+)\)$", re.DOTALL)


def _parse_xludf_fallback(formula):
    """
    Extracts the literal fallback value straight out of a Google Sheets
    export formula's own text - the last-resort path for a cell whose
    cached value (openpyxl's data_only=True) genuinely isn't present
    (can happen if the file was downloaded without ever being opened/
    recalculated by a real spreadsheet program first). Returns None for
    anything that isn't recognizably this exact shape, rather than
    guessing - a caller falls back to treating the cell as blank.
    """
    if not isinstance(formula, str) or not formula.startswith("=IFERROR(__xludf.DUMMYFUNCTION("):
        return None
    match = _XLUDF_FORMULA_RE.match(formula)
    if not match:
        return None
    fallback = match.group(1).strip()
    if len(fallback) >= 2 and fallback.startswith('"') and fallback.endswith('"'):
        return fallback[1:-1].replace('""', '"')
    try:
        return float(fallback) if "." in fallback else int(fallback)
    except ValueError:
        return fallback


def _resolve_cell_value(value_cell, formula_cell):
    """
    value_cell/formula_cell are the SAME cell read from two separate
    load_workbook() calls, one with data_only=True (the last-calculated
    result) and one with data_only=False (the raw formula) - openpyxl
    only exposes one or the other per loaded workbook, never both from a
    single load. Prefers the cached value; falls back to parsing it
    straight out of the formula text (see _parse_xludf_fallback) only when
    the cache is genuinely missing for what really is a formula cell -
    never applied to an actually-blank cell, which has no formula to
    parse in the first place and correctly stays None either way.
    """
    if value_cell.value is not None:
        return value_cell.value
    return _parse_xludf_fallback(formula_cell.value)


def list_sheet_names(data: bytes, suffix: str) -> list:
    """
    Every sheet name in an .xlsx workbook, in file order - [None] for a .csv
    (which has no sheet concept at all), so a caller can loop `for sheet_name
    in list_sheet_names(...)` uniformly across both suffixes without a
    separate branch, passing sheet_name straight through to read_spreadsheet.
    """
    if suffix == ".csv":
        return [None]
    return list(load_workbook(BytesIO(data), read_only=True).sheetnames)


def read_spreadsheet(data: bytes, suffix: str, sheet_name: str = None) -> pd.DataFrame:
    """
    Reads an uploaded provider spreadsheet's raw headers/rows exactly as the
    provider wrote them - unlike storage.file_store.read_xlsx_with_hyperlinks
    (which assumes OUR OWN header format, written by staging_writer.
    write_rows_to_xlsx), headers here are whatever text is actually in the
    sheet's own header row; which one means what is decided later by
    suggest_mapping/the confirm-mapping UI, never assumed from position or a
    fixed label set.

    The header row itself is usually row 1, but not assumed to be - see
    _find_header_row, which scans a few rows down for one with strong
    structural header evidence when row 1 doesn't have it. Confirmed
    necessary against a real UNION by-area export ("City"): row 1 is blank,
    row 2 is introductory prose, and the real header is on row 4.

    sheet_name selects a specific sheet in a multi-sheet .xlsx (see
    list_sheet_names) - defaults to the workbook's active sheet, preserving
    prior single-sheet-file behavior for every existing caller. Ignored for
    .csv, which has only one sheet by construction (and whose header-row
    detection is unchanged - pandas' own row 1 assumption - since a CSV
    export has never been seen with this problem; only a real multi-sheet
    .xlsx with actual title/prose rows above its header has).

    Every .xlsx cell carrying a real hyperlink target uses that target
    instead of its displayed text (e.g. a "Brochure PDF" column showing just
    "Open" or "View") - which column ends up mapped to brochure_link isn't
    known yet at read time, so every column gets this treatment, not just
    ones that already look like a link column.

    Loads the workbook twice - once per openpyxl data_only mode - so every
    cell's cached value is available AND its raw formula text is available
    as a fallback (see _resolve_cell_value): confirmed necessary against a
    real Google Sheets/IMPORTRANGE export (Kitt's Availability), whose
    every header/data cell is a formula wrapping __xludf.DUMMYFUNCTION -
    reading only the formula (the old behavior) surfaced that raw formula
    text as if it were the real header/value.
    """
    if suffix == ".csv":
        return pd.read_csv(BytesIO(data))

    wb_values = load_workbook(BytesIO(data), data_only=True)
    ws_values = wb_values[sheet_name] if sheet_name else wb_values.active
    wb_formulas = load_workbook(BytesIO(data), data_only=False)
    ws_formulas = wb_formulas[sheet_name] if sheet_name else wb_formulas.active

    header_row = _find_header_row(ws_values)
    headers = [
        _resolve_cell_value(vcell, fcell)
        for vcell, fcell in zip(ws_values[header_row], ws_formulas[header_row])
    ]
    records = []
    for value_row, formula_row in zip(
        ws_values.iter_rows(min_row=header_row + 1), ws_formulas.iter_rows(min_row=header_row + 1)
    ):
        record = {}
        for col_idx, (vcell, fcell) in enumerate(zip(value_row, formula_row)):
            header = headers[col_idx]
            if vcell.hyperlink is not None:
                record[header] = vcell.hyperlink.target
            else:
                record[header] = _resolve_cell_value(vcell, fcell)
        records.append(record)
    return pd.DataFrame(records, columns=headers)


# Words that carry no field-identifying meaning on their own - dropped
# before fuzzy word-matching so e.g. "Link to Brochure" and "Brochure"
# compare on {"link", "brochure"} vs {"brochure"} rather than being thrown
# off by "to". Deliberately NOT stripped from the exact-match path above -
# that path already gets normalize_key's full string, and a real header
# genuinely containing one of these words is either an exact synonym
# already (handled) or should fall through to fuzzy matching regardless.
_FUZZY_STOPWORDS = frozenset({"to", "of", "the", "a", "an", "for", "and"})

# A header word and a candidate word must be at least this similar (per
# difflib.SequenceMatcher.ratio(), 0-1) to count as "the same word" for
# fuzzy matching - confirmed against every real UNION/Kitt's header known
# to have no genuine field ("For Sale", "Min. Term", "Floor Plan", "Patch?",
# "Who Onboarded?", etc. - see tests.test_extract_spreadsheet), none of
# which cleared even 0.667 bidirectional similarity against any field, so
# 0.84 leaves comfortable headroom while still catching genuine near-misses
# ("Special Feature" vs special_features -> 0.967, "Rent (PCM)" vs
# rent_pcm -> 1.0). Short words are blocked from this ratio check entirely
# (see _word_similarity) - "To"/"Let" vs "Lat" would otherwise score 0.667
# from small-string noise alone, nowhere near a real match.
_FUZZY_WORD_MATCH_THRESHOLD = 0.84


def _fuzzy_words(key: str) -> list:
    return [w for w in key.split() if w not in _FUZZY_STOPWORDS]


def _word_similarity(a: str, b: str) -> float:
    # A short word (e.g. "to", "let", "lat") almost always scores
    # deceptively high against another short, unrelated word purely from
    # small-string noise (difflib.ratio() on "to"/"lat" is 0.67) - real
    # false positive found during the fuzzy-match investigation ("To Let"
    # scored 0.67 against "lat" this way). Requiring an exact match below
    # 3 characters closes that off without needing a separate length-ratio
    # guard.
    if len(a) < 3 or len(b) < 3:
        return 1.0 if a == b else 0.0
    return difflib.SequenceMatcher(None, a, b).ratio()


def _fuzzy_field_score(header_words: list, candidate_words: list) -> float:
    """
    0 unless EVERY header word has a close match among candidate_words AND
    EVERY candidate word has a close match among header_words - both
    directions. One-directional coverage is what produced the worst false
    positives in the fuzzy-match investigation: "Min. Term" (-> {"min",
    "term"}) sits entirely inside rent_pcm's own synonym "marketing price
    based on min term pcm" (-> {"marketing","price","based","min","term",
    "pcm"}), so a header-words-covered-by-candidate check alone scored it
    a perfect match to the wrong field. Requiring the reverse direction too
    - every candidate word must also be covered - correctly rejects that,
    since "marketing"/"price"/"based"/"pcm" have nothing to match against
    in "Min. Term". Returns the (mean forward, mean backward) average
    similarity purely so multiple full-coverage candidates (rare, given how
    strict this already is) can be ranked against each other, not as a
    softened threshold - full bidirectional coverage is still required for
    any nonzero result.
    """
    if not header_words or not candidate_words:
        return 0.0
    forward = [max(_word_similarity(hw, cw) for cw in candidate_words) for hw in header_words]
    backward = [max(_word_similarity(hw, cw) for hw in header_words) for cw in candidate_words]
    if min(forward) < _FUZZY_WORD_MATCH_THRESHOLD or min(backward) < _FUZZY_WORD_MATCH_THRESHOLD:
        return 0.0
    return (sum(forward) / len(forward) + sum(backward) / len(backward)) / 2


def _fuzzy_match(header_words: list, used_fields: set) -> str:
    best_field, best_score = None, 0.0
    for field_name, options in FIELD_SYNONYMS.items():
        if field_name in used_fields:
            continue
        for option in options:
            score = _fuzzy_field_score(header_words, _fuzzy_words(option))
            if score > best_score:
                best_field, best_score = field_name, score
    return best_field


def header_hash(headers: list) -> str:
    """
    Identifies one exact header set, in order - a provider's recurring
    export (e.g. a monthly UNION availability sheet) keeps the same headers
    in the same order every time, so this is enough to recognize "we've
    already resolved this format's critical fields" on a later re-upload
    (see storage.file_store.get_saved_critical_field_rescue), without
    needing anything fuzzier. ␟ (a control character never legitimately
    present in a spreadsheet header) separates fields so a header itself
    containing "|" or "," can't collide with the join.
    """
    joined = "␟".join(str(h) for h in headers)
    return hashlib.sha256(joined.encode("utf-8")).hexdigest()


def unmapped_critical_fields(mapping: dict) -> list:
    """
    Which of CRITICAL_FIELDS has no header mapped to it, in CRITICAL_FIELDS
    order - the caller (app.py) uses this to decide whether a spreadsheet
    upload can proceed fully automatically, or needs a targeted rescue
    prompt for specifically these fields (never a reason to re-confirm
    every column - every other field is trusted to suggest_mapping either
    way).
    """
    mapped_fields = {field for field in mapping.values() if field}
    return [field for field in CRITICAL_FIELDS if field not in mapped_fields]


def apply_critical_field_rescue(mapping: dict, rescue: dict) -> dict:
    """
    Overlays a human-confirmed {field_name: header_or_None} rescue (see
    storage.file_store.get_saved_critical_field_rescue) on top of an
    automatic mapping - a non-None header is assigned to that field even if
    suggest_mapping had guessed something else for it (a human's explicit
    choice always wins over an automatic guess); None means "confirmed - this
    format genuinely has no such column," left unmapped exactly as
    suggest_mapping already left it, just without asking again next time.
    Returns a new dict; never mutates the mapping passed in.
    """
    mapping = dict(mapping)
    for field, header in rescue.items():
        if header is not None:
            mapping[header] = field
    return mapping


def unresolved_critical_fields(mapping: dict, rescue: dict) -> list:
    """
    Critical fields that still need a human decision - unmapped_critical_
    fields() on the ALREADY-RESCUED mapping (see apply_critical_field_
    rescue), minus whatever `rescue` already has an entry for. A rescue
    entry of None means "confirmed - genuinely no such column," not "still
    needs asking" - without subtracting those out, a confirmed-blank field
    (e.g. a format that truly has no address_1 column) would prompt again
    on every single future upload of that same format, forever, rather than
    staying silently resolved the way a real assignment already does.
    """
    missing = unmapped_critical_fields(mapping)
    return [field for field in missing if field not in rescue]


# UNION's own "by-area" export convention (see CRITICAL_FIELDS' comment
# above) headers what's structurally always the Building column with the
# area's own name instead of a recognizable field name - confirmed against
# every real by-area file seen so far ("Clerkenwell & Farringdon",
# "Fitzrovia & Marylebone", "Soho / Covent Garden", "Mayfair / St James",
# "Midtown", "London Bridge / Southbank", "City", "Aldgate & Whitechapel"):
# always the column immediately BEFORE this exact fixed run of headers -
# Floor, "Current spec" or "Category", a Size column, a lease-term column,
# Monthly Rate, a Price-per-sq.ft column, Brochure. Matched as a contiguous
# run rather than fixed absolute column indices, since a real file
# sometimes has a genuinely blank column before the area-name one and
# sometimes doesn't - "immediately before this run" holds either way,
# "always column N" doesn't. Each position tolerates the wording variants
# actually confirmed across different sheets of the SAME real workbook -
# e.g. "Size sq.ft" on one sheet, bare "Size" on another; "Minimum Term" vs
# bare "Term"; "Price p/sq.ft" vs "Price per sq ft" - never a single fixed
# string per position. Deliberately keyed on the WHOLE surrounding run, not
# "this one column is unmapped" alone, and only ever consulted for a
# filename already recognized as a specific known provider (see
# _leading_known_provider) - a coincidental run match on some other
# provider's genuinely different layout must never misfire into treating an
# unrelated column as Building. UNION's OTHER real export format (the
# current full Shoreditch-style one) already has a real "Building" header
# and maps automatically without ever reaching this.
_PROVIDER_BUILDING_FALLBACK_RUNS = {
    "UNION": (
        ("floor",),
        ("current spec", "category"),
        ("size sqft", "size"),
        ("minimum term", "term"),
        ("monthly rate",),
        ("price psqft", "price per sq ft"),
        ("brochure",),
    ),
}


def _structural_building_header(provider: str, headers: list) -> str:
    """
    Returns the header text sitting immediately before a known provider's
    own fixed run of headers (see _PROVIDER_BUILDING_FALLBACK_RUNS) if that
    exact run appears contiguously in `headers` - None for an unrecognized
    provider, no match at all, or more than one match (genuinely ambiguous
    - stay conservative rather than guess which one).
    """
    run = _PROVIDER_BUILDING_FALLBACK_RUNS.get(provider)
    if not run:
        return None

    normalized = [normalize_key(h) for h in headers]
    matches = [
        i for i in range(1, len(normalized) - len(run) + 1)
        if all(normalized[i + offset] in allowed for offset, allowed in enumerate(run))
    ]
    if len(matches) != 1:
        return None

    building_header = headers[matches[0] - 1]
    if building_header is None or not str(building_header).strip():
        return None
    return building_header


def structural_submarket_fallback(headers: list, filename: str) -> str:
    """
    The same area-name header text used to resolve UNION's by-area
    Building column (see apply_provider_structural_fallback/
    _structural_building_header) IS ALSO the submarket for every row in
    that file - constant per file, by construction (a separate export per
    area, e.g. "Clerkenwell & Farringdon.xlsx"), with no per-row Area/
    submarket column to read one from at all. Returns None under the exact
    same conditions _structural_building_header would (unrecognized
    provider, no fingerprint match) - this only ever makes sense for a file
    that actually needed that fallback in the first place.

    100% reliable when it applies at all - no API call, no dependency on
    Google's own neighbourhood-polygon coverage (confirmed patchy: reliable
    for Mayfair/Fitzrovia/Soho, but Google has no sublocality data at all
    for the real Clerkenwell & Farringdon addresses this was built for -
    see geocode.py's own submarket-backfill comment). The caller should
    apply this BEFORE that geocoded fallback, not after - this is a known-
    correct constant when it's available, that one's only ever a secondary,
    best-effort attempt for whatever this doesn't cover.
    """
    provider = _leading_known_provider(re.findall(r"[A-Za-z0-9']+", Path(filename).stem))
    return _structural_building_header(provider, headers)


def apply_provider_structural_fallback(mapping: dict, headers: list, filename: str) -> dict:
    """
    Automatic, provider-specific fallback for "building" when a provider's
    OWN recognizable export format structurally hides it from any header-
    TEXT-based mapping at all (synonym or fuzzy - see
    _structural_building_header) - e.g. UNION's "by-area" exports, which
    header that column with the area's own name, different per file, with
    no vocabulary in common with "building" any text-based match could ever
    catch. Runs BEFORE the human rescue overlay (apply_critical_field_
    rescue), not instead of it - a saved/human rescue for a genuinely novel
    future layout from the same provider still always wins. Only ever fills
    "building", and only when suggest_mapping's own normal pass left it
    unmapped. Returns a new dict; never mutates the mapping passed in.
    """
    mapping = dict(mapping)
    if "building" in mapping.values():
        return mapping

    provider = _leading_known_provider(re.findall(r"[A-Za-z0-9']+", Path(filename).stem))
    building_header = _structural_building_header(provider, headers)
    if building_header is not None:
        mapping[building_header] = "building"

    return mapping


def suggest_mapping(headers: list) -> dict:
    """
    Best-guess {header: field_name_or_None} mapping - applied automatically
    to every spreadsheet upload with no human confirmation step (see
    build_rows/app.py); a column that ends up genuinely unmapped is simply
    dropped, so both passes below are deliberately conservative rather than
    forcing a guess.

    Pass 1 - exact match via FIELD_SYNONYMS (field name/title-case label/
    known real-header synonyms, see EXTRA_SYNONYMS). Pass 2 - for whatever
    a header didn't exact-match, a conservative word-level fuzzy fallback
    (see _fuzzy_field_score) catches near-miss variants of a field's own
    name or a known synonym (pluralization, punctuation/parenthetical
    noise, minor rewording) without needing every possible phrasing hand-
    added to EXTRA_SYNONYMS - e.g. "Rent (PCM)" or "Special Feature"
    (singular) - while staying strict enough to leave genuinely unrelated
    headers (a real spreadsheet's "Min. Term", "Floor Plan", "Patch?", etc.)
    unmapped rather than guessing wrong; see the fuzzy-match investigation
    for the false positives an earlier, looser design produced against
    these exact columns.

    Never guesses two headers onto the same field, in EITHER pass - first
    match (in header order) wins, every later header that would also match
    that field is left unmapped instead, since a real spreadsheet shouldn't
    have two columns both claiming to be e.g. size_sqft, and guessing one
    of them wrong silently would be worse than dropping it.
    """
    mapping = {}
    used_fields = set()
    for header in headers:
        key = normalize_key(header)
        matched_field = None
        for field_name, options in FIELD_SYNONYMS.items():
            if field_name in used_fields:
                continue
            if key in options:
                matched_field = field_name
                break
        mapping[header] = matched_field
        if matched_field:
            used_fields.add(matched_field)

    for header in headers:
        if mapping[header] is not None:
            continue
        header_words = _fuzzy_words(normalize_key(header))
        matched_field = _fuzzy_match(header_words, used_fields)
        mapping[header] = matched_field
        if matched_field:
            used_fields.add(matched_field)

    return mapping


# Words that recur across many different providers' own export filenames
# and describe the FILE, not who's presenting it - stripped when guessing a
# default provider name (see guess_provider_name). Deliberately a short,
# evidence-grounded list (every entry drawn from a real filename seen in
# this project) rather than a speculative attempt at every possible
# boilerplate word - a guess that keeps a stray word is a minor annoyance
# a human corrects in one edit; this is only ever a pre-filled starting
# point for the one-time provider-name prompt, never used unconfirmed.
_PROVIDER_GUESS_STOPWORDS = frozenset({
    "availability", "external", "export", "extract", "download", "report",
    "update", "updated", "current", "live", "final", "draft", "copy",
    "schedule", "listing", "listings", "spreadsheet", "sheet", "data",
})

# Providers this pipeline already recognizes by name - once a filename's
# leading word(s) match one of these, guess_provider_name trusts that as
# the whole answer and ignores everything after it (area, date,
# "Availability", etc.) rather than trying to strip each irrelevant word
# individually. Replaces an earlier area-name stopword list that had to be
# manually re-extended every time UNION exported a new London area's file
# ("Fitzrovia & Marylebone", then "Clerkenwell & Farringdon", then
# "Shoreditch", then "London Bridge & Southwark", ...) - an open-ended,
# recurring cost for a provider whose filenames are otherwise completely
# predictable. A provider name, unlike a London area name, is a small,
# slow-growing set - once it's here, no future filename from that provider
# needs fixing again, regardless of what area/date/wording follows it.
_KNOWN_PROVIDER_NAMES = ("UNION", "Kitt's")


def _leading_known_provider(words: list) -> str:
    """
    Returns the canonically-cased provider name if `words` (see
    guess_provider_name's own word-splitting) starts with the same word
    sequence as one of _KNOWN_PROVIDER_NAMES, case-insensitively - None
    otherwise. Whole-word sequence matching, not a prefix/substring check,
    so a hypothetical provider like "Unionville" is never mistaken for
    "UNION".
    """
    for provider in _KNOWN_PROVIDER_NAMES:
        provider_words = re.findall(r"[A-Za-z0-9']+", provider)
        if [w.lower() for w in words[:len(provider_words)]] == [w.lower() for w in provider_words]:
            return provider
    return None

# Numeric date-like substrings ("2026-07-17", "2026_07_17", "17.07.2026")
# commonly embedded in a recurring export's filename - describe when the
# file was produced, not who produced it. No \b anchors: a real filename
# routinely butts the date straight up against an underscore separator
# ("Garden_2026-07-17") - "_" is itself a \w character, so \b never fires
# at that boundary and the date would go unstripped. The digit-group/
# separator shape is already specific enough not to need one.
_FILENAME_DATE_RE = re.compile(r"\d{4}[-_./]\d{1,2}[-_./]\d{1,2}|\d{1,2}[-_./]\d{1,2}[-_./]\d{2,4}")

# A month name paired with a short number, in either order ("June 26",
# "Jun 2026", "26 June") - confirmed necessary against the real UNION
# filename convention ("UNION - Availability - June 26 - Fitzrovia &
# Marylebone.xlsx"), whose "June 26" previously survived _FILENAME_DATE_RE
# untouched (that regex only matches an all-numeric date with punctuation
# separators between the groups - "June 26" has neither: no punctuation,
# and "June" isn't a digit group) and leaked straight into the guess as two
# unrelated-looking words. Matched and stripped as one whole phrase, before
# word-splitting, same as _FILENAME_DATE_RE - splitting first would lose
# the ability to recognize "June" + "26" as a single date, since "26" alone
# is far too generic a number to safely strip as a stopword on its own.
_MONTH_NAME_RE = (
    r"(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?"
    r"|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
)
_MONTH_YEAR_RE = re.compile(
    rf"\b{_MONTH_NAME_RE}\.?\s+\d{{1,4}}\b|\b\d{{1,4}}\s+{_MONTH_NAME_RE}\.?\b", re.IGNORECASE
)


def guess_provider_name(filename: str) -> str:
    """
    Best-guess provider name derived from an uploaded spreadsheet's own
    filename - applied automatically to every spreadsheet upload with no
    column mapping to provider (see app.py's fill_missing_provider), never
    confirmed by a human.

    First checks whether the filename starts with an already-known provider
    name (see _KNOWN_PROVIDER_NAMES/_leading_known_provider) - if so, that
    name alone is the answer, and everything else in the filename (area,
    date, "Availability", etc.) is ignored outright rather than stripped
    word-by-word. This is what lets a provider like UNION - whose by-area
    export filenames otherwise change every time - never need fixing again
    for a new area.

    Otherwise falls back to word-by-word stripping: extension, parenthetical/
    bracketed asides (almost always describe the file - "(External)",
    "[DRAFT]" - not who's presenting it), embedded dates (both plain numeric
    - see _FILENAME_DATE_RE - and a month name paired with a number - see
    _MONTH_YEAR_RE), and a small set of generic boilerplate words that recur
    across many providers' own export filenames (see
    _PROVIDER_GUESS_STOPWORDS) - e.g. "Availability Export.xlsx" -> both
    words are stopwords, nothing survives, falls back to the raw stem
    itself (see below) rather than guessing blank.

    Examples: "Kitt's Availability (External).xlsx" -> "Kitt's" (known-
    provider match), "UNION - Availability - June 26 - Fitzrovia &
    Marylebone.xlsx" -> "UNION" (known-provider match, regardless of area),
    "Breezblok.csv" -> "Breezblok" (no known provider, nothing to strip).
    """
    stem = Path(filename).stem

    known = _leading_known_provider(re.findall(r"[A-Za-z0-9']+", stem))
    if known:
        return known

    stripped = re.sub(r"[\(\[][^)\]]*[\)\]]", " ", stem)
    stripped = _FILENAME_DATE_RE.sub(" ", stripped)
    stripped = _MONTH_YEAR_RE.sub(" ", stripped)
    words = re.findall(r"[A-Za-z0-9']+", stripped)
    kept = [w for w in words if w.lower() not in _PROVIDER_GUESS_STOPWORDS]
    guess = " ".join(kept).strip()
    return guess or stem


def _coerce_numeric(value, kind: str):
    """
    None for anything that isn't a real number - confirmed necessary
    against an actual current-format UNION file, whose Lat/Lng columns
    hold the literal text "Needs manual lookup" for rows the provider
    hasn't geocoded yet themselves, sitting in what's otherwise a numeric
    column. Passing that straight into ListingRow's lat: Optional[float]
    would raise a validation error and fail the WHOLE file's extraction
    over one placeholder cell in one column - dropping it to None instead
    (same as a genuinely blank cell) leaves the row's other, valid fields
    intact and lets geocode_rows attempt a real lookup for that row, same
    as it always has for PDF/email sources with no coordinates at all.

    kind is the target ListingRow field's own declared type ("int" or
    "float" - see master_merge.field_kind). For an "int" field, a
    genuinely fractional value is rounded to the nearest whole number
    rather than passed through as-is - confirmed necessary against a real
    Kitt's Availability row, whose desks_max resolved to 38.257143 (almost
    certainly a computed/averaged cell caught up in the mapped column
    range). Pydantic's int validation rejects a fractional float outright,
    which - like the lat/lng case above - would otherwise abort the WHOLE
    file's extraction over one cell; a rounded desk count is a far better
    outcome than that.
    """
    if value is None:
        numeric = None
    elif isinstance(value, (int, float)):
        numeric = value
    else:
        try:
            numeric = float(str(value).replace(",", "").strip())
        except ValueError:
            numeric = None

    # A pandas NaN is itself a float (math.isnan(float("nan")) is True, and
    # isinstance(nan, float) is True) - it slides straight through the
    # isinstance branch above as if it were a real number, rather than
    # hitting the ValueError branch a non-numeric string would. round(nan)
    # raises ValueError - confirmed necessary against a real Kitt's
    # Availability file, whose desks_max column has genuinely blank cells
    # that pandas reads as NaN rather than None. Treated the same as a
    # genuinely blank cell (None), for the same reason as every other
    # branch in this function: one bad cell must not abort the WHOLE
    # file's extraction.
    if isinstance(numeric, float) and math.isnan(numeric):
        numeric = None

    if numeric is not None and kind == "int":
        return round(numeric)
    return numeric


def build_rows(df: pd.DataFrame, mapping: dict, source_file: str) -> list:
    """
    Applies a confirmed {header: field_name_or_None} mapping and converts
    the result to real ListingRows - reuses storage.file_store.
    dataframe_to_listing_rows for the actual row construction (blank-row
    skipping, NaN cleanup, ListingRow(**cleaned) with extra="ignore" for any
    column left unmapped) rather than duplicating that logic here, exactly
    as the earlier design investigation recommended. Every column mapped to
    an int/float ListingRow field is coerced first (see _coerce_numeric) -
    a raw provider spreadsheet's numeric-looking columns aren't guaranteed
    to actually be clean numbers.

    No longer fills in a default provider itself - app.py's
    fill_missing_provider does that uniformly for every upload type (PDF,
    email, spreadsheet, reused rows alike), not just spreadsheets, so
    duplicating the same fallback here would be redundant.
    """
    renamed = {header: field_name for header, field_name in mapping.items() if field_name}
    mapped_df = df[[h for h in df.columns if h in renamed]].rename(columns=renamed)
    for field_name in mapped_df.columns:
        kind = field_kind(field_name)
        if kind in ("int", "float"):
            mapped_df[field_name] = mapped_df[field_name].apply(lambda v: _coerce_numeric(v, kind))
    rows = dataframe_to_listing_rows(mapped_df)
    for row in rows:
        row.source_file = source_file
    return rows
